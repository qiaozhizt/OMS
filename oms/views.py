# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render
import simplejson as json
import urllib2
import re
# Create your views here.
from django.http import HttpResponse, JsonResponse
from django.urls import reverse

from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

from django.utils import timezone

from django.db import connections
from django.db import transaction
from collections import namedtuple

from models.generatelog_models import GenerateLog
from django.core import serializers
from django.contrib.auth.models import User
from django.contrib.auth.models import Group, Permission
from django.contrib.auth.decorators import permission_required
from models.order_tracking_report_cs_models import OrderTrackingReportCS
from models.ordertracking_models import OrderTracking
from models.order_tracking_report_models import OrderTrackingReport
from models.send_comments_models import *
from models.product_models import LabProduct
# from models.order_models import PgOrder, PgOrderItem, OrderActivity
# from models.order_models import LabOrder
# from orderform import PgOrderFormDetail
# from django.forms.formsets import formset_factory
# from pytz import utc
# from pytz import timezone

from models.order_models import CustomerAccountLog
import logging
import time
import math
from util.response import response_message as res_msg, json_response
from django.contrib.contenttypes.models import ContentType
from models.response_models import shipment
from models.application_models import *
from models.actions_models import Action
from django.db.models import Q
from const import *
from decimal import *
import datetime
from models.order_models import *
from django.forms.models import modelformset_factory
from django.forms import widgets as Fwidgets
from models.response_models import *
from models.shipment_models import *
from . import const
from pg_oms.settings import *
from wms.models import inventory_delivery,product_frame
from vendor.models import lens_order, LensSpecmap

from .controllers.order_controller import *
from django.http import HttpResponseRedirect

from .controllers.lab_order_controller import lab_order_controller
from api.controllers.tracking_controllers import tracking_lab_order_controller
from .controllers.pg_order_controller import pg_order_controller
import requests
from .controllers.wx_purchase_controller import wx_purchase_controller
from .controllers.wx_meta_purchase_controller import wx_meta_purchase_controller
from .controllers.wc_purchase_controller import wc_purchase_controller
from wms.web_inventory import *
from wms.models import inventory_struct_warehouse_controller, product_frame, locker_controller,inventory_struct_warehouse,inventory_struct
from .controllers import order_controller
from django.forms.models import model_to_dict
from qc.models import glasses_final_inspection_technique
import MySQLdb
import traceback
from api.controllers.pgorder_frame_controllers import pgorder_frame_controller
from wms.models import LockersItem
from util.usps import get_ship2_tracking_number, read_ups_by_track_numbers
from oms.models import RemakeOrder
from vendor.contollers import WcOrderStatusController, WxOrderStatusController


from io import BytesIO,StringIO
import os
import numpy as np
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

@login_required
def dashboard(request):
    _form_data = {}
    try:
        logging.debug("dashboard ....")
        from report.models import dashboard_controller
        dc = dashboard_controller()
        rm = dc.get_not_inlad()
        _form_data["obj_not_inlab"] = rm.obj

        rm = dc.get_new_laborders()
        _form_data["obj_new"] = rm.obj
        rm = dc.get_processing_laborders()
        _form_data["obj_processing"] = rm.obj
        rm = dc.get_today_complete_laborders()
        _form_data["obj_today_complete"] = rm.obj
        rm = dc.get_today_shipped_laborders()
        _form_data["obj_today_shipped"] = rm.obj

    except Exception as e:
        logging.debug(str(e))

    return render(request, 'dashboard.html',
                  {
                      'form': _form_data
                  },
                  )


# redirect login.html
def loginHtml(request):
    return render(request, 'login.html')


# login
def loginOms(request):
    username = request.POST['username']
    password = request.POST['pwd']
    user = authenticate(username=username, password=password)
    if user is not None and user.is_active:
        # Correct password, and the user is marked "active"
        login(request, user)
        # Redirect to a success page.
        return HttpResponseRedirect("/oms/dashboard/")
    else:
        # Show an error page
        return HttpResponseRedirect("/oms/login/")


# loginout
def loginout(request):
    logout(request)
    return render(request, 'login.html')


# AdminLTE
@login_required
def index(request):
    db = dashboard(request)  # 调用 dashboard方法，跳转dashboard页面
    return db


def namedtuplefetchall(cursor):
    "Return all rows from a cursor as a namedtuple"
    desc = cursor.description
    nt_result = namedtuple('Result', [col[0] for col in desc])
    return [nt_result(*row) for row in cursor.fetchall()]


def dictfetchall(cursor):
    "Return all rows from a cursor as a dict"
    columns = [col[0] for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
        ]


def generatelogmax():
    max_current_entity = 0
    generatelog = GenerateLog.objects.all().order_by("-id")[:1]
    if generatelog:
        max_current_entity = generatelog[0].current_entity
    return max_current_entity


# add by ranhy check magento customer by  email
@login_required
@permission_required('oms.UPDATE_CUSTOMER_INFO_VIEW', login_url='/oms/forbid/')
def search_customer_email(request):
    if request.method == "POST":
        customer_email = request.POST.get('customer_email', '')
        token = getToken()
        http_headers = {
            'Authorization': 'Bearer %s' % token,
            'Content-Type': 'application/json'
        }
        # 修改邮箱
        result = ""
        if (customer_email):
            search_url = base_url + "V1/customers/search?searchCriteria[filterGroups][0][filters][0][field]=email" \
                                    "&searchCriteria[filterGroups][0][filters][0][value]=%s" \
                                    "&searchCriteria[filterGroups][0][filters][0][condition_type]=eq" % customer_email
            api_response = requests.get(url=search_url, headers=http_headers, timeout=60,verify=False)
            dict_customer_data = json.loads(api_response.text)

            if (dict_customer_data.get('total_count')):
                customer_data = dict_customer_data.get('items')[0]
                result = {
                    "code": 0,
                    "lastname": customer_data.get('lastname'),
                    "firstname": customer_data.get('firstname')
                }
            else:
                result = {
                    "code": -1,
                    "msg": "This email(\"%s\") not found!" % customer_email
                }
            return HttpResponse(json.dumps(result))


@login_required
@permission_required('oms.UPDATE_CUSTOMER_INFO_VIEW', login_url='/oms/forbid/')
def update_customer_email(request):
    if request.method == "POST":
        customer_email = request.POST.get('customer_email', '')
        customer_email_new = request.POST.get('customer_email_new', '')
        token = getToken()
        http_headers = {
            'Authorization': 'Bearer %s' % token,
            'Content-Type': 'application/json'
        }
        # 修改邮箱
        result = ""
        if (customer_email_new and customer_email):
            search_url = base_url + "V1/customers/search?searchCriteria[filterGroups][0][filters][0][field]=email" \
                                    "&searchCriteria[filterGroups][0][filters][0][value]=%s" \
                                    "&searchCriteria[filterGroups][0][filters][0][condition_type]=eq" % customer_email
            api_response = requests.get(url=search_url, headers=http_headers,timeout=60,verify=False)
            dict_customer_data = json.loads(api_response.text)
            logging.debug("dict_customer_data:" + api_response.text)

            if (dict_customer_data.get('total_count')):
                customer_data = dict_customer_data.get('items')[0]
                change_email_url = base_url + 'V1/customers/%s' % customer_data.get('id')
                change_email_data = {
                    "customer": {
                        "id": customer_data.get('id'),
                        "email": customer_email_new,
                        "lastname": customer_data.get('lastname'),
                        "firstname": customer_data.get('firstname'),
                        "storeId": customer_data.get('store_id'),
                        "websiteId": customer_data.get('website_id'),
                    }
                    # ,
                    # "password": customer_pass
                }
                try:
                    api_cus_result = requests.put(change_email_url, data=json.dumps(change_email_data),
                                                  headers=http_headers, timeout=60,verify=False)
                    change_email_result = json.loads(api_cus_result.text)
                    cus_log = CustomerAccountLog()
                    cus_log.is_pwd = 0
                    cus_log.user_entity = request.user
                    cus_log.customer_email = customer_email_new
                    cus_log.old_customer_email = customer_email
                    cus_log.comments = ''
                    cus_log.save()
                    result = {
                        "code": 0,
                        "msg": "This email has been changed!"
                    }

                except Exception as e:
                    result = {
                        "code": -2,
                        "msg": "This email change faild!%s" % str(e)
                    }
            else:
                result = {
                    "code": -1,
                    "msg": "This email(\"%s\") not found!" % customer_email
                }
        return HttpResponse(json.dumps(result))
    else:
        return render(request, 'customer_account.html')


@login_required
@permission_required('oms.UPDATE_CUSTOMER_INFO_VIEW', login_url='/oms/forbid/')
def update_customer_pwd(request):
    if request.method == "POST":
        customer_email = request.POST.get('customer_email', '')
        customer_pass = request.POST.get('customer_pass', '')
        token = getToken()
        http_headers = {
            'Authorization': 'Bearer %s' % token,
            'Content-Type': 'application/json'
        }
        # 修改用户密码
        change_cus_pwd_url = base_url + "V1/customer/changepassword"
        data = {
            'email': customer_email,
            'password': customer_pass
        }
        try:
            api_response = requests.post(change_cus_pwd_url, data=json.dumps(data), headers=http_headers, timeout=60,verify=False)
            result = json.loads(api_response.text)
            if (result.has_key('code') and result.get('code') == 0):
                log = CustomerAccountLog()
                log.is_pwd = 1
                log.user_entity = request.user
                log.customer_email = customer_email
                log.old_customer_email = ''
                log.comments = customer_pass
                log.save()
                result = {
                    "code": 0,
                    "msg": "This customer password has been changed!"
                }
            else:
                result = {
                    "code": -1,
                    "msg": "Change password faild!"
                }
        except Exception as e:
            result = {
                "code": -1,
                "msg": "Change password faild!%s" % str(e)
            }
        logging.debug(result)
        return HttpResponse(json.dumps(result))
    else:
        return render(request, 'customer_account.html')


# generate order list
@login_required
@permission_required('oms.NMOL_VIEW', login_url='/oms/forbid/')
def order_list_new(request):
    max_current_entity = generatelogmax()
    logging.debug("max_current_entity==>%s" % max_current_entity)
    page = request.GET.get('page', 1)
    currentPage = int(page)
    with connections['pg_mg_query'].cursor() as cursor:
        sql = '''
        select t0.entity_id       
            ,t0.increment_id
            ,(case when (t9.base_subtotal+t9.base_discount_amount>=50) then 1 else 0 end) as is_vip
            ,t0.status
            ,t0.created_at
            ,(case when t0.payment_method='braintree' then 'credit_card' else t0.payment_method end) as payment_method
            ,(select convert(t9.total_qty_ordered,decimal)) as quantity
            ,(select FORMAT(t0.grand_total,2)) as grand_total
            ,CONCAT(t8.firstname,' ',t8.lastname) as customer_name
            ,(case when t9.shipping_method='standard_standard' then 'standard' when t9.shipping_method='express_express' then 'express' when t9.shipping_method='canada_express_canada_express' then 'ca_express' else t9.shipping_method end ) as shipping_method
            #,t1.email as cusomer_email
        from sales_order_grid t0
        left join customer_grid_flat t1
        on t1.entity_id = t0.customer_id
        left join sales_order t9
        on t9.entity_id = t0.entity_id
        
        left join customer_entity t8
        on t0.customer_id=t8.entity_id
        where t0.entity_id > %s
        order by t0.entity_id DESC
                '''
        logging.info(sql)
        cursor.execute(sql, [max_current_entity])
        # logging.debug(sql, max_current_entity)
        logging.debug(sql, [max_current_entity])
        results = namedtuplefetchall(cursor)
        # logging.debug("results==>%s" % results)
        cursor.close()
        paginator = Paginator(results, 20)  # Show 20 contacts per page

        try:
            contacts = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            contacts = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            contacts = paginator.page(paginator.num_pages)
        return render(request, 'order_list.html',
                      {'list': contacts, 'currentPage': currentPage, 'paginator': paginator,
                       'pgorderitem': 'New Mg Order List', 'requestUrl': '/oms/order_list_new'})
        # return render(request, 'order_list.html', {'list': results})


# generate order list all
@login_required
@permission_required('oms.MOL_VIEW', login_url='/oms/forbid/')
def order_list(request):
    # max_current_entity = generatelogmax()
    logging.debug('--------------------------------order_list_______________')
    page = request.GET.get('page', 1)
    currentPage = int(page)
    with connections['pg_mg_query'].cursor() as cursor:
        sql = '''
            select t0.entity_id       
            ,t0.increment_id
            ,(case when (t9.base_subtotal+t9.base_discount_amount>=50) then 1 else 0 end) as is_vip
            ,t0.status
            ,t0.created_at
            ,(case when t0.payment_method='braintree' then 'credit_card' else t0.payment_method end) as payment_method
            ,(select convert(t9.total_qty_ordered,decimal)) as quantity
            ,(select FORMAT(t0.grand_total,2)) as grand_total
            ,CONCAT(t8.firstname,' ',t8.lastname) as customer_name
            ,(case when t9.shipping_method='standard_standard' then 'standard' when t9.shipping_method='express_express' then 'express' when t9.shipping_method='canada_express_canada_express' then 'ca_express' else t9.shipping_method end ) as shipping_method
            #,t1.email as cusomer_email
            from sales_order_grid t0
            left join customer_grid_flat t1
            on t1.entity_id = t0.customer_id
            left join sales_order t9
            on t9.entity_id = t0.entity_id
            left join customer_entity t8
            on t0.customer_id=t8.entity_id
        '''

        timedel = date_delta()

        sql_ext = ' where t0.created_at>%s order by t0.entity_id DESC '
        para = timedel

        sql_full = sql + sql_ext

        logging.info(sql)
        # 获取URL中除page外的其它参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        cursor.execute(sql_full, [para])
        results = namedtuplefetchall(cursor)
        totalCount = len(results)
        cursor.close()
        paginator = Paginator(results, 20)  # Show 10 contacts per page

        try:
            contacts = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            contacts = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            contacts = paginator.page(paginator.num_pages)
        return render(request, 'order_list.html',
                      {'list': contacts,
                       'currentPage': currentPage,
                       'paginator': paginator,
                       'pgorderitem': 'Mg Order List',
                       'requestUrl': '/oms/order_list',
                       'allpage': json.dumps(totalCount),
                       'nowpage': json.dumps(currentPage),
                       'query_string': query_string})
        # return render(request, 'order_list.html', {'list': results})


"""search mgorderlist by increment_id"""


def serarchMgOrderByNumber(request):
    increment_id = request.POST.get("number")
    logging.debug("increment_id==>" + str(increment_id))
    """sign暂时无用"""
    sign = request.POST.get("sign")
    logging.debug("sign==>" + str(sign))

    try:
        # oid = int(increment_id)
        with connections["pg_mg_query"].cursor() as cursor:
            if len(increment_id) == 3 or len(increment_id) == 4:
                order_number = "%" + increment_id
                logging.info(MgSqlLike)
                cursor.execute(MgSqlLike, [order_number])
            else:
                logging.info(MgSqlEqu)
                cursor.execute(MgSqlEqu, [increment_id])
            queryset = dictfetchall(cursor)
            logging.debug(queryset)
            return JsonResponse(queryset, safe=False)
    except:
        with connections["pg_mg_query"].cursor() as cursor:
            cursor.execute(MgSqlEquName, [increment_id])
            queryset = dictfetchall(cursor)
            logging.debug(queryset)
            return JsonResponse(queryset, safe=False)


# select entity_id
def entity_id_list():
    max_current_entity = generatelogmax()

    with connections['pg_mg_query'].cursor() as cursor:
        cursor.execute("select entity_id from sales_order_grid where entity_id > %s", [max_current_entity])
        results = namedtuplefetchall(cursor)
        return results


# 公共方法
def generate_prescripion_tuple():
    prescripion_data_dict = {}
    prescripion_data_dict['t_sph'] = [x * 0.25 for x in range(-64, 65)]
    prescripion_data_dict['t_cyl'] = [x * 0.25 for x in range(-24, 25)]
    prescripion_data_dict['t_axis'] = range(181)
    prescripion_data_dict['t_add'] = [x * 0.25 for x in range(0, 17)]

    # prescripion_data_dict['binocularPD'] = range(35,81)
    prescripion_data_dict['binocularPD'] = [x * 0.5 for x in range(70, 161)]
    prescripion_data_dict['monocularPD'] = [x * 0.5 for x in range(35, 81)]
    prescripion_data_dict['t_prism'] = [x * 0.25 for x in range(25)]
    prescripion_data_dict['t_base'] = ["", "IN", "OUT", "UP", "DOWN"]  # 支持复合棱镜度后 UP DOWN 应该去掉
    prescripion_data_dict['t_base1'] = ["", "UP", "DOWN"]
    return prescripion_data_dict


# 取消工厂订单
def cancel_order(request, lab_order):
    try:
        loc = lab_order_controller()
        data = loc.cancel_lab_order(lab_order)
        if data.code != 0:
            return HttpResponse("取消失败")
    except Exception as e:
        logging.debug("Error==>%s" % e)
        return HttpResponse(e)


# 更新镜架库存数量及预定数量
def update_inventory_qty(request, laborder):
    logging.debug('更新库存及预订数量')
    rst_data = response_message()
    vd = laborder.vendor
    ws_code = 'W02'
    # if vd == '2' or vd == '4' or vd == '7' or vd == '8':
    #     ws_code = 'W01'
    # elif vd == '5' or vd == '6' or vd == '9':
    #     ws_code = 'W02'
    try:
        from wms.models import inventory_delivery
        delivery_list = inventory_delivery.objects.filter(lab_number=laborder.lab_number)
        if len(delivery_list) > 0:
            delivery = delivery_list[0]
            ws_code = delivery.warehouse_code
    except Exception as e:
        ws_code = 'W02'

    try:
        from wms.models import inventory_struct_contoller
        isc = inventory_struct_contoller()
        if laborder.status == 'R2CANCEL':
            if laborder.current_status == '' or laborder.current_status is None or laborder.current_status == 'REQUEST_NOTES':
                rst_data = isc.subtract_reserver_qty(laborder.frame, laborder.quantity)
            else:
                from wms.models import inventory_receipt_control, inventory_delivery
                # oid = inventory_delivery.objects.get(lab_number=laborder.lab_number)
                # ws_code = oid.warehouse.code
                irc = inventory_receipt_control()
                time_now = time.strftime('%Y%m%d', time.localtime(time.time()))  # 默认【当天日期】
                rst_data = irc.add(request, time_now, ws_code, laborder.frame,
                                   0, 'REFUNDS_IN', laborder.quantity, '',
                                   laborder.lab_number)
        else:
            if laborder.status == '' or laborder.status is None or laborder.status == 'REQUEST_NOTES':
                rst_data = isc.subtract_reserver_qty(laborder.frame, laborder.quantity)
            else:
                from wms.models import inventory_receipt_control, inventory_delivery, inventory_delivery_control
                # oid = inventory_delivery.objects.get(lab_number=laborder.lab_number)
                # ws_code = oid.warehouse.code
                irc = inventory_receipt_control()
                time_now = time.strftime('%Y%m%d', time.localtime(time.time()))  # 默认【当天日期】
                rst_data = irc.add(request, time_now, ws_code, laborder.frame,
                                   0, 'REFUNDS_IN', laborder.quantity, '',
                                   laborder.lab_number)
        return rst_data
    except Exception as e:
        logging.debug('ex:' + str(e))
        rst_data.code = -1
        rst_data.capture_execption(e)
        return rst_data


# generate the pgorder
@login_required
@permission_required('oms.GO_VIEW', login_url='/oms/forbid/')
def generate_pg_orders(request):
    line = '------------------------------------------------------------'
    logging.debug(line)
    msg = 'message'
    logging.debug(msg)

    poc = PgOrderController()

    entity_list = entity_id_list()
    if entity_list:

        max_current_entity = generatelogmax()
        logging.debug("max_current_entity==>%s" % max_current_entity)

        with connections['pg_mg_query'].cursor() as cursor:

            sql = const.sql_generate_pg_orders + const.sql_generate_pg_orders_all

            logging.info(sql)
            cursor.execute(sql, [max_current_entity])

            results = namedtuplefetchall(cursor)

            try:
                with transaction.atomic():
                    for i in range(len(entity_list)):

                        """生成pgorder"""
                        for r in range(len(results)):
                            if results[r].entity_id == entity_list[i].entity_id:
                                resulte = results[r]
                                po, flag = poc.gen_pgorder(resulte)

                                break

                        """生成pgorderitem"""
                        if flag:
                            relatelist = []
                            nonelist = []
                            for j in range(len(results)):
                                if entity_list[i].entity_id == results[j].entity_id:
                                    relatelist.append(results[j])

                            logging.debug('relatelist ---------------------------------------->')
                            """生成pgorderitem"""
                            poc.gen_pgorderitem(relatelist, po)

                    generatelog = GenerateLog()
                    generatelog.last_entity = max_current_entity
                    generatelog.current_entity = entity_list[len(entity_list) - 1].entity_id
                    generatelog.save()
                    # return render(request, 'generatetip.html', {'string': 'Successfully generate the pgorder'})
                    return HttpResponse("Successfully generate the pgorder")
            except Exception as e:
                err_msg = e.message
                logging.error(err_msg)
                # return render(request, 'generatetip.html', {'string': err_msg})
                return HttpResponse(err_msg)
    else:
        # return render(request, 'generatetip.html', {'string': 'No new pgorder is to be generated'})
        return HttpResponse("No new pgorder is to be generated")


# 生成指定的pgorder
@login_required
@permission_required('oms.GO_VIEW', login_url='/oms/forbid/')
def generate_spe_pg_order(request):
    incretment_id = request.POST.get('increment_id')

    po = PgOrder()

    poc = PgOrderController()

    try:
        pgorder = po.query_by_id(incretment_id)
        return HttpResponse('The order has been generated pgorder and can not be generated repeatedly.')
    except:
        with connections['pg_mg_query'].cursor() as cursor:
            sql = const.sql_generate_pg_orders + const.sql_generate_pg_orders_spe

            logging.info(sql)
            cursor.execute(sql, [incretment_id])
            results = namedtuplefetchall(cursor)
            if results:
                try:
                    with transaction.atomic():

                        result = results[0]
                        po, flag = poc.gen_pgorder(result)
                        if flag:
                            poc.gen_pgorderitem(results, po)
                        return HttpResponse('Successfully generate the pgorder')
                except Exception as e:
                    logging.debug("Error==>%s" % e)
                    return HttpResponse(e)
            else:
                return HttpResponse('Can not find this order')


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirectLabOrderList(request):
    type = request.GET.get("type")
    if type == None or type == '':
        type = 'all'
    logging.debug("type==>%s" % type)

    timedel = date_delta()  # N天前的日期

    """查询object_type为OLOR，group为1的所有action"""
    ac = Action()
    actions_one = ac.query_actions("OLOR", 1)
    """查询object_type为OLOR，group为0的所有action"""
    actions_zero = ac.query_actions("OLOR", 0)
    olor_zero = []
    olor_one = []
    for zero in actions_zero:
        olor_zero.append(zero)
    for one in actions_one:
        olor_one.append(one)
    if type == 'all':
        pageNum = LabOrder.objects.filter(create_at__gte=timedel, is_enabled=True).count()
        small = '工厂订单V1'
    elif type == 'expired':
        pageNum = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'), create_at__gte=timedel,
                                          set_time__lt=0, is_enabled=True).count()
        small = '超期订单'
    elif type == 'imminent_expiry':
        pageNum = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'), create_at__gte=timedel,
                                          set_time__lte=24, set_time__gt=0,
                                          is_enabled=True).count()
        small = '即将超期'
    logging.debug("pageNum==>%s" % pageNum)

    return render(request, 'laborder_list.html', {'pageNum': json.dumps(pageNum), 'upActions': json.dumps(olor_zero),
                                                  'downActions': json.dumps(olor_one), 'small': small,
                                                  'type': json.dumps(type)})


# LabOrderList
@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def labOrderList(request):
    currPage = request.POST.get('currPage')
    logging.debug("currPage==>" + str(currPage))

    labnum = request.POST.get('labnum')
    logging.debug("labnum==>" + str(labnum))

    begintime = request.POST.get('begintime')
    logging.debug("begintime==>" + str(begintime))

    endtime = request.POST.get('endtime')
    logging.debug("endtime==>" + str(endtime))

    labstatus = request.POST.get('labstatus')
    logging.debug("labstatus==>" + str(labstatus))

    shipdirection = request.POST.get('shipdirection')
    logging.debug("shipdirection==>" + str(shipdirection))

    type = request.POST.get('type')
    logging.debug("type==>%s" % type)

    start_hour = request.POST.get('start_hour')
    logging.debug("start_hour==>%s" % start_hour)

    end_hour = request.POST.get('end_hour')
    logging.debug("end_hour==>%s" % end_hour)

    lens_receive = request.POST.get('lens_receive')
    logging.debug("lens_receive==>%s" % lens_receive)

    limit = True

    filter = {}
    filter['is_enabled'] = True
    if labnum <> '':
        if labnum[0].upper() == settings.BAR_CODE_PREFIX:
            _entity = labnum.upper().lstrip(settings.BAR_CODE_PREFIX)
            filter["id"] = _entity
        else:
            filter["lab_number__contains"] = labnum
        limit = False
    if shipdirection <> '':
        filter["ship_direction"] = shipdirection
    if begintime <> '':
        filter["order_date__gte"] = begintime
    if endtime <> '':
        filter["order_date__lte"] = endtime
    if labstatus <> '':
        if labstatus == 'PRINTED':
            filter['status'] = ''
        else:
            filter["status"] = labstatus
            if labstatus == 'CANCELLED':
                filter['is_enabled'] = False
    if start_hour <> '' and start_hour <> None:
        filter["set_time__gte"] = start_hour
    if end_hour <> '' and end_hour <> None:
        filter["set_time__lte"] = end_hour

    if request.user.id == 5 or request.user.id == 9 or request.user.id == 43:
        #     # lujh & landlady
        filter['vendor'] = '1'

    curr = int(currPage)

    begin = (curr - 1) * 20
    if begin == 0:
        begin = None
        end = 20
    else:
        end = begin + 20

    timedel = date_delta()  # N天前的日期

    if limit:
        filter["create_at__gte"] = timedel

    # 2018.02.20 by guof
    # 将排序恢复为按id倒序
    # queryset = LabOrder.objects.filter(**filter).filter(final_time=None).order_by("-set_time")[begin:end]
    if type == 'all':
        if lens_receive == 'LENS_RECEIVE':
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               is_enabled=True,
                                               ).filter(
                **filter).filter(
                Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).order_by(
                '-id')  # .order_by("set_time")

            queryset_complete = LabOrder.objects.filter(Q(status='SHIPPING') | Q(status='COMPLETE'),

                                                        is_enabled=True).filter(**filter).filter(
                Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).order_by('-id')
        else:
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               ).filter(
                **filter).order_by('-id')  # .order_by("set_time")
            queryset_complete = LabOrder.objects.filter(Q(status='SHIPPING') | Q(status='COMPLETE'),

                                                        ).filter(**filter).order_by('-id')

        queryset_full = []
        queryset_full.extend(queryset)
        queryset_full.extend(queryset_complete)
        los = queryset_full[begin:end]
        logging.debug("los==>%s" % los)
    if type == 'expired':
        if lens_receive == 'LENS_RECEIVE':
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               set_time__lt=0,
                                               ).filter(
                **filter).filter(
                Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).order_by(
                'set_time')[begin:end]
        else:
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               set_time__lt=0,
                                               ).filter(
                **filter).order_by('set_time')[begin:end]
        los = queryset
    elif type == 'imminent_expiry':
        if lens_receive == 'LENS_RECEIVE':
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               set_time__gt=0,
                                               set_time__lte=24, ).filter(**filter).filter(
                Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).order_by(
                'set_time')[begin:end]
        else:
            queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                               set_time__gt=0,
                                               set_time__lte=24, ).filter(
                **filter).order_by('set_time')[begin:end]
        los = queryset

    results = serializers.serialize('json', los)
    return HttpResponse(results)


# 计算条件过滤后的总条数

def labOrderListcount(request):
    queryset = 0
    try:
        labnum = request.POST.get('labnum')
        begintime = request.POST.get('begintime')
        endtime = request.POST.get('endtime')
        labstatus = request.POST.get('labstatus')
        shipdirection = request.POST.get('shipdirection')
        start_hour = request.POST.get('start_hour')
        logging.debug("start_hour==>%s" % start_hour)

        end_hour = request.POST.get('end_hour')
        logging.debug("end_hour==>%s" % end_hour)

        lens_receive = request.POST.get('lens_receive')
        logging.debug("lens_receive==>%s" % lens_receive)
        type = request.POST.get('type')

        timedel = date_delta()  # N天前的日期

        limit = True

        filter = {}
        filter['is_enabled'] = True
        if labnum <> '':
            logging.debug(labnum[0])
            if labnum[0].upper() == settings.BAR_CODE_PREFIX:
                _entity = labnum.upper().lstrip(settings.BAR_CODE_PREFIX)
                filter["id"] = _entity
                logging.debug('----------------------%s' % _entity)
            else:
                filter["lab_number__contains"] = labnum
            limit = False
        if shipdirection <> '':
            filter["ship_direction"] = shipdirection
        if begintime <> '':
            filter["order_date__gte"] = begintime
        if endtime <> '':
            filter["order_date__lte"] = endtime
        if labstatus <> '':
            if labstatus == 'PRINTED':
                filter['status'] = ''
            else:
                filter["status"] = labstatus
                if labstatus == 'CANCELLED':
                    filter['is_enabled'] = False
        if start_hour <> '' and start_hour <> None:
            filter["set_time__gte"] = start_hour
        if end_hour <> '' and end_hour <> None:
            filter["set_time__lte"] = end_hour

        if limit:
            filter["create_at__gte"] = timedel

        if type == 'all':
            if lens_receive == 'LENS_RECEIVE':
                queryset = LabOrder.objects.filter(is_enabled=True).filter(
                    Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).filter(
                    **filter).count()
            else:
                queryset = LabOrder.objects.filter(**filter).count()
        elif type == 'expired':
            if lens_receive == 'LENS_RECEIVE':
                queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'), is_enabled=True,
                                                   set_time__lt=0).filter(**filter).filter(
                    Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).count()
            else:
                queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                                   set_time__lt=0).filter(**filter).count()
        else:
            if lens_receive == 'LENS_RECEIVE':
                queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                                   set_time__lte=24, set_time__gt=0).filter(**filter).filter(
                    Q(status='') | Q(status=None) | Q(status='PRINT_DATE') | Q(status='FRAME_OUTBOUND')).count()
            else:
                queryset = LabOrder.objects.filter(~Q(status='SHIPPING'), ~Q(status='COMPLETE'),
                                                   set_time__lte=24, set_time__gt=0).filter(**filter).count()

        logging.debug("count==>%s" % queryset)
    except Exception as e:
        logging.debug(e.message)

    return HttpResponse(queryset)


# LabOrder详情
@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def labOrderDetail(request):
    labid = request.POST.get("orderNumber")
    queryset = LabOrder.objects.filter(lab_number=labid)
    results = serializers.serialize('json', queryset)
    return HttpResponse(results)


# laborder操作记录
def labOrderLog(request):
    labNumber = request.POST.get('labNumber')
    logging.debug("------------------------------" + str(labNumber))
    orderTrackings = OrderTracking.objects.filter(order_number=labNumber).order_by('-create_at')
    results = serializers.serialize('json', orderTrackings)
    return HttpResponse(results)


@login_required
@permission_required('oms.OA_VIEW', login_url='/oms/forbid/')
def search_address(request):
    id = request.POST.get('id')
    logging.debug("id--------------------------------------------" + str(id))

    ids = PgOrderItem.objects.filter(
        Q(order_number=id) | Q(lab_order_entity__lab_number=id) | Q(order_number__endswith=id)).order_by("-id")
    logging.debug(ids)
    order_numbers = []
    if len(ids) == 0:
        return HttpResponse("False")
    else:
        for id in ids:
            order_numbers.append(id.order_number)

    logging.debug("order_numbers==>" + str(order_numbers))
    with connections['pg_mg_query'].cursor() as cursor:
        sql = '''
            select t0.increment_id,
            t3.firstname,
            t3.lastname,
            t3.postcode,
            t3.street,
            t3.city,
            t3.region,
            t3.country_id,
            t3.telephone
            from sales_order_grid t0 left join sales_order_address t3
            on t0.entity_id=t3.parent_id and t3.address_type='shipping'
            where t0.increment_id in %s
        '''
        logging.info(sql)
        cursor.execute(sql, [order_numbers])
        results = namedtuplefetchall(cursor)
        logging.debug(results)
        results = utilities.convert_to_dicts(results)
        addresses = json.dumps(results)
        # addresses = serializers.serialize('json', results)
        logging.debug(addresses)
        return HttpResponse(addresses)


@login_required
@permission_required('oms.OTRE_VIEW', login_url='/oms/forbid/')
def redirect_order_tracking_report(request):
    date_time = date_delta()
    countPage = OrderTrackingReport.objects.filter(create_at__gte=date_time).count()
    return render(request, 'order_tracking_report_list.html', {'countPage': json.dumps(countPage)})


@login_required
@permission_required('oms.OTRE_VIEW', login_url='/oms/forbid/')
def ordre_tracking_report(request):
    currpage = request.POST.get('currpage')
    currpage = int(currpage)
    begin = (currpage - 1) * 20
    if begin == 0:
        begin = None
        end = 20
    else:
        end = begin + 20
    date_time = date_delta()
    allData = OrderTrackingReport.objects.filter(create_at__gte=date_time).order_by('-order_date')[begin:end]

    data = serializers.serialize('json', allData)

    return HttpResponse(data)


"""按照单号搜索orderTrackinReport"""


def orderTrackingReportFilter(request):
    order_number = request.POST.get("number")
    logging.debug("order_number==>" + str(order_number))

    sign = request.POST.get("sign")
    logging.debug("sign==>" + str(sign))
    if len(order_number) == 3:
        logging.debug("index=3")
        number = "-" + order_number + "-"
        logging.debug("number==>" + str(number))
        if sign == '1':
            queryset = OrderTrackingReport.objects.filter(lab_order_number__contains=number)
            logging.debug("OrderTrackingReport-->queryset==>" + str(queryset))
            results = serializers.serialize('json', queryset)
            return HttpResponse(results)
        else:
            queryset = OrderTrackingReportCS.objects.filter(order_number__contains=number)
            logging.debug("OrderTrackingReportCS-->queryset==>" + str(queryset))
            results = serializers.serialize('json', queryset)
            return HttpResponse(results)

    else:
        if sign == '1':
            queryset = OrderTrackingReport.objects.filter(lab_order_number=order_number)
            logging.debug(queryset)
            results = serializers.serialize('json', queryset)
            return HttpResponse(results)
        else:
            queryset = OrderTrackingReportCS.objects.filter(order_number=order_number)
            logging.debug(queryset)
            results = serializers.serialize('json', queryset)
            return HttpResponse(results)


"""跳转到forbid页面"""


def forbid(request):
    return render(request, 'forbid.html')


# 跳转到添加权限的roles页面
@login_required
@permission_required('oms.is_superuser', login_url='/oms/forbid/')
def rolesHtml(request):
    models = ContentType.objects.values('id', 'model').all()  # .filter(app_label='oms')
    return render(request, 'roles.html', {'models': models})


# 查询所有权限
@login_required
@permission_required('oms.is_superuser', login_url='/oms/forbid/')
def get_permissions(request):
    queryset = Permission.objects.all().order_by('pk')
    results = serializers.serialize('json', queryset)
    return HttpResponse(results)


# 设置权限
@login_required
@permission_required('oms.is_superuser', login_url='/oms/forbid/')
def permissions(request):
    modelname = request.POST.get('contentType')
    logging.debug("modelname==>" + modelname)

    codename = request.POST.get('codeName')
    logging.debug("codename==>" + codename)

    name = request.POST.get('name')
    logging.debug("name==>" + name)

    content_type = ContentType.objects.get(id=modelname)
    permission = Permission.objects.create(codename=codename,
                                           name=name,
                                           content_type=content_type)

    permission.save()
    models = ContentType.objects.values('id', 'model').filter(app_label='oms')
    return render(request, 'roles.html', {'models': models})


# order_tracking
def add_tracklog(laborder, order_number, sku, order_date, user, remark, alias, is_enable=True):
    logging.debug("-----------------begin add_tracklog method-----------------")
    logging.debug("laboorder==>" + str(laborder))
    logging.debug("order_number==>" + order_number)
    logging.debug("sku==>" + str(sku))
    logging.debug("order_date==>" + str(order_date))
    logging.debug("user==>" + str(user))
    logging.debug("remark==>" + remark)
    logging.debug("alias==>" + alias)
    if user:
        user_id = user.id
        user_name = user.username
    else:
        user_id = 0
        user_name = 'system'
    ordertrack = OrderTracking()
    ordertrack.order_number = order_number
    ordertrack.sku = sku
    ordertrack.order_date = order_date
    ordertrack.lab_order_entity = laborder.id
    ordertrack.user_entity = user_id
    ordertrack.username = user_name
    ordertrack.action = alias

    ordertrack.action_value = remark
    ordertrack.is_enabled = is_enable
    ordertrack.save()


# order_tracking_report
def add_order_tracking_report(id, laborder, aliasname, carrier, ship, esttime):
    # logging.critical("----------------------begin add_order_tracking_report method----------------------")
    # logging.critical("id==>" + id)
    # logging.critical("laborder==>" + str(laborder))
    # logging.critical("aliasname==>" + aliasname)
    # logging.critical("carrier==>" + carrier)
    # logging.critical("ship==>" + ship)

    nowtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))

    # logging.critical("nowtime==>" + str(nowtime))
    aliaslow = aliasname.lower()

    try:
        report = OrderTrackingReport.objects.get(lab_order_number=id)

        if aliaslow == 'print_date':
            report.print_date = nowtime
        elif aliaslow == 'frame_outbound':
            report.frame_outbound = nowtime
        elif aliaslow == 'add_hardened':
            report.add_hardened = nowtime
        elif aliaslow == 'rx_lab':
            report.rx_lab = nowtime
        elif aliaslow == 'coating':
            report.coating = nowtime
        elif aliaslow == 'tint':
            report.tint = nowtime
        elif aliaslow == 'lens_receive':
            report.lens_receive = nowtime
        elif aliaslow == 'assembling':
            report.assembling = nowtime
        elif aliaslow == 'initial_inspection':
            report.initial_inspection = nowtime
        elif aliaslow == 'shaping':
            report.shaping = nowtime
        elif aliaslow == 'purging':
            report.purging = nowtime
        elif aliaslow == 'final_inspection':
            report.final_inspection = nowtime
        elif aliaslow == 'order_match':
            report.order_match = nowtime
        elif aliaslow == 'package':
            report.package = nowtime
        elif aliaslow == 'shipping':
            report.shipping = nowtime
        if esttime <> 'null':
            report.estimated_time = esttime

        if carrier <> 'null' and ship <> 'null':
            report.carriers = carrier
            report.shipping_number = ship
            report.final_time = nowtime

        report.save()

    except Exception as e:
        report = OrderTrackingReport()
        report.lab_order_number = id
        report.sku = laborder.frame
        report.order_date = laborder.order_date
        report.lab_order_entity = laborder

        if aliaslow == 'print_date':
            report.print_date = nowtime
        elif aliaslow == 'frame_outbound':
            report.frame_outbound = nowtime
        elif aliaslow == 'add_hardened':
            report.add_hardened = nowtime
        elif aliaslow == 'rx_lab':
            report.rx_lab = nowtime
        elif aliaslow == 'coating':
            report.coating = nowtime
        elif aliaslow == 'tint':
            report.tint = nowtime
        elif aliaslow == 'lens_receive':
            report.lens_receive = nowtime
        elif aliaslow == 'assembling':
            report.assembling = nowtime
        elif aliaslow == 'initial_inspection':
            report.initial_inspection = nowtime
        elif aliaslow == 'shaping':
            report.shaping = nowtime
        elif aliaslow == 'purging':
            report.purging = nowtime
        elif aliaslow == 'final_inspection':
            report.final_inspection = nowtime
        elif aliaslow == 'order_match':
            report.order_match = nowtime
        elif aliaslow == 'package':
            report.package = nowtime
        elif aliaslow == 'shipping':
            report.shipping = nowtime
        if esttime <> 'null':
            report.estimated_time = esttime

        if carrier <> 'null' and ship <> 'null':
            report.carriers = carrier
            report.shipping_number = ship
            report.final_time = nowtime

        report.save()


# order_tracking_report_cs
def addOrderTrackingReportCs(pgorder_number, shipping_method, id, sku, order_date, field, carrier, ship, esttime):
    # logging.debug("id==>" + id)
    # logging.debug("sku==>" + sku)
    # logging.debug("order_date==>" + str(order_date))
    # logging.debug("field==>" + field)
    # logging.debug("carrier==>" + carrier)
    # logging.debug("ship==>" + ship)

    nowtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
    logging.debug("nowtime==>" + nowtime)

    cs_manufacturing = ['PRINT_DATE', 'FRAME_OUTBOUND', 'RX_LAB', 'ADD_HARDENED', 'COATING', 'TINT', 'LENS_RECEIVE',
                        'ASSEMBLING',
                        'INITIAL_INSPECTION', 'SHAPING', 'PURGING', 'FINAL_INSPECTION', 'ORDER_MATCH']
    try:
        otrc = OrderTrackingReportCS.objects.get(order_number=id)
        if field in cs_manufacturing:
            otrc.cs_status = 'CS_MANUFACTURING'
        elif field == 'PACKAGE':
            otrc.cs_status = 'CS_PACKING'
        elif field == 'SHIPPING':
            otrc.cs_status = 'CS_SHIPPED'
        elif field == 'ONHOLD':
            otrc.cs_status = 'CS_ONHOLD'
        elif field == 'CANCELLED':
            otrc.cs_status = 'CS_CANCELLED'
        elif field == 'REDO':
            otrc.cs_status = 'CS_REDO'

        if carrier <> 'null' and ship <> 'null':
            otrc.carriers = carrier
            otrc.shipping_number = ship
            otrc.final_time = nowtime
        if esttime <> 'null':
            otrc.estimated_time = esttime

        otrc.save()

    except:
        otrc = OrderTrackingReportCS()
        otrc.pgorder_number = pgorder_number
        otrc.shipping_method = shipping_method
        otrc.order_number = id
        otrc.sku = sku
        otrc.order_date = order_date

        if field in cs_manufacturing:
            otrc.cs_status = 'CS_MANUFACTURING'
        elif field == 'PACKAGE':
            otrc.cs_status = 'CS_PACKING'
        elif field == 'SHIPPING':
            otrc.cs_status = 'CS_SHIPPED'
        elif field == 'ONHOLD':
            otrc.cs_status = 'CS_ONHOLD'
        elif field == 'CANCELLED':
            otrc.cs_status = 'CS_CANCELLED'
        elif field == 'REDO':
            otrc.cs_status = 'CS_REDO'

        if carrier <> 'null' and ship <> 'null':
            otrc.carriers = carrier
            otrc.shipping_number = ship
            otrc.final_time = nowtime

        if esttime <> 'null':
            otrc.estimated_time = esttime

        otrc.save()


@login_required
def laborder_status(request):
    logging.debug("-------------------begin laborder_status method--------------------")
    content = request.POST.get('content')
    logging.debug("content==>" + content)
    labid = request.POST.get('labid')
    logging.debug("labid==>" + labid)
    aliasname = request.POST.get('aliasname')
    logging.debug("aliasname==>" + aliasname)
    carrier = request.POST.get('carrier')
    logging.debug("carrier==>" + carrier)
    ship = request.POST.get('ship')
    logging.debug("ship==>" + ship)
    formDate = request.POST.get('form')
    logging.debug("formDate==>%s" % formDate)
    return laborder_status_backend(
        request,
        content,
        labid,
        aliasname,
        carrier,
        ship,
        formDate
    )


def laborder_status_backend(
        request,
        content,
        labid,
        aliasname,
        carrier,
        ship,
        formDate
):
    try:
        ordertracking = OrderTracking.objects.get(order_number=labid, action=aliasname, is_enabled=True)
        return HttpResponse('False')
    except:
        try:
            with transaction.atomic():
                laborder = LabOrder.objects.get(lab_number=labid)
                logging.debug('begin get laborder: %s' % laborder.lab_number)

                poi = PgOrderItem.objects.get(pk=laborder.base_entity)
                logging.debug('pgorder item: %s' % poi.order_number)

                is_formdate_none = (formDate == None)
                logging.debug('is_formdate_none: %s' % is_formdate_none)

                if formDate <> None and formDate <> '':
                    loqc = LabOrder_QualityControl()
                    loqc.add_object(laborder, formDate)
                    add_tracklog(laborder, labid, laborder.frame, laborder.order_date, request.user, content, aliasname,
                                 is_enable=False)
                    ot = OrderTracking()
                    ot.modify_enable(labid)
                    return HttpResponse('Error')

                laborder.status = aliasname
                logging.debug('updated laborder.status ....')
                nowtime = timezone.now().strftime('%Y-%m-%d %H:%M:%S')  # 当前时间
                if aliasname == 'SHIPPING':
                    # nowtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
                    laborder.final_time = nowtime
                    laborder.carriers = carrier
                    laborder.shipping_number = ship

                    poi = PgOrderItem.objects.get(id=laborder.base_entity)
                    poi.final_date = nowtime
                    poi.pg_order_entity.final_date = nowtime
                    poi.pg_order_entity.save()
                    poi.save()

                    # '''判断并修改pgorder的状态为发货'''
                    # pgoi = PgOrderItem()
                    # pgois = pgoi.filter_by_id(poi.order_number)
                    # flag = pgoi.judge_status(pgois)
                    # logging.debug("--------------flag--------->%s"%flag)
                    # if flag:  #flag为True，则改变状态，否则不改变
                    #     po = poi.pg_order_entity
                    #     po.modify_spe_status("shipped")

                elif aliasname == 'FRAME_OUTBOUND':
                    laborder.act_lens_sku = laborder.lens_sku
                    laborder.act_lens_name = laborder.lens_name
                    laborder.lens_delivery_time = nowtime

                laborder.save()

                '''判断并修改pgorder的状态为发货'''
                pgoi = PgOrderItem()
                pgois = pgoi.filter_by_id(poi.order_number)  # 查询当前laborder所对应的pgorder下的所有pgorderitem
                flag = pgoi.judge_status(pgois)  # 判断同一订单下的的所用laborder的状态
                logging.debug("--------------flag--------->%s" % flag)
                if flag:  # flag为True，则改变状态，否则不改变
                    po = poi.pg_order_entity
                    po.modify_spe_status("shipped")

                user = request.user
                logging.debug("user----------------------------------------" + str(user))
                # add order_tracking log

                add_tracklog(laborder, labid, laborder.frame, laborder.order_date, user, content, aliasname)

                # add order_tracking_report
                add_order_tracking_report(labid, laborder, aliasname, carrier, ship, "null")

                # add order_tracking_report_cs
                addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labid, laborder.frame,
                                         laborder.order_date, aliasname, carrier, ship, 'null')

            return HttpResponse("True")
        except Exception as e:

            logging.debug("# Exception #:" + str(e))
            errors = str(e)
            return HttpResponse(errors)


# 未加权限校验
@login_required
def final_inspection(request):
    lab_entity = request.POST.get('lab_entity')
    rm = res_msg.response_dict()
    try:
        lbo = LabOrder.objects.get(pk=lab_entity)
        if lbo.status != "GLASSES_RECEIVE":
            rm["code"] = 4
            rm["message"] = "只有【成镜收货】状态才能变更为【终检】状态"
            return JsonResponse(rm)
        else:
            with transaction.atomic():
                lbo.status = "FINAL_INSPECTION"
                lbo.save()
                # 添加订单记录
                tloc = tracking_lab_order_controller()
                tloc.tracking(lbo, request.user, "FINAL_INSPECTION", "终检")
                rm["message"] = "订单已进入【终检】状态"
                return JsonResponse(rm)

    except Exception as e:
        logging.debug("Error====>%s" % e)
        rm["code"] = "-1"
        rm["message"] = "Error: %s" % e
        return JsonResponse(rm)


def action_status(request):
    # logging.debug("****************************************&&&&&&&&&&&&")
    orderNumber = request.POST.get('order_number')
    logging.debug(orderNumber)
    queryset = OrderTracking.objects.filter(order_number=orderNumber, is_enabled=True).order_by("-create_at")

    results = serializers.serialize('json', queryset)

    return HttpResponse(results)


@login_required
@permission_required('oms.OTRC_VIEW', login_url='/oms/forbid/')
def redirectReportCSList(request):
    date_time = date_delta()
    pageNum = OrderTrackingReportCS.objects.filter(create_at__gte=date_time).count()
    return render(request, 'order_track_report_cs.html', {'pageNum': json.dumps(pageNum)})


@login_required
@permission_required('oms.OTRC_VIEW', login_url='/oms/forbid/')
def orderTrackingReportCsList(request):
    curr = request.POST.get('currPage')
    currPage = int(curr)
    begin = (currPage - 1) * 20
    if begin == 0:
        begin = None
        end = 20
    else:
        end = begin + 20
    date_time = date_delta()
    queryset = OrderTrackingReportCS.objects.filter(create_at__gte=date_time).order_by('-order_date')[begin:end]
    logging.debug("queryset==>%s" % queryset)

    results = serializers.serialize('json', queryset)
    return HttpResponse(results)


# laborder点击预计时间方法
def esTime(request):
    logging.debug("----------------------------------------begin esTime method----------------------------------------")

    try:
        estTime = request.POST.get('shiptime')
        logging.debug("estTime----------------------------------------" + str(estTime))

        content = request.POST.get('content')
        logging.debug("content----------------------------------------" + content)
        alias = request.POST.get('alias')
        logging.debug("alias----------------------------------------" + alias)
        labid = request.POST.get('labid')
        logging.debug("labid----------------------------------------" + labid)
        with transaction.atomic():
            laborder = LabOrder.objects.get(lab_number=labid)
            poi = PgOrderItem.objects.get(pk=laborder.base_entity)
            '''------'''
            laborder.estimated_date = estTime
            laborder.save()
            lab_number_split = labid.split("-")
            lab_number_contain = "-" + lab_number_split[1] + "-"
            laborders = LabOrder.objects.filter(lab_number__contains=lab_number_contain).values("estimated_date")
            logging.debug("--------laborders==>%s" % laborders)
            # poi = PgOrderItem.objects.values("order_number").get(pk=laborder.base_entity)
            poi = PgOrderItem.objects.get(pk=laborder.base_entity)
            '''----'''
            poi.estimated_ship_date = estTime
            poi.save()

            logging.debug("---------poi==>%s" % poi)
            po = PgOrder.objects.get(order_number=poi.order_number)
            date_list = []
            for lo in laborders:
                if lo['estimated_date'] <> '' and lo['estimated_date'] <> None:
                    date_list.append(lo['estimated_date'])
            logging.debug("--------date_list==>%s" % date_list)
            if len(date_list) == laborders.count():
                max_date = max(date_list)
                logging.debug("--------max_date==>%s" % max_date)
                if po.ship_direction == 'STANDARD':
                    logging.debug("!!!!!!!!")
                    po.estimated_ship_date = (max_date + datetime.timedelta(days=3)).strftime("%Y-%m-%d %H:%M:%S")
                    po.save()
                else:
                    logging.debug("@@@@@@@@@@")
                    po.estimated_ship_date = (max_date + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")
                    po.save()

            user = request.user

            tloc = tracking_lab_order_controller()
            tloc.tracking(laborder, request.user, alias, '预计完成时间', estTime)

            # add_tracklog(laborder, labid, laborder.frame, laborder.order_date, user, content, alias)
            add_order_tracking_report(labid, laborder, alias, 'null', 'null', estTime)
            addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labid, laborder.frame, laborder.order_date,
                                     alias, 'null', 'null', estTime)
        return HttpResponse('True')
    except Exception as e:
        logging.debug("Error==>%s" % e)
        return HttpResponse(str(e))


# laborder点击完成操作
def confirmAction(request):
    logging.debug("------------------------------begin confirmAction method----------------------------------")
    labid = request.POST.get('labid')
    logging.debug("labid==>" + labid)

    alias = request.POST.get('alias')
    logging.debug("alias==>" + alias)

    content = request.POST.get('content')
    logging.debug("content==>" + content)

    try:
        with transaction.atomic():

            laborder = LabOrder.objects.get(lab_number=labid)
            laborder.status = alias
            laborder.save()
            user = request.user

            add_tracklog(laborder, labid, laborder.frame, laborder.order_date, user, content, alias)
        return HttpResponse('True')
    except Exception as e:
        return HttpResponse(str(e))


# laborder暂停取消操作
@login_required
@permission_required('oms.ACT_ONHOLD', login_url='/oms/forbid/')
def labOrderHold(request):
    logging.debug("begin labOrderHold method")

    labNumber = request.POST.get("labid")  # 获取order_number
    logging.debug("labNumber==>" + str(labNumber))

    alias = request.POST.get('alias')  # 获取点击状态
    logging.debug("alias==>" + str(alias))

    content = request.POST.get('content')  # 获取点击按钮的内容
    logging.debug("content==>" + str(content))

    # ot = OrderTracking.objects.get(order_number=labNumber,action=alias)

    user = request.user  # 获取操作人的信息
    laborder = LabOrder.objects.get(lab_number=labNumber)  # 获取当前点击的LabOrder对象
    poi = PgOrderItem.objects.get(pk=laborder.base_entity)

    with transaction.atomic():
        try:
            if alias == 'ONHOLD':

                laborder.status = alias  # 改变laborder的状态为当前点击状态
                laborder.save()

                add_tracklog(laborder, labNumber, laborder.frame, laborder.order_date, user, content,
                             alias)  # 在ordertracking中添加操作记录

                addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labNumber, laborder.frame,
                                         laborder.order_date, alias, 'null', 'null',
                                         'null')  # 修改orderTrackingReportCs中记录
                # addOrderTrackingReportCs(labNumber, laborder.frame, laborder.order_date, alias, 'null','null', 'null')

            else:
                statusList = ["ONHOLD", "CANCEL_HOLD", "ESTIMATED_TIME"]
                ordertracks = OrderTracking.objects.filter(order_number=labNumber).order_by(
                    "-create_at")  # 获取当前LabOrder的操作记录

                toggle = True
                for ordertrack in ordertracks:
                    if ordertrack.action not in statusList:
                        toggle = False
                        status = ordertrack.action
                        laborder.status = status  # 将laborder的状态换到上一个状态
                        laborder.save()

                        add_tracklog(laborder, labNumber, laborder.frame, laborder.order_date, user, content,
                                     alias)  # 在ordertracking中添加操作记录

                        addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labNumber, laborder.frame,
                                                 laborder.order_date, status, 'null', 'null',
                                                 'null')  # 修改orderTrackingReportCs中记录

                        break
                if toggle:
                    laborder.status = ''
                    laborder.save()

                    add_tracklog(laborder, labNumber, laborder.frame, laborder.order_date, user, content, alias)

            return HttpResponse("True")  # 操作成功  返回标识
        except Exception as e:
            logging.error("Error==>" + str(e))
            return HttpResponse(str(e))


# Lab Order V2 暂停/取消暂停 操作
@login_required
@permission_required('oms.ACT_ONHOLD', login_url='/oms/forbid/')
def lab_order_hold_v2(request):
    # 获取 Lab order number / 点击时状态属性 / 按钮内容 / 操作人的信息
    labNumber = request.POST.get('lab_num')
    reason = request.POST.get('reason')
    exclude_days = request.POST.get('exclude_days', '0')
    user = request.user
    # poi = PgOrderItem.objects.get(pk=laborder.base_entity)
    try:
        laborder = LabOrder.objects.get(lab_number=labNumber)
        lbc = lab_order_control()
        tloc = tracking_lab_order_controller()
        if laborder.status != "ONHOLD" and laborder.status != "R2HOLD":
            comments_inner_n = laborder.comments_inner + ";" + reason
            data_dict = {
                "lab_number": labNumber,
                "comments_inner": comments_inner_n,
                'status': 'R2HOLD',
                'current_status': laborder.status
            }
            # 修改状态
            lbc.change_status(data_dict)
            # 写入申请列表
            cosac = hold_cancel_request_control()
            res = cosac.add(request, labNumber, laborder.status, 'ONHOLD', reason)
            # 写操作记录
            tloc.tracking(laborder, user, "R2HOLD", "申请暂停", reason)
        else:
            comments_inner_n = laborder.comments_inner + ";" + reason
            new_exclude_days = float(laborder.exclude_days) + float(exclude_days)
            data_dict = {
                "lab_number": labNumber,
                "comments_inner": comments_inner_n,
                'status': laborder.current_status,
                'current_status': '',
                'exclude_days': new_exclude_days
            }
            # 修改状态
            lbc.change_status(data_dict)

            if laborder.status == 'R2HOLD':  # 如果是申请暂停状态
                # 写tracking
                tloc.tracking(laborder, user, "CANCEL_R2HOLD", '取消申请暂停', reason)
                # 更新暂停申请列表
                with connections['default'].cursor() as cursor:
                    update_sql = u"""
                    UPDATE oms_hold_cancel_request SET `is_handle`='%s', reply="%s", reply_username='%s' WHERE lab_number='%s'
                    """ % (1, reason, user.username, labNumber)
                    cursor.execute(update_sql)
            else:
                # 写tracking
                tloc.tracking(laborder, user, "CANCEL_HOLD", '取消暂停', reason)

        tracking = tloc.get_order_history(laborder.lab_number).obj
        newlaborder = LabOrder.objects.get(lab_number=labNumber)
        return render(request, "laborder_detail_v2.pspf.html", {
            'item': newlaborder,
            'tracking': tracking
        })
    except Exception as e:
        logging.error("Error==>" + str(e))
        return HttpResponse(str(e))


# lab_order订单取消操作
@login_required
@permission_required('oms.PORL_CANCELLED', login_url='/oms/forbid/')
def labOrderCancelled(request):
    logging.debug("begin labOrderCancelled method")

    labNumber = request.POST.get("labid")  # 获取order_number
    logging.debug("labNumber==>" + str(labNumber))

    alias = request.POST.get('alias')  # 获取点击状态
    logging.debug("alias==>" + str(alias))

    content = request.POST.get('content')  # 获取点击按钮的内容
    logging.debug("content==>" + str(content))

    reason = request.POST.get('reason')
    logging.debug("reason==>%s" % reason)

    user = request.user  # 获取操作人的信息
    laborder = LabOrder.objects.get(lab_number=labNumber)  # 获取当前点击的LabOrder对象
    poi = PgOrderItem.objects.get(pk=laborder.base_entity)
    with transaction.atomic():
        try:
            laborder.status = 'CANCELLED'  # 修改laborder对象的状态
            laborder.is_enabled = False  # 修改laborder可用状态
            laborder.comments = reason
            laborder.save()

            # 修改对应的pgorderitem的状态为取消
            poi = PgOrderItem()
            pgorderitem = poi.query_by_id(laborder.base_entity)
            pgorderitem.status = 'canceled'
            pgorderitem.save()

            pois = poi.filter_by_id(pgorderitem.order_number)
            flag = True
            for pgi in pois:
                if pgi.status <> 'canceled':
                    flag = False
            if flag:
                pgorder = pgorderitem.pg_order_entity
                pgorder.status = 'canceled'
                pgorder.save()

            # 在ordertracking中添加操作记录
            add_tracklog(laborder, labNumber, laborder.frame, laborder.order_date, user, content,
                         alias)
            addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labNumber, laborder.frame,
                                     laborder.order_date, alias, 'null', 'null', 'null')
            return HttpResponse('True')
        except Exception as e:
            logging.debug("Error==>%s" % e)
            return HttpResponse('Error')


# lab order v2订单取消操作
@login_required
@permission_required('oms.PORL_CANCELLED', login_url='/oms/forbid/')
def lab_order_cancelled_v2(request):
    try:
        lab_number = request.POST.get("lab_num")
        logging.debug(lab_number)
        reason = request.POST.get("reason")
        action = request.POST.get('action')
        user = request.user
        laborder = LabOrder.objects.get(lab_number=lab_number)
        lbc = lab_order_control()
        tloc = tracking_lab_order_controller()
        if laborder.status == 'R2CANCEL':
            comments_inner_n = laborder.comments_inner + ";" + reason
            data_dict = {
                "lab_number": lab_number,
                "comments_inner": comments_inner_n,
                'status': laborder.current_status,
                'current_status': ''
            }
            # 修改状态
            lbc.change_status(data_dict)
            # 写tracking
            tloc.tracking(laborder, user, "CANCEL_R2CANCEL", '取消申请取消', reason)
            with connections['default'].cursor() as cursor:
                update_sql = u"""
                                UPDATE oms_hold_cancel_request SET `is_handle`='%s', reply="%s", reply_username='%s' WHERE lab_number='%s'
                                """ % (1, reason, user.username, lab_number)
                cursor.execute(update_sql)
        else:
            comments_inner_n = laborder.comments_inner + ";" + reason
            data_dict = {
                "lab_number": lab_number,
                "comments_inner": comments_inner_n,
                'status': 'R2CANCEL',
                'current_status': laborder.status
            }
            # 修改状态
            lbc.change_status(data_dict)
            # 添加到申请列表
            cosac = hold_cancel_request_control()
            res = cosac.add(request, laborder.lab_number, laborder.status, 'CANCELLED', reason)

            # 获取 tracking_lab_order_controller 并跟踪操作

            tloc.tracking(laborder, user, "R2CANCEL", '申请取消', reason)
            # 重新查询获得订单状态
        tracking = tloc.get_order_history(laborder.lab_number).obj
        newlaborder = LabOrder.objects.get(lab_number=lab_number)
        return render(request, "laborder_detail_v2.pspf.html", {
            'item': newlaborder,
            'tracking': tracking
        })
    except Exception as e:
        logging.debug("Error==>%s" % e)
        return HttpResponse('Error')


# laborder订单重做
@login_required
@permission_required('oms.PORL_REDO', login_url='/oms/forbid/')
def labOrderRedo(request):
    _form_data = {}
    _form_data['request_feature'] = 'All LAB Orders V2'
    try:
        logging.debug("begin labOrderCancelled method")
        labNumber = request.GET.get("labid")  # 获取order_number
        logging.debug("labNumber==>" + str(labNumber))
        alias = request.GET.get('alias')  # 获取点击状态
        logging.debug("alias==>" + str(alias))
        content = request.GET.get('content')  # 获取点击按钮的内容
        logging.debug("content==>" + str(content))

        lo = LabOrder()
        if request.method == 'GET':

            laborder = lo.query_by_id(labNumber)
            laborderlist = lo.filter_by_id(labNumber)

            laborder.is_remake_order = True
            laborder.status = ''

            laborderform = LabOrderForm(instance=laborder)
            return render(request, 'laborder_redo.html', {
                'formset': laborderform,
                'origin': laborderlist,
                'alias': alias,
                'content': content
            })
        else:
            with transaction.atomic():
                user = request.user
                del_old = request.POST.get("del_old")
                to_new = request.POST.get("to_new")
                labNumber = request.POST.get("lab_num")  # 获取order_number
                alias = request.POST.get('alias')  # 获取点击状态
                content = request.POST.get('content')  # 获取点击按钮的内容
                tloc = tracking_lab_order_controller()
                original_laborder = lo.query_by_id(labNumber)  # 获取原订单
                # 验证仓位
                lc = locker_controller()
                lockers_data = lc.redo_job_order(labNumber)
                if lockers_data.code == 0:
                    # 先创建新订单
                    newLabOrderForm = LabOrderForm(request.POST)
                    if newLabOrderForm.is_valid():
                        new_lab_order = newLabOrderForm.save()
                        # 保存用户信息
                        new_lab_order.user_id = user.id
                        new_lab_order.user_name = user.username
                        new_lab_order.ship_direction = original_laborder.ship_direction
                        if to_new == "True":
                            new_lab_order.status = ""
                        else:
                            if del_old != "True":
                                raise DataError
                            # 增加 laborder_request_note
                            lrn = laborder_request_notes()
                            lrn.lab_number = new_lab_order.lab_number
                            lrn.laborder_id = new_lab_order.id
                            lrn.count = 1
                            lrn.vendor = new_lab_order.vendor
                            lrn.save()

                            lrn_lines = laborder_request_notes_line()
                            lrn_lines.lrn = lrn
                            lrn_lines.laborder_id = new_lab_order.id
                            lrn_lines.lab_number = new_lab_order.lab_number
                            lrn_lines.index = 1
                            lrn_lines.frame = new_lab_order.frame
                            lrn_lines.laborder_entity = new_lab_order
                            lrn_lines.lens_type = new_lab_order.lens_type
                            lrn_lines.quantity = new_lab_order.quantity
                            lrn_lines.order_date = new_lab_order.order_date
                            lrn_lines.order_created_date = new_lab_order.create_at
                            lrn_lines.save()

                            new_lab_order.act_ship_direction = original_laborder.act_ship_direction
                            new_lab_order.ship_direction = original_laborder.ship_direction
                            new_lab_order.status = "FRAME_OUTBOUND"
                        new_lab_order.save()
                        try:
                            from oms.controllers.lab_order_controller import lab_order_controller
                            loc = lab_order_controller()
                            loc.post_mrp(new_lab_order)
                        except Exception as ex:
                            logging.critical(str(ex))
                        # 如果是重做单  re_or_z内容用来标识申请取消类型，不可随意更改
                        if alias == 'REDO':
                            re_or_z = '重做订单取消原单(自动申请)'
                            # 设置原订单状态
                            original_laborder.has_remake_orders = True
                            original_laborder.save()
                            tloc.tracking(original_laborder, user, alias, content, "")
                        else:
                            re_or_z = '转单取消原单(自动申请)'

                        # 取消原单
                        # wj 2019.06.14 去掉网站库存同步
                        # 2019.09.28 改成申请取消原单 --by gaoyu
                        if del_old == 'True':
                            # 更新lab_order状态为 申请取消
                            data_dict = {
                                "lab_number": original_laborder.lab_number,
                                "status": 'R2CANCEL',
                                "current_status": original_laborder.status
                            }
                            lbc = lab_order_control()
                            lbc.change_status(data_dict)
                            # 添加到申请列表
                            cosac = hold_cancel_request_control()
                            cosac.add(request, original_laborder.lab_number, original_laborder.status, 'CANCELLED',
                                      re_or_z)
                            # 写原单的日志
                            tloc.tracking(original_laborder, user, "R2CANCEL", "申请取消", re_or_z)

                        # 处理对应PgOrder
                        poi = PgOrderItem.objects.get(pk=original_laborder.base_entity)
                        if alias == 'REDO':
                            po = PgOrder()
                            pgorder = po.query_by_id(poi.order_number)
                            pgorder.is_required_return_lable = True
                            pgorder.save()
                            addOrderTrackingReportCs(poi.order_number, poi.shipping_method, labNumber,
                                                     original_laborder.frame,
                                                     original_laborder.order_date, alias, 'null', 'null', 'null')

                        rm = redirect_laborder_list_v2(request)
                        return rm

                        # if alias == 'REDO':
                        #     laborderlist = redirectLabOrderList(request)
                        #     return laborderlist
                        # else:

                    else:
                        _form_data['error_message'] = newLabOrderForm
                        return render(request, "exceptions.html", {
                            'form_data': _form_data,
                        })

                else:
                    _form_data['error_message'] = lockers_data.message
                    return render(request, "exceptions.html", {
                        'form_data': _form_data,
                    })
            '''
            alias=='TRANSFER'
            '''
    except Exception as e:
        logging.debug('Exception: %s' % e)
        _form_data['exceptions'] = e
        _form_data['error_message'] = str(e)
        return render(request, "exceptions.html", {
            'form_data': _form_data,
        })


def redirectPgOrderList(request):
    page = request.GET.get('page', 1)
    currentPage = int(page)
    in_lab_status = request.GET.get('filter', 'a')
    logging.debug("in_lab_status==>%s" % in_lab_status)

    timedel = date_delta()  # 获取N天前的日期
    filter = {}
    filter["create_at__gte"] = timedel
    filter['is_enabled'] = True
    if in_lab_status == 'false':  # 查询没有生成laborder的pgorder
        filter["is_inlab"] = False
        results = PgOrder.objects.filter(Q(status='processing') | Q(status='holded')).filter(**filter).order_by("-id")
    else:

        if in_lab_status == "true":  # 查询生成laborder的pgorder
            filter["is_inlab"] = True

        elif in_lab_status == 'issue':  # 查询生成laborder但是没有发送地址的pgorder
            filter['is_inlab'] = True
            filter['is_shiped_api'] = False
            filter['status'] = 'processing'
        elif in_lab_status == 'lable':  # 查询有回退标签的pgorder
            filter['is_required_return_lable'] = True

        elif in_lab_status == 'reviewed':
            filter['status_control'] = 'REVIEWED'
            filter['status'] = 'processing'
            filter['is_inlab'] = False

    logging.debug(filter)
    logging.debug(in_lab_status)

    # if in_lab_status 为a 则查询所有的订单
    results = PgOrder.objects.filter(**filter).order_by("-id")

    paginator = Paginator(results, 20)  # Show 20 contacts per page

    try:
        contacts = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        contacts = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        contacts = paginator.page(paginator.num_pages)
    return render(request, 'pgorder_list.html',
                  {'list': contacts, 'currentPage': currentPage, 'paginator': paginator,
                   'pgorderitem': 'New Mg Order List', 'requestUrl': '/oms/redirectPgOrderList',
                   'filter': in_lab_status})


def pgOrderList(request):
    currPage = request.POST.get("currPage")
    inlab_status = request.POST.get("inlabstatus")
    logging.debug("status==>%s" % inlab_status)
    po = PgOrder()
    results = po.pgOrderList(currPage)
    return HttpResponse(results)


def poiFormset(request):
    formsets = modelformset_factory(PgOrderItem, fields=(
        'comments_inner', 'order_number', 'lab_status', 'item_id', 'order_date', 'shipping_method',
        'attribute_set_name',
        'pg_order_entity', 'lab_order_entity', 'lab_order_number',
        'comments', 'size', 'instruction',
        'lens_height', 'lens_width', 'thumbnail',
        'quantity', 'frame', 'lens_sku', 'lens_name',
        'coating_sku', 'coating_name', 'coating_sku', 'coating_name',
        'coating_sku', 'coating_name',
        'pal_design_sku', 'pal_design_name',
        'od_sph', 'od_cyl', 'od_axis',
        'os_sph', 'os_cyl', 'os_axis', 'pd', 'is_singgle_pd', 'od_pd',
        'os_pd',
        'tint_sku', 'tint_name', 'od_add', 'os_add', 'od_prism',
        'od_base', 'os_prism', 'os_base', 'od_prism1', 'od_base1',
        'os_prism1', 'os_base1', 'tag', 'so_type', 'product_index'), extra=0, widgets={
        'comments': Fwidgets.Textarea(attrs={'rows': 3, 'cols': 50})
    })
    return formsets


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirectPgOrderItemList(request):
    number = request.GET.get("number")
    logging.debug("numbr==>" + str(number))

    currentpage = request.GET.get('page')
    logging.debug("currentpage==>%s" % currentpage)

    filter = request.GET.get('filter')
    logging.debug("filter==>%s" % filter)

    po = PgOrder()
    pgorder = po.query_by_id(number)
    pgorderformdetail = PgOrderFormDetail(instance=pgorder)
    formsets = poiFormset(request)
    formset = formsets(queryset=PgOrderItem.objects.filter(order_number=number))

    product_image_path_prefix = const.PRODUCT_IMAGE_PREPATH
    # product_image_path_prefix = 'https://test.kyoye.com/media/catalog/product'

    return render(request, 'pgorderitem_list.html', {'order_number': number, 'base_entity': pgorder.base_entity,
                                                     'form': pgorderformdetail, 'formsets': formset,
                                                     'currentpage': currentpage, 'filter': json.dumps(filter),
                                                     'prd_img_pre': product_image_path_prefix
                                                     })


# pgorder更新
@login_required
@permission_required('oms.OPOR_SAVE', login_url='/oms/forbid/')
def pgOrderUpdate(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>%s" % order_number)
    po = PgOrder()
    pgorder = po.query_by_id(order_number)
    logging.debug("pgorder==>%s" % pgorder)
    if request.method == 'GET':
        return render(request, 'pgorder_list.html')
    elif request.method == 'POST':
        fm_result = PgOrderFormDetail(request.POST, instance=pgorder)
        if fm_result.is_valid():
            fm_result.save()
            formsets = poiFormset(request)
            formset = formsets(queryset=PgOrderItem.objects.filter(order_number=order_number))
            return render(request, 'pgorderitem_list.html',
                          {'order_number': order_number, 'form': fm_result, 'formsets': formset})


# 更新pgorderitem
def pgOrderItemUpdate(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>%s" % order_number)

    pgorder = PgOrder.objects.get(order_number=order_number)
    pgorderformdetail = PgOrderFormDetail(instance=pgorder)
    if request.method == 'GET':
        return render(request, 'pgorder_list.html')
    elif request.method == 'POST':
        formsets = poiFormset(request)
        formset = formsets(request.POST)
        if formset.is_valid():
            formset.save()

            return render(request, 'pgorderitem_list.html',
                          {'order_number': order_number, 'form': pgorderformdetail, 'formsets': formset})


@login_required
@permission_required('oms.ACT_REVIEW', login_url='/oms/forbid/')
def pg_order_review(request):
    order_number = request.POST.get('order_number')
    act = 'review'
    poc = PgOrderController()
    rv = poc.actions(request, order_number, act)

    # review 时调用mg接口 以阻止用户重复下单  post_order_comment()
    try:
        pgo = PgOrder.objects.get(order_number=order_number)
        # 添加订单历史信息
        post_order_comment_v2(pgo.base_entity, 'PG_ORDER_REVIEWED', 'processing', pgo.create_at)
        # 更新MG订单的状态
        # order_change_status(pgo.base_entity, pgo.order_number, 'in_lab', 'in_lab')
    except:
        pass

    return HttpResponse(rv)


@login_required
@permission_required('oms.ACT_CANCEL_REVIEW', login_url='/oms/forbid/')
def pg_order_cancel_review(request):
    order_number = request.POST.get('order_number')
    act = 'cancel_review'
    poc = PgOrderController()
    rv = poc.actions(request, order_number, act)
    return HttpResponse(rv)


@login_required
@permission_required('oms.ACT_APPROVE', login_url='/oms/forbid/')
def pg_order_approve(request):
    order_number = request.POST.get('order_number')
    act = 'approve'
    poc = PgOrderController()
    rv = poc.actions(request, order_number, act)
    return HttpResponse(rv)


def pg_order_actions(order_number, action):
    order_number = eval(order_number)

    po = PgOrder.objects.get(order_number=order_number)

    if po.status <> 'processing':
        return HttpResponse('The order status is not processing, and Laborder cannot be generated.')

    if po.is_inlab == False:
        return HttpResponse("Success")


@login_required
@permission_required('oms.ACT_GLO', login_url='/oms/forbid/')
def generateLabOrders(request):
    order_number = request.POST.get('order_number')
    # order_number = eval(order_number)
    ignore_fraud_check = request.POST.get('ignore_fraud_check', '')

    data = {}
    if ignore_fraud_check:
        data['ignore_fraud_check'] = ignore_fraud_check

    poc = PgOrderController()
    rv = poc.generate_lab_orders(order_number, data)

    try:
        origin_value = ''
        new_value = rv

        pgo = PgOrder.objects.get(order_number=order_number)
        fields = ''
        action_value = "generate_lab_orders"
        from api.controllers.tracking_controllers import tracking_operation_controller
        tc = tracking_operation_controller()
        tc.tracking(
            pgo.type, pgo.id,
            pgo.order_number,
            action_value,
            fields,
            request.user,
            origin_value,
            new_value,
            None,
            None,
        )
    except Exception as e:
        logging.debug(str(e))

    return HttpResponse(rv)


def queryActivities(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>" + str(order_number))

    obj_type = request.POST.get("obj_type")
    logging.debug("obj_type==>" + str(obj_type))
    oa = OrderActivity()
    results = oa.queryActivities(order_number, obj_type)
    return HttpResponse(results)


# def queryActivities(request):
#     order_number = request.POST.get("order_number")
#     logging.debug("order_number==>" + str(order_number))
#
#     obj_type = request.POST.get("obj_type")
#     logging.debug("obj_type==>" + str(obj_type))
#     oa = OrderActivity()
#     results = oa.queryActivities(order_number, obj_type)
#     pgOrder = PgOrder.objects.get(order_number)
#     data = {"results":results,"pgOrder":pgOrder}
#     datas = serializers.serialize('json',data)
#     return HttpResponse(datas)

@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def pgOrderItemList(request):
    logging.debug("------------------------------ begin pgOrderItemList method ------------------------------")

    currPage = request.POST.get('currPage')
    logging.debug("currPage---------------------------------------->" + str(currPage))

    curr = int(currPage)

    # curr = 1

    begin = (curr - 1) * 20
    if begin == 0:
        begin = None
        end = 20
    else:
        end = begin + 20

    logging.debug("begin---------------------------------------->" + str(begin))
    logging.debug("end---------------------------------------->" + str(end))

    queryset = PgOrderItem.objects.all().order_by('-order_number')[begin:end]

    results = serializers.serialize('json', queryset)
    return HttpResponse(results)


def pgOrderItemFilter(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>" + str(order_number))

    queryset = PgOrder.objects.filter(
        Q(order_number=order_number) | Q(order_number__endswith=order_number) | Q(customer_name=order_number) | Q(
            email=order_number))
    if len(queryset) == 0:
        return HttpResponse('False')
    else:
        results = serializers.serialize('json', queryset)
        return HttpResponse(results)


# @login_required
# @permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def shipmentServer(request, orderNumber=None, toggle=True):
    orderNumber = request.POST.get('order_number', orderNumber)
    poc = PgOrderController()
    rv = poc.shipment_server(orderNumber, toggle)
    return HttpResponse(rv)


# PgOrderItem详情
@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def pgorderItemDetail(request):
    order_number = request.POST.get('orderNumber')
    logging.debug("order_number==>" + str(order_number))

    productIndex = request.POST.get('productIndex')
    logging.debug("productIndex==>" + str(productIndex))
    response_value = []
    pgorderitems = PgOrderItem.objects.get(order_number=order_number, product_index=productIndex)
    # pgorderitem = pgorderitems[0]
    response_value.append(pgorderitems)
    orderactivities = OrderActivity.objects.filter(order_number=order_number, object_type='PORL').order_by('-id')

    if len(orderactivities) <> 0:
        for orderactivity in orderactivities:
            response_value.append(orderactivity)

    results = serializers.serialize('json', response_value)
    return HttpResponse(results)


# 跳转到order_address页面
@login_required
@permission_required('oms.OA_VIEW', login_url='/oms/forbid/')
def redirectOrderAddress(request):
    return render(request, 'order_address.html')


# 跳转到修改密码页面
@login_required
def redirectChangePassword(request):
    return render(request, 'change_password.html')


# 修改密码
@login_required
def changePassword(request):
    oldPassword = request.POST.get('oldPassword')
    logging.debug("oldPassword==>" + oldPassword)

    newPasswordOne = request.POST.get('newPasswordOne')
    logging.debug("newPasswordOne==>" + newPasswordOne)

    newPasswordTwo = request.POST.get('newPasswordTwo')
    logging.debug("newPasswordTwo==>" + newPasswordTwo)

    user = request.user
    flag = user.check_password(oldPassword)  # 验证旧密码是否正确
    logging.debug("flag==>" + str(flag))

    if flag:
        if newPasswordOne == newPasswordTwo:
            user.set_password(newPasswordOne)  # 修改成新密码
            user.save()
            return HttpResponse('Success')
        else:
            return HttpResponse('Wrong')
    else:
        return HttpResponse('Error')


# PgOrder添加备注
def addComments(request):
    is_f = request.POST.get('is_f')
    ordertype = request.POST.get('ordertype')
    if is_f == '1':
        if ordertype == 'OPOR':
            pgorder_id = request.POST.get('id')
            lbo_cmt = PgOrder.objects.only('comments').get(pk=pgorder_id)
            return HttpResponse(lbo_cmt.comments)
        elif ordertype == 'PORL':
            pgorder_id = request.POST.get('id')
            lbo_cmt = PgOrderItem.objects.only('comments').get(pk=pgorder_id)
            return HttpResponse(lbo_cmt.comments)

    try:
        order_number = request.POST.get('order_number')
        action_value = request.POST.get('action_value')
        object_entity = request.POST.get('id')
        new_value = request.POST.get('textarea')

        inner = request.POST.get('inner', 0)

        user = request.user
        user_entity = user.id
        user_name = user.username

        oa = OrderActivity()
        responseValue = oa.add_activity(ordertype, object_entity, order_number, action_value, user_entity, user_name,
                                        new_value)
        if ordertype == 'OPOR':
            po = PgOrder()
            object = po.query_by_id(order_number)
        elif ordertype == 'OLOR':
            lo = LabOrder()
            object = lo.query_by_id(order_number)

        elif ordertype == 'PORL':
            poi = PgOrderItem()
            object = poi.query_by_id(object_entity)

        if int(inner) == 0:
            object.comments = new_value
        else:
            if ordertype == 'OLOR' or ordertype == 'PORL':
                if object.comments_inner == None or object.comments_inner == '':
                    object.comments_inner = new_value
                else:
                    object.comments_inner = object.comments_inner + "；" + new_value

        object.save()
        # add by ranhy 2018-09-20
        if (int(inner) == 0):
            content = object.comments
            id_content = 'comments_' + str(object_entity)
        else:
            content = object.comments_inner
            id_content = 'comments_inner_' + str(object_entity)
        data = {
            "id_content": id_content,
            "content": content,
            "status": responseValue
        }
        result = json.dumps(data)
        return HttpResponse(result)
        # return HttpResponse(responseValue)
    except Exception as e:
        logging.debug(e.message)
        return HttpResponse(e.message)


# 修改comments
def changeComments(request):
    logging.debug("start==>changeComments")

    pk = request.POST.get("pk")
    logging.debug("pk==>" + str(pk))

    comments = request.POST.get("comments_value")
    logging.debug("comments==>" + str(comments))

    try:
        oa = OrderActivity.objects.get(pk=pk)
        oa.change_activity(comments)
        return HttpResponse("Success")
    except Exception as e:
        logging.debug("Error==>" + str(e))
        return HttpResponse(str(e))


def pgorder_status(request):
    """在pgorderitem页面点击hold,unhold,cancle"""
    order_number = request.POST.get('order_number')
    logging.debug("order_number==>%s" % order_number)

    content = request.POST.get("content")
    logging.debug("content==>%s" % content)
    tag = request.POST.get("tag")
    logging.debug("tag==>%s" % tag)
    user = request.user
    po = PgOrder()
    pgorder = po.query_by_id(order_number)
    ol = OperationLog()
    if pgorder.is_inlab:  # 判断该订单是否已经生成LabOrder 如果 is_inlab为true,则返回，否则执行else
        ol.log(pgorder.type, pgorder.id, content, "status", user)  # 添加操作记录
        return HttpResponse("该订单已生成LabOrder")
    else:
        with transaction.atomic():
            if tag == 'unhold':
                tag = ol.query_specific_log(pgorder.id, pgorder.status)  # 取出pgorder原先的状态
                if tag == False:
                    tag = 'processing'
                logging.debug("action_value==>%s" % tag)

            po.modify_status(order_number, content, tag, user)  # 修改pgorder的状态并做记录

        return HttpResponse("True")


@login_required
@permission_required('oms.PORL_CANCELLED', login_url='/oms/forbid/')
def pgorder_status_v3(request):
    '''PG Order的暂停/取消暂停/取消 操作'''
    order_number = request.POST.get('order_number')
    content = request.POST.get('content')
    reason = request.POST.get('reason', '')
    logging.debug(reason);

    rm = res_msg()
    rm_f = rm.response_dict()

    if order_number in (None, ''):
        rm.code = -1
        rm.message = 'order_number is None'
        return rm

    # pg_order操作处理类
    pgoc = pg_order_controller(order_number)
    if content == 'Cancel':
        pgoc.cancel_order(content, reason, request.user, rm)
    elif content == 'Hold':
        # Hold的处理还在上面的pgorder_status视图处理函数中
        # hold处理方法 pgoc.hold_order()
        pass
    else:
        rm.code = -1
        rm.message = 'unknown status: %s' % content
        return rm

    logging.debug(rm)
    rm_f['code'] = rm.code
    rm_f['message'] = {'status': rm.obj.status}
    return JsonResponse(rm_f)


def print_lab_order(request):
    """laborder打印"""
    lab_number = request.GET.get('p')
    data = LabOrder.objects.get(lab_number=lab_number)
    # 生成一维码
    flag = utilities.generate_code128(data)
    if flag:
        data = LabOrder.objects.get(lab_number=lab_number)
    print_date = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
    data.print_date = print_date
    return render(request, "print_order.html",
                  {"data": data, 'media_base_url': settings.SSH_MEDIA_SERVER.get('MEDIA_BASE_URL')})


'''向magento发送请求'''


def post_order_comment(order_entry_id, order_comment):
    # -------------------------------获取token-----------------------------------------------

    token = getToken()
    send_comment_url = BASE_URL + url_prefix + order_entry_id + comment_url_suffix
    http_headers = {
        'Authorization': 'Bearer %s' % token,
        'Content-Type': 'application/json'
    }
    comment = {
        'comment': order_comment.comment,
        'is_visible_on_front': order_comment.is_visible_on_front,
        'is_customer_notified': order_comment.is_customer_notified,
        'status': 'processing',
    }

    send_data = {
        'statusHistory': comment
    }

    try:
        api_response = requests.post(send_comment_url, data=json.dumps(send_data), headers=http_headers)
        return api_response.text

    except Exception as e:
        print(e)
    return None


# 不知道post_order_comment是否有用
# 新增加一个请求magento接口的函数 用于订单的review操作
def post_order_comment_v2(order_entry_id, comment, status, created_at, is_cn=0, is_vof=0):
    # 获取token
    token = getToken()
    send_comment_url = BASE_URL + url_prefix + order_entry_id + comment_url_suffix

    http_headers = {
        'Authorization': 'Bearer ' + token,
        'Content-Type': 'application/json'
    }

    review_json = {
        "comment": comment,
        "created_at": created_at.strftime('%Y-%m-%d %H:%M:%S'),
        "parent_id": order_entry_id,
        "is_customer_notified": is_cn,
        "is_visible_on_front": is_vof,
        "status": status
    }

    send_data = {
        "statusHistory": review_json
    }

    try:
        resp = requests.post(send_comment_url, data=json.dumps(send_data), headers=http_headers)
        return resp
    except Exception as e:
        print(e)
    return None


'''封装更改MG订单状态的接口'''


def order_change_status(order_entry_id, increment_id, status, state):
    # 获取token
    token = getToken()
    send_comment_url = BASE_URL + url_prefix

    http_headers = {
        'Authorization': 'Bearer ' + token,
        'Content-Type': 'application/json'
    }

    entity_json = {
        "entity_id": order_entry_id,
        "increment_id": increment_id,
        "status": status,
        "state": state
    }

    send_data = {
        "entity": entity_json
    }

    send_data = json.dumps(send_data)

    try:
        req = urllib2.Request(url=send_comment_url, data=send_data, headers=http_headers)
        res = urllib2.urlopen(req)
        resp = res.read()
        return resp
    except Exception as e:
        print(e)
    return None


'''send_message API'''


def order_comment(request):
    if request.method == 'POST':
        comment = request.POST.get('comment')
        comments = Order_comment(comment)
        is_customer_notified = request.POST.get('is_customer_notified')
        if is_customer_notified == 'on':
            comments.is_customer_notified = 1
        is_visible_on_front = request.POST.get('is_visible_on_front')
        if is_visible_on_front == 'on':
            comments.is_visible_on_front = 1
        entity = request.POST.get('entity')
        if post_order_comment(entity, comments) == 'true':
            return HttpResponse("成功")
    return HttpResponse("失败")


'''send_message模板页面'''


def send_message(request):
    comments = SendComment.objects.filter(type='SECO')
    if request.method == 'GET':
        order_number = request.GET.get("p")
        print
        order_number
        return render(request, 'send_message.html', {"order_num": order_number, "comments": comments})
    return render(request, 'send_message.html', {'comments': comments})


'''获取entity_id API'''


def get_entity(request):
    if request.method == 'POST':
        order_number = request.POST.get("order_number")
        try:
            orders = PgOrder.objects.get(order_number=order_number)
            return HttpResponse(orders.base_entity)
        except:
            return HttpResponse("error")
    return HttpResponse("error")


'''change action'''


def deal_comments(request):
    if request.method == 'POST':
        object_entity = request.POST.get("object_entity")
        order_number = request.POST.get("order_number")
        action = request.POST.get("action")
        type = request.POST.get("type")
        id = request.POST.get("id")
        order_activity = OrderActivity()
        oa = order_activity.query_by_id(id)
        oa.change_status(action)
        user = request.user
        order_activity.add_activity(type, object_entity, order_number, action, user.id, user.username)

        if type == "OPOR":
            return HttpResponse("suceess")
        else:
            return HttpResponse("成功")


# dashboard获取 pg comment 和 lab comment 的个数
def getNumber(request):
    order_activity = OrderActivity()
    filter = request.GET.get("f")
    if filter == "all":
        po_queryset = order_activity.query_by_type("OPOR")
        lo_queryset = order_activity.query_by_type("OLOR")
    elif filter == "new":
        po_queryset = order_activity.query_by_type("OPOR").filter(status='NEW')
        lo_queryset = order_activity.query_by_type("OLOR").filter(status='NEW')
    elif filter == "processing":
        po_queryset = order_activity.query_by_type("OPOR").filter(status='PROCESSING')
        lo_queryset = order_activity.query_by_type("OLOR").filter(status='PROCESSING')
    elif filter == "complete":
        po_queryset = order_activity.query_by_type("OPOR").filter(status='COMPLETE')
        lo_queryset = order_activity.query_by_type("OLOR").filter(status='COMPLETE')
    else:
        po_queryset = order_activity.query_by_type("OPOR").filter(Q(status='NEW') | Q(status='PROCESSING'))
        lo_queryset = order_activity.query_by_type("OLOR").filter(Q(status='NEW') | Q(status='PROCESSING'))

    data = {"lab": lo_queryset.count(), 'pg': po_queryset.count()}
    return HttpResponse(json.dumps(data))


# dashboard获取pg comment 和 lab comment 分页数据方法
def getPageInfo(request):
    order_activity = OrderActivity()

    if request.method == 'POST':
        types = request.POST.get("type")
        page = request.POST.get("page")
        filter = request.POST.get("filter")

        if filter == "default":
            print
            filter
            po_queryset = order_activity.query_by_type("OPOR").filter(Q(status='NEW') | Q(status='PROCESSING'))
            lo_queryset = order_activity.query_by_type("OLOR").filter(Q(status='NEW') | Q(status='PROCESSING'))
        elif filter == "new":
            po_queryset = order_activity.query_by_type("OPOR").filter(status='NEW')
            lo_queryset = order_activity.query_by_type("OLOR").filter(status='NEW')
        elif filter == "processing":
            po_queryset = order_activity.query_by_type("OPOR").filter(status='PROCESSING')
            lo_queryset = order_activity.query_by_type("OLOR").filter(status='PROCESSING')
        elif filter == "complete":
            po_queryset = order_activity.query_by_type("OPOR").filter(status='COMPLETE')
            lo_queryset = order_activity.query_by_type("OLOR").filter(status='COMPLETE')
        else:
            po_queryset = order_activity.query_by_type("OPOR")
            lo_queryset = order_activity.query_by_type("OLOR")

        po_paginator = Paginator(po_queryset, 10)
        lo_paginator = Paginator(lo_queryset, 10)

        if types == 'lab':
            try:
                lo_contacts = lo_paginator.page(int(page))
            except PageNotAnInteger:
                lo_contacts = lo_paginator.page(1)
            except EmptyPage:
                lo_contacts = lo_paginator.page(lo_contacts.num_pages)
            lo_contacts = serializers.serialize("json", lo_contacts)
            return HttpResponse(lo_contacts)
        elif types == 'pg':
            try:
                po_contacts = po_paginator.page(int(page))
            except PageNotAnInteger:
                po_contacts = po_paginator.page(1)
            except EmptyPage:
                po_contacts = po_paginator.page(po_contacts.num_pages)
            po_contacts = serializers.serialize("json", po_contacts)
            return HttpResponse(po_contacts)
    return HttpResponse(None)


def getToken():
    logging.debug("-------------------gettoken---------------------")
    get_token_url = BASE_URL + token_url
    logging.debug("gettoken_url====>" + get_token_url)
    api_response = requests.post(get_token_url, data=json.dumps(token_data), headers=token_header, timeout=10,
                                 verify=False)
    return api_response.text.replace('"', '')


# pg Order pushDate function
def pushDate(request):
    logging.debug("----------------------------------------")
    rm = response_message()

    rmjs = {}
    if request.method == 'POST':
        data = request.POST
        poc = PgOrderController()
        rm = poc.post_estimate_date(request, data)
    # 转换成列表或者数组才能转换成json
    rmjs['error_code'] = rm.error_code
    rmjs['error_message'] = rm.error_message
    return HttpResponse(json.dumps(rmjs))


# 获取pgorder
def getPgOrder(request):
    if request.method == 'POST':
        pgs = []
        order_num = request.POST.get("order_number")
        print
        order_num
        pg = PgOrder.objects.get(order_number=order_num)
        pgs.append(pg)
        pgs = serializers.serialize('json', pgs)
        return HttpResponse(pgs)
    return HttpResponse(None)


# 跳转到shipment页面
@login_required
@permission_required('oms.ACT_SHIPMENT', login_url='/oms/forbid/')
def redirectShipment(request):
    lab = LabOrder().find_order_match()
    return render(request, 'order_shipment.html', {"lab": lab})


# 发货处理
def shipmentController(request):
    try:
        if request.method == 'POST':
            lalist = request.POST.get("lalist")
            pglist = request.POST.get("pglist")
            lalist = json.loads(lalist)
            pglist = json.loads(pglist)
            carrier = request.POST.get("carrier")
            post_orderNum = request.POST.get("post_orderNum")
            remark = request.POST.get("remark")
            shipMethod = request.POST.get("shipMethod")
            la = LabOrder()
            # lalist = list(set(lalist))
            pglist = list(set(pglist))

            qty_inbox_list = []

            delivery_entity = request.POST.get("dilivery_entity", "0")
            region = request.POST.get("region", "")
            print(delivery_entity)
            print(region)
            logging.debug('shipmentController 接收参数完成')
            if not int(delivery_entity) == 0:
                from shipment.models import pre_delivery, pre_delivery_line
                pd = pre_delivery.objects.get(id=delivery_entity)

                chk = pd.check_orders_status(region)
                if not chk == '0':
                    return HttpResponse(chk)

                pglist = pd.get_orders(region)
                lalist = pd.get_lab_orders(region)
                qty_inbox_list = pd.get_qty_inbox_list(region)
                if len(lalist) == 0:
                    return HttpResponse("数据选择错误")

            if not region == "":
                if remark:
                    remark = '%s|%s-%s' % (region, delivery_entity, remark)
                else:
                    remark = region

            '''
            qty = {}
            qty['order_number'] = '202020'
            qty['qty_inbox'] = 15
    
            '''

            # qty_inbox_list.append(qty)
            # 2020.02.25 by guof. OMS-628
            # 停止Ship2系统同步数据

            '''
            pb = PushBox()
            data = pb.getJson(carrier, shipMethod, post_orderNum, remark, pglist, lalist, qty_inbox_list)
    
            requrl = PIM_BOX
            headers = {'Content-Type': 'application/json'}
    
            req = urllib2.Request(url=requrl, data=data, headers=headers)
            try:
                res = urllib2.urlopen(req)
    
            except Exception as e:
                return HttpResponse("error")
            '''

            for x in lalist:
                logging.debug('shipmentController labnumber %s start' % x)
                lab = la.query_by_id(x)
                lab.status = 'SHIPPING'
                lab.carriers = carrier
                lab.shipping_number = post_orderNum
                lab.save()
                logging.debug('shipmentController labnumber %s end' % x)

            logging.debug('shipmentController lalist end')
            for y in pglist:
                logging.debug('shipmentController order_number %s start' % y)
                pg = PgOrder.objects.get(order_number=y)
                pg.status = 'shipped'
                pg.save()
                logging.debug('shipmentController order_number %s end' % y)
            logging.debug('shipmentController pglist end')
            try:
                sh = Shipment()
                sh.carrier = carrier
                sh.remark = remark
                sh.carrierNumber = post_orderNum
                sh.save()
            except Exception, e:
                logging.debug("############################")
                logging.error(str(e))
                logging.debug("############################")

            return HttpResponse("success")

        return HttpResponse("error")

    except Exception as e:
        logging.debug("*************************")
        logging.error(str(e))
        logging.debug("*************************")
        return HttpResponse("error")


# add by ranhy stripe账单
def send_invoice(request):
    status = 0
    msg = ""
    request_info = ""
    response_info = ""
    inv_amount = 0.0
    if request.method == 'POST':
        orderNum = request.POST.get("order_number")
        json_data = request.POST.get("inv_list")
        days_until_due = request.POST.get("days_until_due")
        ticket_no = request.POST.get("ticket_no")
        invoice_type = request.POST.get("invoice_type")

        for item in eval(json_data):
            inv_amount = inv_amount + float(item['unit_amount']) * int(item['quantity'])

        try:
            http_headers = {
                'Authorization': 'Bearer %s' % getToken(),
                'Content-Type': 'application/json'
            }
            data = {
                "order_number": orderNum,
                "days_until_due": days_until_due,
                "invoice_description": "order no:" + orderNum,
                "invoice_type": invoice_type,
                "items": json.loads(json_data)
            }
            request_info = json.dumps(data)
            api_url = settings.BASE_URL + "V1/orders/invoice"

            api_response = requests.post(api_url, data=json.dumps(data), headers=http_headers, timeout=5, verify=False)
            logging.debug(api_response.text)
            response_info = api_response.text
            res_json = json.loads(api_response.text)
            pgo = PgOrder.objects.get(order_number=orderNum)
            if res_json.get("code") == 0:
                # 写备注
                nowTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                reture_id = res_json["message"]['id']
                amount_due = res_json["message"]['amount_due']
                invoic_pdf = res_json["message"]['invoice_pdf']
                # if (pgo.comments):
                #     pgo.comments += "\nSend Invoice  Success Info(Day until due:%s,Create Time:%s), Return Info:(id:%s,amount due:%s,invoice pdf:%s)" \
                #                     % (days_until_due, nowTime, reture_id, amount_due, invoic_pdf)
                #
                # else:
                #     pgo.comments = "Send Invoice Success  Info(Day until due:%s,Create Time:%s), Return Info:(id:%s,amount due:%s,invoice pdf:%s)\n" \
                #                    % (days_until_due, nowTime, reture_id, amount_due, invoic_pdf)
                if (pgo.send_invoic_info):
                    pgo.send_invoic_info += "\nrequest info:%s\n=======\nrespons info:%s" % (
                        request_info, response_info)
                else:
                    pgo.send_invoic_info = "request info:%s\n=======\nrespons info:%s\n" % (request_info, response_info)

                pgo.save()

                pg_order_inv = PgOrderInvoice()
                pg_order_inv.pg_order_entity_id = pgo.id
                pg_order_inv.order_number = orderNum
                pg_order_inv.inv_type = invoice_type
                pg_order_inv.inv_amount = inv_amount
                pg_order_inv.ticket_no = ticket_no
                pg_order_inv.invoice_id = res_json["message"]['id']

                if pg_order_inv.comments:
                    pg_order_inv.comments += "\nSend Invoice  Success Info(Day until due:%s,Create Time:%s), Return Info:(id:%s,amount due:%s,invoice pdf:%s)" \
                                             % (days_until_due, nowTime, reture_id, amount_due, invoic_pdf)
                else:
                    pg_order_inv.comments = "Send Invoice Success  Info(Day until due:%s,Create Time:%s), Return Info:(id:%s,amount due:%s,invoice pdf:%s)\n" \
                                            % (days_until_due, nowTime, reture_id, amount_due, invoic_pdf)

                pg_order_inv.save()
                return json_response(code=0, msg='successs!')
            else:
                pgo.comments += "\nSend Invoice Faild Info:%s\n" % (res_json.get("message"))
                pgo.send_invoic_info += "request info:%s\n=======\nrespons info:%s" % (request_info, response_info)
                pgo.save()
                return json_response(code=-1, msg=res_json.get("message"))
        except Exception as e:
            return json_response(code=-1, msg=traceback.format_exc())


# get shipment info
def getShipmentInfo(request):
    if request.method == 'POST':
        orderNum = request.POST.get("orderNum")

        try:
            labOrder = LabOrder().query_by_id(orderNum)
            pgOrerItems = PgOrderItem().query_by_labOrder(labOrder.base_entity)
            data = {
                "region": pgOrerItems.pg_order_entity.region,
                "ship_region": pgOrerItems.pg_order_entity.ship_region,
                "labNum": labOrder.lab_number,
                "pgNum": pgOrerItems.order_number,
                "sku": labOrder.frame,
                "comments_ship": labOrder.comments_ship
            }
            data = json.dumps(data)
            return HttpResponse(data)
        except:
            return HttpResponse(None)
    return HttpResponse(None)


# search Order Info
def searchOrderInfo(request):
    if request.method == 'POST':
        orderNum = request.POST.get("order")
        try:
            labOrder = LabOrder().query_by_id(orderNum)
            data = {
                "order_ship_region": labOrder.order_ship_region,
                "order_state": labOrder.order_state,
                "labNum": labOrder.lab_number,
                "sku": labOrder.frame,
                "status": labOrder.status
            }
            data = json.dumps(data)
            return HttpResponse(data)
        except:
            return HttpResponse(None)
    else:

        try:

            labOrder = LabOrder.objects.all()

            data = []
            for lab in labOrder:
                la = {
                    "order_ship_region": lab.order_ship_region,
                    "order_state": lab.order_state,
                    "labNum": lab.lab_number,
                    "sku": lab.frame,
                    "status": lab.status
                }
                data.append(la)

            data = json.dumps(data)
            return HttpResponse(data)
        except Exception as e:
            logging.debug(e)
            return HttpResponse(None)
    return HttpResponse(None)


# search MatchOrder Info
def searchMatchOrderInfo(request):
    labOrder = LabOrder().find_order_match()

    data = []
    for lab in labOrder:
        la = {
            "order_ship_region": lab.order_ship_region,
            "order_state": lab.order_state,
            "labNum": lab.lab_number,
            "sku": lab.frame,
            "status": lab.status
        }
        data.append(la)

    data = json.dumps(data)
    return HttpResponse(data)


# 转发到send_comment页面
def redirectComment(request):
    if request.method == 'GET':
        order_num = request.GET.get("num")
        comment = request.GET.get("comment")
        id = request.GET.get("pk")
        return render(request, "send_message.html",
                      {"order_num": order_num, "comment": comment, "redirect": True, "pk": id})
    return render(request, "send_message.html")


# change ActivitySendStatus
def changeSendStatus(request):
    if request.method == 'POST':
        pk = request.POST.get("pk")
        orderActivity = OrderActivity().query_by_id(pk)
        orderActivity.changeSendStatus()
        return HttpResponse("success")
    return HttpResponse("error")


# 2018.06.22 by guof.
# pgorder list v2
@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_list(request):
    page = request.GET.get('page', 1)
    currentPage = int(page)
    in_lab_status = request.GET.get('filter', 'a')
    logging.debug("in_lab_status==>%s" % in_lab_status)

    timedel = date_delta()  # 获取N天前的日期
    filter = {}
    filter["create_at__gte"] = timedel
    filter['is_enabled'] = True
    if in_lab_status == 'false':  # 查询没有生成laborder的pgorder
        filter["is_inlab"] = False
        results = PgOrder.objects.filter(Q(status='processing') | Q(status='holded')).filter(**filter).order_by("-id")
    else:

        if in_lab_status == "true":  # 查询生成laborder的pgorder
            filter["is_inlab"] = True

        elif in_lab_status == 'issue':  # 查询生成laborder但是没有发送地址的pgorder
            filter['is_inlab'] = True
            filter['is_shiped_api'] = False
            filter['status'] = 'processing'
        elif in_lab_status == 'lable':  # 查询有回退标签的pgorder
            filter['is_required_return_lable'] = True

        elif in_lab_status == 'reviewed':
            filter['status_control'] = 'REVIEWED'
            filter['status'] = 'processing'
            filter['is_inlab'] = False

    logging.debug(filter)
    logging.debug(in_lab_status)

    # if in_lab_status 为a 则查询所有的订单
    results = PgOrder.objects.filter(**filter).order_by("-id")

    paginator = Paginator(results, 20)  # Show 20 contacts per page

    try:
        contacts = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        contacts = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        contacts = paginator.page(paginator.num_pages)
    return render(request, 'pgorder_list_v2.html',
                  {'list': contacts, 'currentPage': currentPage, 'paginator': paginator,
                   'pgorderitem': 'New Mg Order List', 'requestUrl': '/oms/pgorder_list',
                   'filter': in_lab_status})


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_detail(request):
    number = request.GET.get("number")
    logging.debug("numbr==>" + str(number))

    currentpage = request.GET.get('page')
    logging.debug("currentpage==>%s" % currentpage)

    filter = request.GET.get('filter')
    logging.debug("filter==>%s" % filter)

    po = PgOrder()
    pgorder = po.query_by_id(number)
    pgorderformdetail = PgOrderFormDetail(instance=pgorder)
    formsets = poiFormset(request)
    formset = formsets(queryset=PgOrderItem.objects.filter(order_number=number))

    product_image_path_prefix = const.PRODUCT_IMAGE_PREPATH
    # product_image_path_prefix = 'https://test.kyoye.com/media/catalog/product'

    return render(request, 'pgorder_detail.html', {'order_number': number, 'base_entity': pgorder.base_entity,
                                                   'form': pgorderformdetail, 'formsets': formset,
                                                   'currentpage': currentpage, 'filter': json.dumps(filter),
                                                   'prd_img_pre': product_image_path_prefix
                                                   })


def redirect_pgorder_item_detail(request):
    m_order_number = request.POST.get("order_number")
    m_item_id = request.POST.get("item_id")

    logging.debug('request:')
    logging.debug(request)

    logging.debug('order_number: %s' % m_order_number)

    poc = PgOrderController()
    poi = poc.get_pgorder_item_detail(m_order_number, m_item_id)
    pois = []
    pois.append(poi)
    json_poi = serializers.serialize('json', pois)
    return HttpResponse(json_poi)


@login_required
@permission_required('oms.OPOR_SAVE', login_url='/oms/forbid/')
def redirect_pgorder_update(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>%s" % order_number)
    po = PgOrder()
    pgorder = po.query_by_id(order_number)
    logging.debug("pgorder==>%s" % pgorder)
    if request.method == 'GET':
        return render(request, 'pgorder_list.html')
    elif request.method == 'POST':
        fm_result = PgOrderFormDetail(request.POST, instance=pgorder)
        if fm_result.is_valid():
            fm_result.save()
            formsets = poiFormset(request)
            formset = formsets(queryset=PgOrderItem.objects.filter(order_number=order_number))
            return render(request, 'pgorder_detail.html',
                          {'order_number': order_number, 'form': fm_result, 'formsets': formset})


# 2018.07.11 by guof.
# pgorder list v3
@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_list_v3(request):
    _form_data = {}
    try:
        page = request.GET.get('page', 1)
        currentPage = int(page)
        order_number = request.GET.get('order_number', '')
        in_lab_status = request.GET.get('filter', 'a')
        # 获取加急订单筛选标记
        express_state = request.GET.get('express', '0')
        # 获取替换单筛选标记
        re_order_state = request.GET.get('re_order', '0')
        logging.debug("in_lab_status==>%s" % in_lab_status)

        timedel = date_delta()  # 获取N天前的日期
        filter = {}
        filter["create_at__gte"] = timedel
        filter['is_enabled'] = True

        _page_info = {}
        if order_number:
            results = PgOrder.objects.filter(
                Q(order_number=order_number) | Q(order_number__endswith=order_number) | Q(
                    customer_name=order_number) | Q(
                    email=order_number))
        else:
            if in_lab_status == 'false':  # 查询没有生成laborder的pgorder
                filter["is_inlab"] = False
                results = PgOrder.objects.filter(Q(status='processing') | Q(status='holded')).filter(**filter).order_by(
                    "-id")
            else:

                if in_lab_status == "true":  # 查询生成laborder的pgorder
                    filter["is_inlab"] = True
                if in_lab_status == "fraud":  # 查询生成laborder的pgorder
                    filter["status"] = 'fraud'

                elif in_lab_status == 'issue':  # 查询生成laborder但是没有发送地址的pgorder
                    filter['is_inlab'] = True
                    filter['is_shiped_api'] = False
                    filter['status'] = 'processing'
                elif in_lab_status == 'lable':  # 查询有回退标签的pgorder
                    filter['is_required_return_lable'] = True

                elif in_lab_status == 'reviewed':
                    filter['status_control'] = 'REVIEWED'
                    filter['status'] = 'processing'
                    filter['is_inlab'] = False
                elif in_lab_status == 'approved':
                    filter['status_control'] = 'APPROVED'
                    filter['status'] = 'processing'
                    filter['is_inlab'] = False

                elif in_lab_status == 'holded':
                    filter['status'] = 'holded'

                elif in_lab_status == 'canceled':
                    filter['status'] = 'canceled'


            if express_state == '1':
                filter['ship_direction'] = 'Express'
            if re_order_state == '1':
                filter['coupon_code__startswith'] = 'REPLACE'  # 模糊查询

            # if in_lab_status 为a 则查询所有的订单
            results = PgOrder.objects.filter(**filter).order_by("-id")

            if not in_lab_status == 'canceled':
                results = results.exclude(status='canceled').exclude(status='closed')  # 过滤取消&关闭状态的订单

        # 获取URL中除page外的其它参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        logging.debug(filter)
        logging.debug(in_lab_status)

        _page_info['total'] = results.count()
        # 分页对象，设置每页20条数据
        paginator = Paginator(results, 20)

        try:
            contacts = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            contacts = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            contacts = paginator.page(paginator.num_pages)
        return render(request, 'pgorder_list_v3.html',
                      {'page_info': _page_info, 'list': contacts, 'currentPage': currentPage,
                       'paginator': paginator,
                       'pgorderitem': 'All PgOrder List V3', 'request_url': '/oms/pgorder_list_v3/',

                       'filter': in_lab_status, 'express': express_state, 'query_string': query_string,
                       're_order': re_order_state}, )

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        _form_data['request_feature'] = 'ALL PG Orders V3'
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                          'requestUrl': reverse('pgorder_detail_v3'),
                      })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_detail_v3(request):
    _form_data = {}
    try:
        number = request.GET.get("number")
        logging.debug("numbr==>" + str(number))

        currentpage = request.GET.get('page', 0)
        logging.debug("currentpage==>%s" % currentpage)

        filter = request.GET.get('filter')
        logging.debug("filter==>%s" % filter)

        po = PgOrder()
        pgorder = po.query_by_id(number)
        if not pgorder.reorder_number == None:
            reorder_list = pgorder.reorder_number.split(';')
        else:
            reorder_list = []

        blue_glasses_list = [ item.frame for item in BlueGlasses.objects.filter(is_enabled=True)]
        pgorderformdetail = PgOrderFormDetail(instance=pgorder)
        pgorderitem_list = PgOrderItem.objects.filter(order_number=number)
        nrx_list = []
        rx_list = []
        poc = pgorder_frame_controller()
        for item in pgorderitem_list:
            if item.lens_sku in blue_glasses_list:
                res_rm = poc.get_lab_frame({"pg_frame": item.frame})
                lab_frame = res_rm.obj['lab_frame'] + 'U'
                isws = inventory_struct_warehouse.objects.filter(sku=lab_frame, warehouse_code='US-AC01')
                if len(isws) > 0:
                    isw = isws[0]
                    quantity = isw.quantity
                    diff_quantity = quantity - item.quantity
                    if diff_quantity >= 0:
                        nrx_list.append(item.id)
                    else:
                        rx_list.append(item.id)
                else:
                    rx_list.append(item.id)

            elif item.so_type == 'frame_only':
                res_rm = poc.get_lab_frame({"pg_frame": item.frame})
                lab_frame = res_rm.obj['lab_frame']
                isws = inventory_struct_warehouse.objects.filter(sku=lab_frame, warehouse_code='US-AC01')
                if len(isws) > 0:
                    isw = isws[0]
                    quantity = isw.quantity
                    diff_quantity = quantity - item.quantity
                    if diff_quantity >= 0:
                        nrx_list.append(item.id)
                    else:
                        rx_list.append(item.id)
                else:
                    rx_list.append(item.id)

            elif item.attribute_set_name not in ['Glasses', 'Goggles']:
                nrx_list.append(item.id)
            else:
                rx_list.append(item.id)
                
        formsets = poiFormset(request)
        formset = formsets(queryset=PgOrderItem.objects.filter(id__in=rx_list))

        formsets_nrx = poiFormset(request)
        nrx_pgi = PgOrderItem.objects.filter(id__in=nrx_list)
        # nrx_pgi_l = PgOrderItem.objects.filter(order_number=number,is_nonPrescription=True).filter(Q(lens_sku__in=blue_glasses_list) | Q(so_type='frame_only'))
        # nrx_pgi = nrx_pgi.union(nrx_pgi_l)
        formset_nrx = formsets_nrx(queryset=nrx_pgi)
        logging.debug('customer id: %s' % pgorder.customer_id)
        # add 2019-09-15
        warranty = 0
        # delete by guof. 2020.03.25 OMS-686
        # if (not pgorder.has_warranty):
        #     warranty_obj = get_warranty_by_order(number)
        #     logging.debug(warranty_obj)
        #     if (warranty_obj.has_key('code') and warranty_obj['code'] == 0):
        #         warranty = warranty_obj['message']['warranty']
        # end
        order_history = PgOrder.objects.filter(customer_id=pgorder.customer_id, id__lt=pgorder.id).only(
            'order_number').order_by('-id')[:5]

        product_image_path_prefix = const.PRODUCT_IMAGE_PREPATH
        # product_image_path_prefix = 'https://test.kyoye.com/media/catalog/product'

        remake_order = RemakeOrder.objects.filter(remake_order=number)
        origin_order_number = ''
        if (len(remake_order)):
            origin_order_number = remake_order[0].order_number

        # 获取Invoice信息
        pg_order_inv = PgOrderInvoice.objects.filter(order_number=number).order_by("-id")
        invoice_status = ''
        invoice_type = ''
        if len(pg_order_inv) > 0:
            invoice_status = pg_order_inv[0].status
            invoice_type = pg_order_inv[0].inv_type

        return render(request, 'pgorder_detail_v3.html', {'order_number': number, 'base_entity': pgorder.base_entity,
                                                          'priority': pgorder.priority,
                                                          'order_history': order_history, 'reorder_list': reorder_list,
                                                          'form': pgorderformdetail,
                                                          'formsets': formset, 'formsets_nrx': formset_nrx,
                                                          'currentpage': currentpage, 'filter': json.dumps(filter),
                                                          'prd_img_pre': product_image_path_prefix,
                                                          'warranty': warranty,
                                                          'origin_order_number': origin_order_number,
                                                          'pg_order_inv': pg_order_inv,
                                                          'invoice_status': invoice_status,
                                                          'invoice_type': invoice_type,
                                                          'request_url': reverse('pgorder_detail_v3'),
                                                          })
    except Exception as e:
        logging.debug('Exception: %s' % e)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e
        _form_data['request_feature'] = 'ALL PG Orders V3'
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                          'requestUrl': reverse('pgorder_detail_v3'),
                      })


from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def redirect_pgorder_get_warranty_by_order(request):
    order_number = request.POST.get('order_number', '0')
    warranty_obj = _get_warranty_by_order(order_number)
    logging.debug(warranty_obj)
    if (warranty_obj.has_key('code') and warranty_obj['code'] == 0):
        warranty = warranty_obj['message']['warranty']
    else:
        warranty = '0'
    return HttpResponse(warranty)


def _get_warranty_by_order(order_number):
    token = getToken()
    http_headers = {
        'Authorization': 'Bearer %s' % token,
        'Content-Type': 'application/json'
    }
    url = base_url + "V1/sales/mine/calc/warrant"
    try:
        poi = PgOrderItem.objects.values_list('item_id', flat=True).filter(order_number=order_number)
        data = {"order_number": order_number, "items": list(poi)}
        api_response = requests.post(url=url, data=json.dumps(data), headers=http_headers, timeout=5, verify=False)
        result = json.loads(api_response.text)
    except Exception as e:
        logging.debug(e)
        result = {"code": -1, "msg": "faild"}
    return result


@csrf_exempt
def redirect_pgorder_set_priority(request):
    pgoc = pg_order_controller()

    data = {}
    data['user_id'] = request.user.id
    data['user_name'] = request.user.username
    data['order_number'] = request.POST.get('order_number', '')
    data['priority'] = request.POST.get('priority', 0)

    rm = pgoc.set_priority(request, data)
    jrm = json.dumps(rm.__dict__)
    return HttpResponse(jrm)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_item_detail_v3(request):
    m_order_number = request.POST.get("order_number")
    m_item_id = request.POST.get("item_id")
    m_product_index = request.POST.get("product_index")

    logging.debug('order_number: %s' % m_order_number)
    logging.debug('item_id: %s' % m_item_id)

    poc = PgOrderController()
    poi = poc.get_pgorder_item_detail_v3(m_order_number, m_item_id, m_product_index)

    item = PgOrderItemFormDetailParts(instance=poi)
    prd_img_pre = const.PRODUCT_IMAGE_PREPATH
    order_image_urls = poi.order_image_urls

    # get lab order info corresponding to PG order
    lab_orders = LabOrder.objects.filter(base_entity=poi.id)

    tracking = None

    if lab_orders:
        lbo = lab_orders[lab_orders.count() - 1]
        ot = tracking_lab_order_controller()
        rm = ot.get_order_history_en(lbo.lab_number)
        tracking = rm.obj
        lbo.tracking_log = tracking

    assemble_height_tuple = tuple(range(-10, 11))

    # 重做单不能进行重做单 ，ranhy，必须在原单处理
    ro = RemakeOrder.objects.filter(remake_order=m_order_number)
    is_remake = False
    if (len(ro) > 0):
        is_remake = True

    return render(
        request, "pgorder_item_detail_v3.pspf.html", {
            'obj_type': poi.type,
            'obj_id': poi.id,
            'instance': poi,
            'item': item,
            'is_remake': is_remake,
            'prd_img_pre': prd_img_pre,
            'order_image_urls': order_image_urls,
            'lab_orders': lab_orders,
            'tracking_en': tracking,
        }
    )


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def edit_process(request):
    # ##############################################
    # 0: item未生成工厂订单 或 对应工厂订单有未到镜片生产的
    # 1: item已生成工厂订单 且 所有工厂订单都已过镜片生产
    # -1: 异常 item_id 没传过来
    # ##############################################

    # 要记录未更新的订单  反馈到前端
    rm = {'code': '0', 'message': '', 'unchange_list': []}
    # from_info
    form_info = request.POST.get('form_info')
    process_dict = json.loads(form_info)
    logging.debug("===================")
    logging.debug(process_dict)
    itm_id = process_dict.pop('item_id')
    order_number = process_dict.pop('order_number')
    product_index = process_dict.pop('product_index')
    logging.debug(process_dict)

    # 获取PgOrderItem 通过item_id
    if itm_id != '' or itm_id != None:
        cur_items = PgOrderItem.objects.filter(item_id=itm_id,order_number=order_number,product_index=product_index)
        item = cur_items[0]
    else:
        rm['code'] = '-1'
        rm['message'] = 'item error.'
        return JsonResponse(rm)

    # 如果PgOrder没生成工厂订单
    if not item.pg_order_entity.is_inlab:
        item.__dict__.update(**process_dict)
        item.save()
    else:
        # 生成了工厂订单
        # 查找相对应的Lab Order 限定状态范围
        status_tuple = (None, '', 'REQUEST_NOTES', 'FRAME_OUTBOUND')
        cur_lbos = LabOrder.objects.filter(base_entity=item.id)
        cur_chg_lbos = cur_lbos.filter(status__in=status_tuple)
        # 如果queryset不为空
        if cur_chg_lbos.exists():
            cur_chg_lbos.update(**process_dict)
            item.__dict__.update(**process_dict)
            item.save()

            # 查询所有没有更改状态的lab order反馈到前端
            for lbo in cur_lbos:
                if not (lbo.status in status_tuple):
                    rm['unchange_list'].append(lbo.lab_number)

        else:
            # 如果所有对应工厂订单都已镜片生产
            rm['code'] = '1'
            rm['message'] = '所有对应工厂订单镜片都已生产 无法更改'
            return JsonResponse(rm)

    return JsonResponse(rm)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def edit_lab_comments(request):
    rm = {'code': '0', 'message': ''}
    is_ship = request.POST.get("ship", '0');  # 1为发货备注
    op_type = request.POST.get("op_type");
    if op_type == '0':
        m_order_number = request.POST.get("order_number")
        m_item_id = request.POST.get("item_id")
        product_index = request.POST.get("product_index")
        try:
            poi = PgOrderItem.objects.get(order_number=m_order_number, item_id=m_item_id, product_index=product_index)
            lab_orders = LabOrder.objects.filter(base_entity=poi.id)
            return render(request, 'lab_comments_edit.html', {
                'lbos': lab_orders,
                'is_ship': is_ship
            })
        except:
            return HttpResponse('Pg Order Item not found.')

    if op_type == '1':
        lbo = None
        m_lbo_id = request.POST.get("lbo_id")
        m_comments = request.POST.get("comments")
        try:
            lbo = LabOrder.objects.get(pk=m_lbo_id)
        except Exception as e:
            logging.debug('error==========>%s' % e)
            rm['code'] = '-1'
            rm['message'] = 'LabOrder not found.'
            return JsonResponse(rm)

        if not lbo == None:
            if is_ship == '1':
                lbo.comments_ship = m_comments
            else:
                lbo.comments = m_comments
            lbo.save()

            order_activity = OrderActivity()
            order_activity.add_activity(
                'OPOR', m_lbo_id, lbo.order_number, 'EDIT COMMENT', request.user.id, request.user.username,
                'EDIT SHIP COMMENTS' if is_ship == '1' else 'EDIT LAB COMMENTS'
            )

            return JsonResponse(rm)

    rm['code'] = '-2'
    return JsonResponse(rm)


# prescripion edit view
@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_prescripion_update_v3(request):
    clk_flag = request.POST.get('clk_flag', '0')
    m_order_number = request.POST.get('order_number')
    m_item_id = request.POST.get('item_id')
    product_index = request.POST.get('product_index')
    rm = res_msg()

    try:
        poi = PgOrderItem.objects.get(order_number=m_order_number, item_id=m_item_id, product_index=product_index)
    except Exception as e:
        rm.capture_execption(e)
        return JsonResponse({'code': rm.code, 'message': rm.message})

    if clk_flag == '1':
        prescripion_num = generate_prescripion_tuple()
        item = PgOrderItemFormDetailParts(instance=poi)
        return render(request, 'prescripion_detail.html', {
            'instance': poi,
            'item': item,
            'prescripion_num': prescripion_num,
        })

    form_data = request.POST.get("form_data")
    form_json = json.loads(form_data)
    lab_orders = LabOrder.objects.filter(base_entity=poi.id)

    if len(lab_orders) > 1:
        return JsonResponse({'flag': '0', 'content': '已有多个工厂订单 需手动修改'})

    if len(form_json.get('frame', '')) < 8:
        return JsonResponse({'flag': '0', 'content': '镜架信息不正确'})

    # 镜架SKU
    poc = pgorder_frame_controller()
    res_rm = poc.get_lab_frame({"pg_frame": form_json.get('frame')})
    frame_sku = res_rm.obj['lab_frame']
    # frame_lens = len(form_json.get('frame'))
    # if frame_lens == 9:  # 镜架 太阳镜
    # frame_sku = form_json.get('frame')[1:-1]
    # elif frame_lens == 8:
    # frame_sku = form_json.get('frame')[1:]

    # 找到新镜架图片
    pf_entity = product_frame.objects.filter(sku=frame_sku)
    if pf_entity.count() == 0:
        return JsonResponse({'flag': '0', 'content': '镜架图片未找到'})
    frame_img = pf_entity[0].image
    # 镜架图片加入字典
    form_json['image'] = frame_img
    form_json['thumbnail'] = frame_img

    with transaction.atomic():
        # wj 2019.06.14 去掉网站库存同步

        # Pg Order Item数据不做处理
        poi.__dict__.update(**form_json)
        poi.save()

        # 如果有工厂订单
        if not len(lab_orders) == 0:

            lbo = lab_orders[0]
            if lbo.status not in ('', None, 'REQUEST_NOTES', 'FRAME_OUTBOUND'):
                return JsonResponse({"flag": "0", "content": "当前状态【%s】不允许修改验光单数据" % lbo.get_status_display()})

            # cyl为正数时做特殊处理 准备lab_order lens_order数据
            if float(form_json.get('od_cyl')) > 0:
                pres = PrescriptionSwap()
                pres = poi.convert_cyl_plus(form_json.get('od_sph'), form_json.get('od_cyl'), form_json.get('od_axis'))
                form_json['od_sph'] = pres.sph
                form_json['od_cyl'] = pres.cyl
                form_json['od_axis'] = pres.axis

            if float(form_json.get('os_cyl')) > 0:
                pres = PrescriptionSwap()
                pres = poi.convert_cyl_plus(form_json.get('os_sph'), form_json.get('os_cyl'), form_json.get('os_axis'))
                form_json['os_sph'] = pres.sph
                form_json['os_cyl'] = pres.cyl
                form_json['os_axis'] = pres.axis

            form_json['frame'] = frame_sku

            lbo.__dict__.update(**form_json)
            lbo.save()

            # 处理lens_order数据
            lods = lens_order.objects.filter(lab_number=lbo.lab_number)
            for lo in lods:
                if lo.rl_identification == 'R':
                    lo.sph = form_json.get('od_sph')
                    lo.cyl = form_json.get('od_cyl')
                    lo.add = form_json.get('od_add')
                    if form_json.get('is_singgle_pd') == 'True':
                        lo.pd = math.floor(float(form_json.get('pd'))) / 2
                    else:
                        lo.pd = form_json.get('od_pd')
                    lo.prism = form_json.get('od_prism')
                    lo.base = form_json.get('od_base')
                    lo.prism1 = form_json.get('od_prism1')
                    lo.base1 = form_json.get('od_base1')
                    lo.axis = form_json.get('od_axis')

                elif lo.rl_identification == 'L':
                    lo.sph = form_json.get('os_sph')
                    lo.cyl = form_json.get('os_cyl')
                    lo.add = form_json.get('os_add')
                    if form_json.get('is_singgle_pd') == 'True':
                        lo.pd = math.floor(float(form_json.get('pd'))) / 2
                    else:
                        lo.pd = form_json.get('os_pd')
                    lo.prism = form_json.get('os_prism')
                    lo.base = form_json.get('os_base')
                    lo.prism1 = form_json.get('os_prism1')
                    lo.base1 = form_json.get('os_base1')
                    lo.axis = form_json.get('os_axis')

                lo.save()

        order_activity = OrderActivity()
        order_activity.add_activity('OPOR', poi.pg_order_entity.id, m_order_number, 'CHANGE_PRESCRIPION',
                                    request.user.id, request.user.username, 'prescripion(RX) has changed...')

    # generate model
    try:
        logging.debug('sssssssssssssssssssssssssssssssss')
        poi = PgOrderItem.objects.get(order_number=m_order_number, item_id=m_item_id)
        item = PgOrderItemFormDetailParts(instance=poi)
        prd_img_pre = const.PRODUCT_IMAGE_PREPATH
        order_image_urls = poi.order_image_urls
    except Exception as e:
        logging.debug('errerrerrerrerererererererererererererererererer')
        rm.capture_execption(e)
        return JsonResponse({'code': rm.code, 'message': rm.message})

    return render(request, "pgorder_item_detail_v3.pspf.html", {
        'obj_type': poi.type,
        'obj_id': poi.id,
        'instance': poi,
        'item': item,
        'prd_img_pre': prd_img_pre,
        'order_image_urls': order_image_urls,
        'lab_orders': lab_orders,
        'handling_choices': PgOrderItem.HANDLING_CHOICES
    })


@login_required
@permission_required('oms.OPOR_SAVE', login_url='/oms/forbid/')
def redirect_pgorder_update_v3(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>%s" % order_number)
    po = PgOrder()
    pgorder = po.query_by_id(order_number)
    logging.debug("pgorder==>%s" % pgorder)
    if request.method == 'GET':
        return render(request, 'pgorder_list_v3.html')
    elif request.method == 'POST':
        fm_result = PgOrderFormDetail(request.POST, instance=pgorder)
        if fm_result.is_valid():
            fm_result.save()
            formsets = poiFormset(request)
            formset = formsets(queryset=PgOrderItem.objects.filter(order_number=order_number))
            return render(request, 'pgorder_detail_v3.html',
                          {'order_number': order_number, 'form': fm_result, 'formsets': formset})


# 已废弃
@login_required
@permission_required('oms.OPOR_SAVE', login_url='/oms/forbid/')
def redirect_pgorder_item_update_v3(request):
    order_number = request.POST.get("order_number")
    logging.debug("order_number==>%s" % order_number)
    item_id = request.POST.get("item_id")

    poi = PgOrderItem.objects.get(order_number=order_number, item_id=item_id)

    if request.method == 'GET':
        return render(request, 'pgorder_list_v3.html')
    elif request.method == 'POST':
        form = PgOrderItemFormDetailParts(request.POST, instance=poi)
        if form.is_valid():
            form.save()
            return render(request, "pgorder_item_detail_v3.pspf.html",
                          {
                              'item': form,
                              'prd_img_pre': None,
                              'order_image_urls': None,
                          })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes(request):
    page_info = {}

    form_data = None
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filter', 'new')
    priority = request.GET.get('priority', '4')
    warehouse = request.GET.get('warehouse', '')
    ship = request.GET.get('ship', 'ALL')
    overdue = request.GET.get('overdue', 'ALL')
    items = []
    paginator = None

    try:
        filter = {}

        # filter["status"] = ''
        # filter["status"] = 'PRINT_DATE'
        if laborder_filter != 'new':
            filter['vendor'] = laborder_filter
        if ship != '' and ship != 'ALL':
            filter['act_ship_direction'] = ship

        lbos = []
        lbos_c = []
        lbos_k = []
        lbos_full = []
        lbos_three = []
        lbos_two = []
        total_lab_five = 0
        total_lab_forth = 0
        total_lab_three = 0
        total_lab_two = 0

        vendors_choice_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            if int(vc.key) > 0:
                vendors_choice_list.append(vc)

        filter['priority'] = priority

        if laborder_filter == 'new':
            lbos_full = LabOrder.objects.filter(is_enabled=True) \
                .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('vendor', 'frame')
        else:
            lbos_full = LabOrder.objects.filter(is_enabled=True) \
                .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('frame')

        iswc = inventory_struct_warehouse_controller()

        for lbo in lbos_full:
            if not laborder_filter == 'new':
                location = iswc.get_location(lbo.frame, warehouse)
                lbo.location = location
            lbo.overdue_day = (timezone.now() - lbo.create_at).days
            if lbo.overdue_day >= 5:
                total_lab_five = total_lab_five + 1

            if lbo.overdue_day == 4:
                total_lab_forth = total_lab_forth + 1

            if lbo.overdue_day == 3:
                total_lab_three = total_lab_three + 1

            if lbo.overdue_day <= 2:
                total_lab_two = total_lab_two + 1

            if lbo.lens_type == 'C':
                lbos_c.append(lbo)
            else:
                lbos_k.append(lbo)

        lbos.extend(lbos_c)
        lbos.extend(lbos_k)

        if overdue == 'ALL' or overdue == '':
            items = lbos
        elif overdue == 'THREE':
            for lbo in lbos:
                print(lbo.overdue_day)
                if lbo.overdue_day >= 3:
                    lbos_three.append(lbo)
            items = lbos_three
        elif overdue == 'TWO':
            for lbo in lbos:
                if lbo.overdue_day <= 2:
                    lbos_two.append(lbo)
            items = lbos_two
        else:
            items = lbos

        count = len(items)
        if count > 0:
            page_info['total'] = count

        paginator = Paginator(items, 100)  # Show 20 contacts per page
        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)

        return render(request, "laborder_request_notes.html", {
            'page_info': page_info,
            'form_data': form_data,
            'list': items,
            'currentPage': currentPage, 'paginator': paginator,
            'requestUrl': '/oms/laborder_request_notes/',
            'filter': laborder_filter,
            'warehouse': warehouse,
            'priority': priority,
            'ship': ship,
            'vendors_choices': vendors_choice_list,
            'total_lab_five': total_lab_five,
            'total_lab_forth': total_lab_forth,
            'total_lab_three': total_lab_three,
            'total_lab_two': total_lab_two,
            'overdue': overdue
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request_notes.html", {
            'page_info': page_info,
            'form_data': form_data,
            'list': items,
            'currentPage': currentPage, 'paginator': paginator,
            'requestUrl': '/oms/laborder_request_notes/',
            'filter': laborder_filter,
            'priority': priority,
            'warehouse': warehouse,
        })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_archives(request):
    page_info = {}

    form_data = None
    page = request.GET.get('page', 1)
    currentPage = int(page)
    vendor = request.GET.get('vendor', 'all')

    items = []

    paginator = None
    vendors_choice_list = []
    for vcl in LabOrder.VENDOR_CHOICES:
        vc = status_choice()
        vc.key = vcl[0]
        vc.value = vcl[1]
        vendors_choice_list.append(vc)
    target_day = const.date_delta()
    logging.debug(target_day)
    target_day = const.date_delta()
    logging.debug(target_day)
    # 获取URL中除page外的其它参数
    query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
    if query_string:
        query_string = '&' + query_string
    try:

        items = laborder_request_notes.objects.filter(created_at__gte=target_day).order_by('-id')

        if not vendor == 'all':
            items = items.filter(vendor=vendor)
        page_info['total'] = items.count()

        paginator = Paginator(items, 30)  # Show 20 contacts per page

        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)

        return render(request, "laborder_request_notes_archives.html",
                      {
                          'page_info': page_info,
                          'form_data': form_data,
                          'list': items,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/laborder_request_notes_archives/',
                          'query_string': query_string,
                          'vendors_choices': vendors_choice_list,
                          'vendor': vendor
                      })
    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        return render(request, "laborder_request_notes_archives.html",
                      {
                          'page_info': page_info,
                          'form_data': form_data,
                          'list': items,
                          'currentPage': currentPage, 'paginator': paginator,
                          'requestUrl': '/oms/laborder_request_notes_archives/',
                          'vendor': vendor
                      })


# qulee1118
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_frame_delivery_list(request):
    lbos = []
    time_now = request.POST.get("time")
    if time_now == None or time_now == "":
        time_now = timezone.now().date()
    else:
        int_time_arr = []
        time_arr = time_now.split('-')
        for tm in time_arr:
            int_time_arr.append(int(tm))
        time_now = datetime.date(int_time_arr[0], int_time_arr[1], int_time_arr[2])

    time_tom = time_now + datetime.timedelta(days=1)
    time_cur = time_now.strftime("%Y-%m-%d")
    time_next = time_tom.strftime("%Y-%m-%d")
    lab_num_list = []

    invds = inventory_delivery.objects.filter(user_id=request.user.id).filter(created_at__range=[time_cur, time_next])
    if len(invds) != 0:
        for invd in invds:
            lab_num_list.append(invd.comments)
        lbos = LabOrder.objects.filter(lab_number__in=lab_num_list).order_by('-id')
        for lbo in lbos:
            logging.debug(lbo.lab_number)

    if request.method == "POST":
        url = "frame_delivery_list_part.html"
    else:
        url = "frame_delivery_list.html"

    return render(request, url, {
        "lbos": lbos
    })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_frame_delivery_print(request):
    id_list = request.POST.get("id_list")
    date = request.POST.get("date")
    if date == None or date == "":
        date = timezone.now().date()
    lbos = []
    id_arr = id_list.split(',')
    logging.debug("====================%s" % len(id_arr))
    try:
        for id in id_arr:
            logging.debug("id================%s" % id)
            lbos.append(LabOrder.objects.get(pk=id))

    except Exception as e:
        return HttpResponse("请勾选要打印的订单")

    return render(request, "frame_delivery_list_print.html", {
        "lbos": lbos,
        "date": date
    })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_detail(request):
    page_info = {}

    form_data = {}
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filters', 'all')
    id = request.GET.get('id', '')
    form_data['id'] = id
    try:
        id = int(id)
    except:
        return HttpResponse('订单编号错误!')

    items = []
    paginator = None

    try:
        lrn_entity = laborder_request_notes.objects.get(id=id)
        form_data['lrn_entity'] = lrn_entity
        filter = {}

        filter["status"] = ''
        # filter["status"] = 'PRINT_DATE'

        items = laborder_request_notes_line.objects.filter(lrn_id=id).order_by('laborder_entity__vendor', 'location')

        count = items.count()
        if count > 0:
            page_info['total'] = count

        paginator = Paginator(items, 20)  # Show 20 contacts per page
        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)

        return render(request, "laborder_request_notes_detail.html",
                      {
                          'page_info': page_info,
                          'form_data': form_data,
                          'list': items,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/laborder_request_notes_detail/',
                          'filter': laborder_filter,
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request_notes_detail.html",
                      {
                          'page_info': page_info,
                          'form_data': form_data,
                          'list': items,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/laborder_request_notes_detail/',
                          'filter': laborder_filter,
                      })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_change_vendor(request):
    form_data = {}
    try:
        data = {}
        data['id'] = request.POST.get('id', 0)
        data['origin_vendor'] = request.POST.get('origin_vendor', '')
        data['new_vendor'] = request.POST.get('new_vendor', '')

        from oms.controllers.lab_order_request_notes_controllers import laborder_request_notes_controller
        lrnc = laborder_request_notes_controller()
        rm = lrnc.change_vendor(request, data)
        rm_dict = dict_helper.convert_to_dict(rm)
        r_json = json.dumps(rm_dict)
        logging.debug(r_json)
        return HttpResponse(r_json)

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse(str(e))


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_create(request):
    logging.debug('method: %s' % request.method)
    if request.method == 'GET':
        logging.debug('redirect to print')
        return render(request, "laborder_request_notes_print.html")

    laborder_filter = request.POST.get('filter', 'new')
    warehouse = request.POST.get('warehouse', '')
    ship = request.POST.get('ship', 'ALL')
    overdue = request.POST.get('overdue', 'ALL')
    entities = request.POST.get('entities', '')
    logging.debug('no parameters ....')

    lines = request.POST.get('lines', 30)
    items = []
    print (request.method)
    print(entities)
    try:
        if entities:
            lbos = []
            lbos_c = []
            lbos_k = []
            lbos_list = []
            arry_entities = entities.split(',')
            if laborder_filter == 'new':
                lbos_init = LabOrder.objects.filter(id__in=arry_entities).order_by('-id')[:1]
                lbos_full = LabOrder.objects.filter(id__in=arry_entities).order_by('vendor', 'frame')

            else:
                lbos_init = LabOrder.objects.filter(id__in=arry_entities).order_by('-id')[:1]
                lbos_full = LabOrder.objects.filter(id__in=arry_entities).order_by('frame')
            for lbo in lbos_full:
                if lbo.lens_type == 'C':
                    lbos_c.append(lbo)
                else:
                    lbos_k.append(lbo)

            lbos_list.extend(lbos_c)
            lbos_list.extend(lbos_k)
            lbos = lbos_list
        else:
            filter = {}

            # filter["status"] = ''
            # filter["status"] = 'PRINT_DATE'
            if laborder_filter != 'new':
                filter['vendor'] = laborder_filter
            if ship != '' and ship != 'ALL':
                filter['act_ship_direction'] = ship
            lbos = []
            lbos_c = []
            lbos_k = []
            lbos_two = []
            lbos_three = []
            lbos_list = []
            lbos_init = []
            lbos_full = []

            if warehouse == '':
                return HttpResponse("请先指定仓库!")

            if laborder_filter == 'new':
                lbos_init = LabOrder.objects.filter(is_enabled=True) \
                                .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('-id')[:1]
                lbos_full = LabOrder.objects.filter(is_enabled=True) \
                    .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('vendor', 'frame')

            else:
                lbos_init = LabOrder.objects.filter(is_enabled=True) \
                                .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('-id')[:1]
                lbos_full = LabOrder.objects.filter(is_enabled=True) \
                    .filter(Q(status='') | Q(status=None)).filter(**filter).order_by('frame')

            # Create Request Notes

            for lbo in lbos_full:
                if lbo.lens_type == 'C':
                    lbos_c.append(lbo)
                else:
                    lbos_k.append(lbo)

            lbos_list.extend(lbos_c)
            lbos_list.extend(lbos_k)

            if overdue == 'ALL' or overdue == '':
                lbos = lbos_list
            elif overdue == 'THREE':
                for lbo in lbos_list:
                    if (timezone.now() - lbo.create_at).days >= 3:
                        lbos_three.append(lbo)
                lbos = lbos_three
            elif overdue == 'TWO':
                for lbo in lbos:
                    if (timezone.now() - lbo.create_at).days <= 2:
                        lbos_two.append(lbo)
                lbos = lbos_two
            else:
                lbos = lbos_list

        lrn_id = 0
        for lb in lbos:
            lrnls = laborder_request_notes_line.objects.filter(lab_number=lb.lab_number)
            if lrnls.count() > 0:
                return HttpResponse("订单: [%s] 已经生成过 [出库申请单], 请先暂停." % lb.lab_number)

        try:
            with transaction.atomic():
                lrn = laborder_request_notes()
                count = len(lbos)
                lbo = lbos_init[0]
                lrn.lab_number = lbo.lab_number
                lrn.warehouse_code = warehouse
                lrn.laborder_id = lbo.id
                lrn.count = count

                if not laborder_filter == 'new':
                    lrn.vendor = laborder_filter
                else:
                    lrn.vendor = lbo.vendor
                # lrn.user_id = request.user.id
                # lrn.user_name = request.user_name
                lrn.save()
                lrn_id = lrn.id
                iswc = inventory_struct_warehouse_controller()

                index = 0
                for lbo in lbos:
                    index += 1
                    lrn_lines = laborder_request_notes_line()
                    lrn_lines.lrn = lrn
                    lrn_lines.laborder_id = lbo.id
                    lrn_lines.lab_number = lbo.lab_number
                    lrn_lines.index = index
                    lrn_lines.frame = lbo.frame
                    location = iswc.get_location(lbo.frame, warehouse)
                    lrn_lines.location = location
                    lrn_lines.warehouse_code = warehouse
                    lrn_lines.laborder_entity = lbo
                    lrn_lines.lens_type = lbo.lens_type
                    lrn_lines.quantity = lbo.quantity
                    lrn_lines.order_date = lbo.order_date
                    lrn_lines.order_created_date = lbo.create_at
                    lrn_lines.vendor = lbo.vendor
                    lrn_lines.save()

                    # lbo.status = 'REQUEST_NOTES'
                    # lbo.save()

                    # add log
                    tloc = tracking_lab_order_controller()
                    tloc.tracking(lbo, request.user, "REQUEST_NOTES", "出库申请")

                    # lbo_new = LabOrder.objects.get(id=lbo.id)
                    # lbo_new.status = 'REQUEST_NOTES'
                    # lbo_new.save()
                    # 改成事务外SQL更新
            with connections['default'].cursor() as cursor:
                try:
                    sql = '''
                        UPDATE oms_laborder SET STATUS='REQUEST_NOTES' 
                        WHERE id IN (SELECT laborder_id FROM oms_laborder_request_notes_line WHERE lrn_id = %s)
                    ''' % lrn_id
                    cursor.execute(sql)
                except Exception as e:
                    return HttpResponse('lab_order状态更新失败： ' + e)
        except Exception as e:
            return HttpResponse('系统认为你可能已经创建了今天的出库申请单，如果需要重新打印，请转至 Request Notes Archives - ' + e)

        # Print Start  ========================================

        items = laborder_request_notes_line.objects.filter(lrn__id=lrn.id).order_by('location')

        count = len(items)
        lines = int(lines)

        copies = count // lines
        remainder = count % lines
        if remainder > 0:
            copies += 1

        index = 0
        items_n = []

        for i in range(copies):
            pitms = ParentItems()
            pitms.index = i + 1
            pitms.count = copies
            pitms.created_at = datetime.date.today()
            pitms.items = []
            for j in range(lines):
                itm = Items()
                itm.index = index

                obj = Items()
                obj.index = index + 1
                obj.obj = items[index]
                itm.left = obj
                pitms.items.append(itm)

                index += 1
                logging.debug('index: %s' % index)
                if index == count:
                    break

            items_n.append(pitms)
            if index == count:
                break

        return render(request, "laborder_request_notes_print_part.html", {
            'list': items_n,
            'requestUrl': '/oms/laborder_request_notes_print_part/',
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request_notes_print_part.html", {
            'list': items_n,
            'requestUrl': '/oms/laborder_request_notes_print_part/',
        })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_generate_barcode(request):
    page_info = {}
    form_data = None
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filter', 'all')

    id = request.GET.get('id', '')
    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    lines = request.GET.get('lines', 30)

    items = []
    paginator = None

    try:
        filter = {}

        lrn = laborder_request_notes.objects.get(id=id)
        # lbos = lrn.laborder_entities

        # Print Start
        lbos = laborder_request_notes_line.objects.filter(lrn__id=lrn.id).order_by('location')

        for lbo in lbos:
            items.append('%s|B%d' % (lbo.lab_number, lbo.laborder_id))

        logging.debug('items count: %s' % len(items))

        response = HttpResponse(content_type='text')
        file_name = 'laborder_numbers_barcode'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.txt'

        count = len(items)
        index = 0

        for item in items:
            index += 1
            if index < count:
                response.write(item + '\n')
            else:
                response.write(item)
        return response

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse('生成条码标签清单遇到异常[ %s ], 请暂时手动生成，并联系系统支持....' % e.message)


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_generate_barcode_special(request):
    page_info = {}
    form_data = None
    items = []
    paginator = None
    lbos = None
    # 获取参数
    entities = request.GET.get('entities', '')
    filter = request.GET.get('filter', '')
    status = request.GET.get('status', '')
    vendor = request.GET.get('vendor', '')
    ltype = request.GET.get('ltype', '')
    sorted = request.GET.get('sorted', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    try:
        if entities == '':
            # 筛选条件处理
            _filter = {}
            timedel = date_delta()
            timedel_week = date_delta_week()
            timedel_month = date_delta_month()
            lbos = LabOrder.objects.filter(
                is_enabled=True,
                create_at__gte=timedel)
            if filter == 'week':
                lbos = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel_week)

            if filter == 'month':
                lbos = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel_month)

            if filter == 'all':
                lbos = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel)

            lbos = lbos.order_by(sorted)
            if filter == 'new':
                lbos = lbos.filter(**_filter).filter(Q(status='') | Q(status=None))
            else:
                if not vendor == 'all':
                    _filter['vendor'] = vendor

                if not status == 'all':
                    _filter['status'] = status

                lbos = lbos.filter(**_filter)
            if sorted == 'set_time':
                _filter['is_enabled'] = True
                _filter['create_at__gte'] = timedel
                _filter['set_time__lt'] = 0

                _filter_exceed_dimentions = {}
                _filter_exceed_dimentions['is_enabled'] = True
                _filter_exceed_dimentions['create_at__gte'] = timedel
                _filter_exceed_dimentions['set_time__lt'] = 0
                items_origin = LabOrder.objects.filter(**_filter_exceed_dimentions)
                from report.models import exceed_laborders
                el = exceed_laborders()

                lbos = LabOrder.objects.filter(**_filter) \
                    .filter(
                    ~Q(status='SHIPPING'),
                    ~Q(status='CANCELLED')
                )
                lbos.order_by('set_time')

            if not ltype == 'all':
                if ltype == 'c':
                    lbos = lbos.filter(lab_number__contains='C')
                elif ltype == 'r':
                    lbos = lbos.filter(lab_number__contains='R')
                elif ltype == 'z':
                    lbos = lbos.filter(lab_number__contains='Z')
                else:
                    lbos = lbos.filter(lab_number__contains='B')
            # 按时间区间筛选
            if not start_date == '':
                lbos = lbos.filter(create_at__range=(start_date, end_date))
            # 按镜架排序
            lbos = lbos.order_by('frame')
        else:
            ids = entities.split(',')
            filter = {}
            lbos = LabOrder.objects.filter(id__in=ids).order_by('frame')
        # Print Start
        for lbo in lbos:
            items.append('%s|B%d' % (lbo.lab_number, lbo.id))

        logging.debug('items count: %s' % len(items))
        response = HttpResponse(content_type='text')
        file_name = 'laborder_numbers_barcode'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.txt'

        count = len(items)
        index = 0

        for item in items:
            index += 1
            if index < count:
                response.write(item + '\n')
            else:
                response.write(item)
        return response

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse('生成条码标签清单遇到异常[ %s ], 请暂时手动生成，并联系系统支持....' % e.message)


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_generate_csv(request):
    page_info = {}
    form_data = None
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filter', 'all')

    id = request.GET.get('id', '')
    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    lines = request.GET.get('lines', 30)

    items = []
    paginator = None

    try:
        filter = {}

        lrn = laborder_request_notes.objects.get(id=id)
        lbos = lrn.laborder_entities_frame_outbound

        # Print Start

        import csv, codecs

        for lbo in lbos:
            items.append(lbo.lab_number)

        logging.debug('items count: %s' % len(items))

        response = HttpResponse(content_type='text/csv')
        file_name = 'laborder_numbers_csv'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'

        response.write(codecs.BOM_UTF8)

        writer = csv.writer(response)

        writer.writerow(['订单信息'])
        writer.writerow([
            '', '', '', '', '', '', '', '', '', '',
            '',
            '验光单', '', '', 'PD', '', 'ADD', 'prism-H', 'base-H', 'prism-V', 'base-V'
        ])

        writer.writerow([
            '序号', '配送', '下达日期', '订单号', '镜架尺寸', '框型', '名称', '设计', '涂层名称', '染色名称',
            'R/L', 'SPH', 'CYL', 'AXIS', '单眼', '双眼', '', '', '', '', '', '备注'
        ])

        count = len(items)
        index = 0

        for lbo in lbos:
            index += 1
            lbo = __formate_prescripion(lbo)

            writer.writerow([
                index, '库房自提', lbo.order_date, lbo.lab_number, lbo.size, '', lbo.act_lens_name, lbo.pal_design_name,
                lbo.coating_name,
                lbo.tint_name,
                'OD-右眼',
                lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_pd, lbo.pd, lbo.od_add, lbo.od_prism, lbo.od_base,
                lbo.od_prism1, lbo.od_base1, lbo.comments
            ])

            writer.writerow([
                '', '', '', '', '', '', '', '', '', '', 'OS-左眼',
                lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_pd, '', lbo.os_add, lbo.os_prism, lbo.os_base,
                lbo.os_prism1, lbo.os_base1, ''
            ])

        return response

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse('生成采购订单遇到异常[ %s ], 请暂时手动生成，并联系系统支持....' % e.message)


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_generate_csv_wx(request):
    page_info = {}
    form_data = None
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filter', 'all')

    id = request.GET.get('id', '')
    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    lines = request.GET.get('lines', 30)

    items = []
    paginator = None

    try:
        filter = {}

        lrn = laborder_request_notes.objects.get(id=id)
        lbos = lrn.laborder_entities_frame_outbound

        # Print Start

        import csv, codecs

        for lbo in lbos:
            items.append(lbo.lab_number)

        logging.debug('items count: %s' % len(items))

        response = HttpResponse(content_type='text/csv')
        file_name = 'laborder_numbers_csv'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'

        response.write(codecs.BOM_UTF8)

        writer = csv.writer(response)

        writer.writerow([
            '', '', '', '', 'Frame',
            '', '', '', '', '', '', '', '', 'Right Eye',
            '', '', '', '', '', '', '', '', '', '', '', '', 'Left Eye'
        ])

        writer.writerow([
            'Imported', 'Customer', 'Order_num', 'Frame Type', 'Model', 'Color',
            'Size', 'SizeB', 'Index', 'Material', 'Dia[ref only]', 'Products', 'Treament',
            'SPH', 'CYL', 'Axis', 'ADD', 'Prism', 'Direction', 'BASE', 'Decenter1', 'CT', 'ET', 'PD1', 'PH1', 'Qty',
            'SPH', 'CYL', 'Axis', 'ADD', 'Prism', 'Direction', 'BASE', 'Decenter1', 'CT', 'ET', 'PD1', 'PH1', 'Qty',
            'Tinting', 'Remark'
        ])

        count = len(items)
        index = 0

        for lbo in lbos:
            index += 1
            lbo = __formate_prescripion(lbo)

            logging.debug(lbo.__dict__)
            #
            # writer.writerow([
            #
            #     lbo.create_at, 'ZJ', lbo.lab_number, lbo.frame_type, 'Model', 'Color',
            #     lbo.size, 'SizeB', lbo.lens_index, 'Material', lbo.dia_1, lbo.lens_name, 'HMC|' + lbo.coating_name,
            #     # 'SPH', 'CYL', 'Axis', 'ADD', 'Prism', 'Direction', 'BASE', 'Decenter1', 'CT', 'ET', 'PD1', 'PH1', 'Qty',
            #     lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_add, lbo.od_prism, 'Direction', lbo.od_base, 'Decenter1',
            #     'CT', 'ET', lbo.od_pd, 'PH1', '1',
            #     lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_add, lbo.os_prism, 'Direction', lbo.os_base, 'Decenter1',
            #     'CT', 'ET', lbo.os_pd, 'PH1', '1',
            #
            #     # 'Tinting', 'Remark',
            #     lbo.tint_name, lbo.comments
            # ])

            writer.writerow([

                lbo.create_at, 'ZJ', lbo.lab_number, lbo.frame_type, lbo.frame[:-3], lbo.frame[-3:],
                lbo.size, lbo.lens_height, lbo.lens_index, '-', lbo.dia_1, lbo.act_lens_name, 'HMC|' + lbo.coating_name,
                # 'SPH', 'CYL', 'Axis', 'ADD', 'Prism', 'Direction', 'BASE', 'Decenter1', 'CT', 'ET', 'PD1', 'PH1', 'Qty',
                lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_add, lbo.od_prism, '-', lbo.od_base, '-',
                '-', '-', lbo.od_pd, '-', '1',
                lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_add, lbo.os_prism, '-', lbo.os_base, '-',
                '-', '-', lbo.os_pd, '-', '1',

                # 'Tinting', 'Remark',
                lbo.tint_name, lbo.comments
            ])

        return response

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse('生成采购订单遇到异常[ %s ], 请暂时手动生成，并联系系统支持....' % e.message)


# 镜片采购单生成 列表页
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_purchase_orders(request):
    page_info = {}
    page = request.GET.get('page', 1)
    laborder_vd = request.GET.get('filter', '1')
    currentPage = int(page)
    p_type = {}
    try:
        vendors_choice_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            if int(vc.key) > 0:
                vendors_choice_list.append(vc)

        q1 = Q()
        q1.connector = 'OR'
        q1.children.append(('status', 'FRAME_OUTBOUND'))
        q1.children.append(('status', 'REQUEST_NOTES'))
        # 查出所有状态为‘出库申请’的订单
        frame_outbound_lab_orders = LabOrder.objects.filter(is_enabled=True).filter(
            vendor=laborder_vd).filter(q1).order_by("-id")

        # 获取采购订单类型choices
        for pt in laborder_purchase_order_line.PURCHASE_TYPE_CHOICES:
            p_type[pt[0]] = pt[1]

        # 分页
        count = 0
        count_e = len(frame_outbound_lab_orders)
        for lbo in frame_outbound_lab_orders:
            if not lbo.is_procured:
                count += 1
        page_info['total'] = count
        page_info['all_num'] = count_e
        page_info['vd'] = laborder_vd
        paginator = Paginator(frame_outbound_lab_orders, 20)
        try:
            frame_outbound_lab_orders = paginator.page(page)
        except PageNotAnInteger:
            frame_outbound_lab_orders = paginator.page(1)
        except EmptyPage:
            frame_outbound_lab_orders = paginator.page(paginator.num_pages)

        return render(request, "laborder_purchase_orders.html", {
            'page_info': page_info,
            'lbos': frame_outbound_lab_orders,
            'currentPage': currentPage,
            'paginator': paginator,
            'requestUrl': "/oms/laborder_purchase_orders/",
            'filter': laborder_vd,
            'p_type': p_type,
            'vendors_choices': vendors_choice_list,
        })

    except Exception as e:
        logging.debug('Exception: %s' % e.message)


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_purchase_orders(request):
    page_info = {}
    page = request.GET.get('page', 1)
    laborder_vd = request.GET.get('filter', '1')
    currentPage = int(page)

    try:
        vendors_choice_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            if int(vc.key) > 0:
                vendors_choice_list.append(vc)

        lopos = laborder_purchase_order.objects.filter(vendor=laborder_vd).order_by("-id")

        # 分页
        count = 0
        if len(lopos) > 0:
            count = lopos.count()
        page_info['total'] = count
        paginator = Paginator(lopos, 20)
        try:
            lopos = paginator.page(page)
        except PageNotAnInteger:
            lopos = paginator.page(1)
        except EmptyPage:
            lopos = paginator.page(paginator.num_pages)

        return render(request, "purchase_orders.html", {
            'page_info': page_info,
            'lopos': lopos,
            'currentPage': currentPage,
            'paginator': paginator,
            'requestUrl': "/oms/purchase_orders/",
            'filter': laborder_vd,
            'vendors_choices': vendors_choice_list,
        })

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse(e.message)


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_purchase_order_line(request):
    page_info = {}
    page = request.GET.get('page', 1)
    lopo_id = request.GET.get('lid')
    currentPage = int(page)
    can_print_wx = True

    lbo_entities = []
    try:
        lopo = laborder_purchase_order.objects.filter(pk=lopo_id)
        lopols = laborder_purchase_order_line.objects.filter(lpo=lopo).order_by("-id")
        for lopol in lopols:
            lbo_entities.append((lopol.laborder_entity, lopol.is_purchased))
            # 如果有任何一个purchase_type为ASSEMBLED 不能打印伟星的
            if lopol.purchase_type == "ASSEMBLED":
                can_print_wx = False

        # 分页
        count = 0
        if len(lbo_entities) > 0:
            count = len(lbo_entities)
        page_info['total'] = count
        page_info['lid'] = lopo_id
        paginator = Paginator(lbo_entities, 20)
        try:
            lbo_entities = paginator.page(page)
        except PageNotAnInteger:
            lbo_entities = paginator.page(1)
        except EmptyPage:
            lbo_entities = paginator.page(paginator.num_pages)

        return render(request, "purchase_order_line.html", {
            'page_info': page_info,
            'lbos': lbo_entities,
            'lpo_id': lopo_id,
            'currentPage': currentPage,
            'paginator': paginator,
            'requestUrl': "/oms/purchase_order_line/",
            'can_print_wx': can_print_wx,
        })

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse(e.message)


# 设置出库申请单某一项为下单成功，并推送到MRP
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def set_purchase_order_is_in_rx(request):
    res = {}
    rx_lab_number = request.POST.get('rx_lab_number', '')
    td_lab_number = request.POST.get('td_lab_number', '')

    try:
        # 获取采购订单
        lopols = laborder_purchase_order_line.objects.filter(lab_number=td_lab_number).order_by("-id")
        logging.debug('lab_num=%s' % td_lab_number)
        if lopols.count() == 0:
            res['code'] = -1
            res['message'] = '采购单该条记录不存在'
            json_res = json.dumps(res)
            return HttpResponse(json_res)
        lopol = lopols[0]
        # 查询出对应的lab_order
        lbos = LabOrder.objects.filter(lab_number=td_lab_number)
        if lbos.count() == 0:
            res['code'] = -1
            res['message'] = 'lab_order不存在'
            json_res = json.dumps(res)
            return HttpResponse(json_res)
        lbo = lbos[0]
        # 写入车房单号
        lopol.vendor_order_reference = rx_lab_number
        lopol.save()
        # 推送到MRP
        url = MRP_BASE_URL + 'api/LabOrder/'
        headers = {'Content-Type': 'application/json'}
        # 封装数据
        values = {}
        values['id'] = lbo.id
        values['purchase_order_created_at'] = str(lopol.created_at)
        values['vendor_order_reference'] = lopol.vendor_order_reference
        values['chanel'] = lbo.chanel
        values['create_at'] = str(lbo.create_at)
        values['vendor'] = lbo.vendor
        data = json.dumps(values)
        logging.debug(data)
        # 调用接口
        req = requests.put(url=url, data=data, headers=headers)
        resp = req.text
        respjs = json.loads(resp)
        logging.debug(respjs)
        if respjs['code'] == 400:
            res['code'] = -1
            res['message'] = '推送到MRP失败，请记录下单号，并联系IT'
            json_res = json.dumps(res)
            return HttpResponse(json_res)

        res['code'] = 0
        res['message'] = '成功'
        logging.debug('成功')
        json_res = json.dumps(res)
        return HttpResponse(json_res)

    except Exception as e:
        logging.debug('Exception:' + str(e))
        res['code'] = -1
        res['message'] = str(e)
        json_res = json.dumps(res)
        return HttpResponse(json_res)


# 更新订单状态为print date
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def updata_laborder_status_print_date(request):
    logging.debug('开始更新状态')
    lopo_id = request.POST.get('lid', 0)
    logging.debug('id%s' % lopo_id)
    try:
        if lopo_id == 0:
            res = {}
            res['code'] = -1
            res['message'] = '错误'
            json_res = json.dumps(res)
            return HttpResponse(json_res)
        lopo = laborder_purchase_order.objects.filter(pk=int(lopo_id))
        lopols = laborder_purchase_order_line.objects.filter(lpo=lopo).order_by("-id")
        # add(self, id, user_entity=None, user_id=-1, user_name='system', comments='')
        cvc = construction_voucher_control()
        for lopol in lopols:
            cvc.add(lopol.lab_number, request.user, request.user.id, request.user.username, '批量更新状态为镜片生产')
        res = {}
        res['code'] = 0
        res['message'] = '成功'
        json_res = json.dumps(res)
        return HttpResponse(json_res)

    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        res = {}
        res['code'] = 2
        res['message'] = str(e)
        json_res = json.dumps(res)
        return HttpResponse(json_res)


# 推送采购订单到MRP
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def push_purchase_order_to_mrp(request):
    logging.debug('开始推送到MRP')
    lopo_id = request.POST.get('lid', 0)
    logging.debug('id%s' % lopo_id)
    try:
        if lopo_id == 0:
            res = {}
            res['code'] = -1
            res['message'] = '错误'
            json_res = json.dumps(res)
            return HttpResponse(json_res)
        lopo = laborder_purchase_order.objects.filter(pk=int(lopo_id))  # 采购订单
        lopols = laborder_purchase_order_line.objects.filter(lpo=lopo).order_by("-id")  # 采购订单子项
        # get参数
        url = MRP_BASE_URL + 'api/LabOrder/'
        headers = {'Content-Type': 'application/json'}
        message = '失败数量:0'
        num = 0
        for lopol in lopols:
            # 查询出对应的lab_order
            lbos = LabOrder.objects.filter(id=lopol.laborder_id)
            if lbos.count() == 0:
                logging.debug('lbo不存在，跳过')
                continue
            lbo = lbos[0]
            # 封装数据
            values = {}
            values['id'] = lopol.laborder_id
            values['purchase_order_created_at'] = str(lopol.created_at)
            values['vendor_order_reference'] = lopol.vendor_order_reference
            values['chanel'] = lbo.chanel
            values['create_at'] = str(lbo.create_at)
            values['vendor'] = lbo.vendor
            data = json.dumps(values)
            logging.debug(data)
            # 调用接口
            req = requests.put(url=url, data=data, headers=headers)
            resp = req.text
            respjs = json.loads(resp)
            logging.debug(respjs)
            if respjs['code'] == 400:
                num += 1
                message = '失败数量:%s' % str(num)
        res = {}
        res['code'] = 0
        res['message'] = '完成。' + message
        json_res = json.dumps(res)
        return HttpResponse(json_res)

    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        res = {}
        res['code'] = -1
        res['message'] = str(e)
        json_res = json.dumps(res)
        return HttpResponse(json_res)


# 伟星下单
@login_required
@permission_required('oms.WX_ORDER', login_url='/oms/forbid/')
def wx_purchasing_order(request):
    m_pur_list = []
    rm = res_msg.response_dict()
    line_id = request.POST.get('line_id')
    sub_list = request.POST.getlist('sub_list')
    is_ready = request.POST.get('is_ready')

    try:
        # wx controller类
        wx_pur_ctrl = wx_purchase_controller()

        # 先获取所有line到前端
        if is_ready == '1':
            pur_list = wx_pur_ctrl.get_purchase_list(line_id)
            for item in pur_list:
                m_pur_list.append(item[0])

            return JsonResponse({'m_pur_list': m_pur_list, 'code': '1'})

        # 获取账套
        acc_success = wx_pur_ctrl.get_account()

        if acc_success == True or acc_success == 'true':
            # 下单流程 开始 & 获取结果集
            stat_dict = wx_pur_ctrl.run_purchase(sub_list, request)

            # logging.debug(json.dumps(stat_dict))
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })

        else:
            rm['code'] = -1
            rm['message'] = 'Get Account Fail...'
    except Exception as e:
        rm['code'] = -1
        rm['message'] = 'Error Message: %s' % str(e)

    return JsonResponse(rm)


# 五彩自动下单
@login_required
@permission_required('oms.WX_ORDER', login_url='/oms/forbid/')
def wc_purchasing_order(request):
    # 定义返回参数
    stat_dict = {}
    # 获取必要参数
    line_id = request.POST.get('sub_list', '')
    try:
        # 获取采购订单line对象
        purchase_order_line_entitys = laborder_purchase_order_line.objects.filter(id=line_id)

        # 检查
        if purchase_order_line_entitys.count() < 1:  # 采购订单不存在
            stat_dict[line_id] = {'Success': False, 'Message': '采购订单不存在'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })
        # 取出对象
        purchase_order_line_entity = purchase_order_line_entitys[0]
        if purchase_order_line_entity.is_purchased:  # 已经下单
            stat_dict[purchase_order_line_entity.lab_number] = {'Success': False, 'Message': '该订单已下单'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })
        if purchase_order_line_entity.laborder_entity.status not in (
                'PRINT_DATE', 'LENS_REGISTRATION', 'LENS_RECEIVE', 'ASSEMBLING', 'FRAME_OUTBOUND', 'REQUEST_NOTES'):
            stat_dict[purchase_order_line_entity.lab_number] = {'Success': False, 'Message': '该状态不能下单'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })

        # 调用controller，开始下单
        wcpc = wc_purchase_controller()
        stat_dict = wcpc.run_purchase(request, purchase_order_line_entity)

        # 返回
        return render(request, 'purchase_res_template.html', {
            'stat_list': stat_dict
        })
    except Exception as e:
        logging.debug(str(e))
        stat_dict[line_id] = {'Success': False, 'Message': str(e)}
        return render(request, 'purchase_res_template.html', {
            'stat_list': stat_dict
        })


# 生成一般采购订单
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_purchase_generate_csv(request):
    lab_order_list = []
    lab_order_pur_type_list = []
    rm = res_msg.response_dict()
    if request.method == "POST":
        is_full = request.POST.get("is_full")
        laborder_vd = request.POST.get("vd")
        id_list = request.POST.getlist("id_list")
        all_type = request.POST.get("all_type")

        try:
            with transaction.atomic():
                if is_full == "True":
                    q1 = Q()
                    q1.connector = 'OR'
                    q1.children.append(('status', 'FRAME_OUTBOUND'))
                    q1.children.append(('status', 'REQUEST_NOTES'))
                    laborders = LabOrder.objects.filter(is_enabled=True).filter(q1).filter(
                        vendor=laborder_vd).order_by("-id")
                    for lbo in laborders:
                        if not lbo.is_procured:
                            lab_order_list.append(lbo)
                else:
                    for purchase_order_id in id_list:
                        po_info_arr = purchase_order_id.split('.')
                        lbo = LabOrder.objects.get(pk=po_info_arr[0])
                        lab_order_pur_type_list.append(po_info_arr[1])
                        lab_order_list.append(lbo)

                if len(lab_order_list) == 0:
                    rm["code"] = -1
                    rm["message"] = "请先选择要生成采购订单的订单项"
                    return JsonResponse(rm)

                # 生成 laborder_purchase_order
                lopo = laborder_purchase_order()
                lopo.count = len(lab_order_list)
                lopo.vendor = laborder_vd
                lopo.user_id = request.user.id
                lopo.user_name = request.user.username
                lopo.save()

                for idx in range(len(lab_order_list)):
                    lopol = laborder_purchase_order_line()
                    lopol.lpo = lopo
                    lopol.laborder_entity = lab_order_list[idx]
                    lopol.laborder_id = lab_order_list[idx].id
                    lopol.frame = lab_order_list[idx].frame
                    lopol.lab_number = lab_order_list[idx].lab_number
                    lopol.quantity = lab_order_list[idx].quantity
                    lopol.lens_type = lab_order_list[idx].lens_type
                    lopol.order_date = lab_order_list[idx].order_date
                    lopol.order_created_date = lab_order_list[idx].create_at
                    if is_full == "True":
                        lopol.purchase_type = all_type
                    else:
                        lopol.purchase_type = lab_order_pur_type_list[idx]
                    lopol.save()

                # 返回laborder_purchase_order的id 前端再通过id去get
                rm["message"] = lopo.id
                return JsonResponse(rm)

        except Exception as e:
            logging.debug('Exception: %s' % str(e))
            rm["code"] = -1
            rm["message"] = "%s" % str(e)
            return JsonResponse(rm)
    # GET
    else:
        try:
            # import csv, codecs
            # # Print Start
            # lopo_id = request.GET.get("lopo_id")
            # lrn = laborder_purchase_order.objects.get(id=lopo_id)
            # lab_order_list = lrn.laborder_entities_frame_outbound
            #
            # items = []
            # for lbo in lab_order_list:
            #     items.append(lbo.lab_number)
            #
            # logging.debug('items count: %s' % len(items))
            #
            # response = HttpResponse(content_type='text/csv')
            # file_name = 'laborder_numbers_csv'
            # response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'
            #
            # response.write(codecs.BOM_UTF8)
            #
            # writer = csv.writer(response)
            #
            # writer.writerow(['订单信息'])
            # writer.writerow([
            #     '', '', '', '', '', '', '', '', '', '', '', '',
            #     '',
            #     '验光单', '', '', 'PD', '', 'ADD', 'prism-H', 'base-H', 'prism-V', 'base-V'
            # ])
            #
            # writer.writerow([
            #     '序号', '配送', '下达日期', '订单号', '镜架号', '镜架尺寸', '框型', '框高',
            #     '名称', '设计', '涂层名称', '染色名称',
            #     'R/L',
            #     'SPH', 'CYL', 'AXIS', '单眼', '双眼', '', '', '', '', '',
            #     '备注', '加工瞳高', '装配瞳高', '子镜高度', '特殊处理', '采购类型'
            # ])
            #
            # index = 0
            # for lbo in lab_order_list:
            #     index += 1
            #     lbo = __formate_prescripion(lbo)
            #     lopol = laborder_purchase_order_line.objects.get(laborder_entity=lbo.id)
            #     pal_design_name_done = lbo.pal_design_name
            #     if pal_design_name_done == 'IOT 渐进 Alpha H45':
            #         pal_design_name_done = 'IOT 渐进 Alpha H45NC'
            #     if pal_design_name_done == 'IOT 渐进 Alpha H25':
            #         pal_design_name_done = 'IOT 渐进 Alpha H25NC'
            #     writer.writerow([
            #         index, '库房自提', lbo.order_date, lbo.lab_number, lbo.frame, lbo.size, lbo.frame_type, lbo.lens_height,
            #         lbo.act_lens_name,
            #         pal_design_name_done,
            #         lbo.coating_name,
            #         lbo.tint_name,
            #         'OD-右眼',
            #         lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_pd, lbo.pd, lbo.od_add, lbo.od_prism, lbo.od_base,
            #         lbo.od_prism1, lbo.od_base1,
            #         lbo.comments, "%smm" % lbo.lab_seg_height, "按标准瞳高%smm" % lbo.assemble_height,
            #                       "%smm" % lbo.sub_mirrors_height,
            #         lbo.special_handling, lopol.get_purchase_type_display()
            #     ])
            #
            #     writer.writerow([
            #         '', '', '', '', '', '', '', '', '', '', '', '', 'OS-左眼',
            #         lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_pd, '', lbo.os_add, lbo.os_prism, lbo.os_base,
            #         lbo.os_prism1, lbo.os_base1,
            #         ''
            #     ])
            # return response
            return generate_po_xls(request)
        except Exception as e:
            logging.debug('Exception: %s' % str(e))
            rm["code"] = -1
            rm["message"] = "生成采购订单遇到异常[ %s ], 请暂时手动生成，并联系系统支持...." % str(e)
            return JsonResponse(rm)

@login_required
def generate_po_xls(request):
    lopo_id = request.GET.get("lopo_id")
    lrn = laborder_purchase_order.objects.get(id=lopo_id)
    lab_order_list = lrn.laborder_entities_frame_outbound
    vd = lrn.vendor

    # add by ranhy 2020-09-22
    str_time = time.strftime("%Y-%m-%d", time.localtime(time.time()))
    file_name = 'vd%s_po_%s.xlsx' % (vd, str_time)

    # 设置HTTPResponse的类型
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename=%s' % file_name
    byteio = BytesIO()

    wb = openpyxl.Workbook()  # 新建工作簿
    ws = wb.active
    ws.title = '总表'
    ws.append(['订单信息'])
    ws.append([
        '', '', '', '', '', '', '', '', '', '', '', '',
        '',
        '验光单', '', '', 'PD', '', 'ADD', 'prism-H', 'base-H', 'prism-V', 'base-V'
    ])

    ws.append([
        '序号', '配送', '下达日期', '订单号', '镜架号', '镜架尺寸', '框型', '框高',
        '名称', '设计', '涂层名称', '染色名称',
        'R/L',
        'SPH', 'CYL', 'AXIS', '单眼', '双眼', '', '', '', '', '',
        '备注', '加工瞳高', '装配瞳高', '子镜高度', '特殊处理', '采购类型'
    ])

    list_lab_lens = []
    list_lab_lens_index = []
    index = 0
    for lbo in lab_order_list:
        index += 1
        lbo = __formate_prescripion(lbo)
        lopol = laborder_purchase_order_line.objects.get(laborder_entity=lbo.id)
        pal_design_name_done = lbo.pal_design_name
        if pal_design_name_done == 'IOT 渐进 Alpha H45':
            pal_design_name_done = 'IOT 渐进 Alpha H45NC'
        if pal_design_name_done == 'IOT 渐进 Alpha H25':
            pal_design_name_done = 'IOT 渐进 Alpha H25NC'
        ws.append([
            index, '库房自提', lbo.order_date, lbo.lab_number, lbo.frame, lbo.size, lbo.frame_type, lbo.lens_height,
            lbo.act_lens_name,
            pal_design_name_done,
            lbo.coating_name,
            lbo.tint_name,
            'OD-右眼',
            lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_pd, lbo.pd, lbo.od_add, lbo.od_prism, lbo.od_base,
            lbo.od_prism1, lbo.od_base1,
            lbo.comments, "%smm" % lbo.lab_seg_height, "按标准瞳高%smm" % lbo.assemble_height,
                          "%smm" % lbo.sub_mirrors_height,
            lbo.special_handling, lopol.get_purchase_type_display()
        ])

        ws.append([
            '', '', '', '', '', '', '', '', '', '', '', '', 'OS-左眼',
            lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_pd, '', lbo.os_add, lbo.os_prism, lbo.os_base,
            lbo.os_prism1, lbo.os_base1,
            ''
        ])
        if (str(vd) == '2'):
            if float(lbo.od_sph)>0:
                list_lab_lens.append([lbo.act_lens_sku, lbo.act_lens_name,'老花',lbo.od_sph, lbo.od_cyl, lbo.id])
                if [lbo.act_lens_sku,lbo.act_lens_name,"老花"] not in list_lab_lens_index:
                    list_lab_lens_index.append([lbo.act_lens_sku,lbo.act_lens_name,'老花'])
            elif float(lbo.od_sph)<0:
                list_lab_lens.append([lbo.act_lens_sku, lbo.act_lens_name,'近视', lbo.od_sph, lbo.od_cyl, lbo.id])
                if [lbo.act_lens_sku, lbo.act_lens_name,'近视'] not in list_lab_lens_index:
                    list_lab_lens_index.append([lbo.act_lens_sku, lbo.act_lens_name,'近视'])

            if float(lbo.os_sph) >= 0:
                list_lab_lens.append([lbo.act_lens_sku, lbo.act_lens_name,'老花', lbo.os_sph, lbo.os_cyl, lbo.id])
                if [lbo.act_lens_sku, lbo.act_lens_name,'老花'] not in list_lab_lens_index:
                    list_lab_lens_index.append([lbo.act_lens_sku, lbo.act_lens_name,'老花'])
            elif float(lbo.os_sph) < 0:
                list_lab_lens.append([lbo.act_lens_sku, lbo.act_lens_name,'近视', lbo.os_sph, lbo.os_cyl, lbo.id])
                if [lbo.act_lens_sku,lbo.act_lens_name ,'近视'] not in list_lab_lens_index:
                    list_lab_lens_index.append([lbo.act_lens_sku, lbo.act_lens_name,'近视'])
    if str(vd) == '2':
        for lens in list_lab_lens_index:
            X1 = np.arange(0, 201, 25)
            print(lens[0]+lens[2])
            ws = wb.create_sheet(lens[0]+"-"+lens[2])  # 添加页
            ws.title = lens[0]+"-"+lens[2]
            ws.cell(row=1, column=10)  #
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
            ws.cell(1, 1, lens[1]+"(%s)"%(lens[0]+"-"+lens[2]))
            ws.append(['sph/cyl'] + list(X1))
            M = np.zeros((25, 9))
            print(list_lab_lens)
            for lab_lens in list_lab_lens:
                if lens[0] == lab_lens[0] and lens[1] == lab_lens[1] and lens[2] == lab_lens[2]:
                        k = abs(float(lab_lens[3]) * 100 / 25)
                        j = abs(float(lab_lens[4]) * 100 / 25)
                        M[int(k), int(j)] += 1
            m_list = M.tolist()
            i = 0
            for k in m_list:
                ws.append([i * 25] + k)
                i = i + 1

            fill = PatternFill("solid", fgColor="1874CD")
            font = Font(name='Calibri',
                        size=16,
                        color='FF000000',
                        bold=True,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False)

            font1 = Font(name='Calibri',
                        size=16,
                        color='FF000000',
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False)
            from   openpyxl.styles import borders
            border = Border(left=Side(border_style=borders.BORDER_THIN,color = "000000"),
                            right = Side(border_style=borders.BORDER_THIN,color = '000000'),
                            top = Side(border_style=borders.BORDER_THIN,color = '000000'),
                            bottom = Side(border_style=borders.BORDER_THIN,color = '000000'),
                            # diagonal = Side(border_style=None,color = '0099CC00'),
                            # diagonal_direction = 0,
                            # outline = Side(border_style=None,color = '0099CC00'),
                            # vertical = Side(border_style=None,color = '0099CC00'),
                            # horizontal = Side(border_style=None,color = '0099CC00')
                            )
            alignment = Alignment(horizontal='general',
                                  vertical = 'bottom',
                                  text_rotation = 0,
                                  wrap_text = False,
                                  shrink_to_fit = False,
                                  indent = 0)
            number_format = 'General'
            for row in ws.iter_rows():
                for i in row:
                    i.number_format = number_format
                    i.border = border
                    i.alignment = alignment

                    if(i>=3 and i.value>0):
                        # i.fill = fill
                        i.font = font
                    else:
                        i.font = font1

                # fill = PatternFill("solid", fgColor="1874CD")

    wb.save(byteio)  # 一定要保存
    byteio.seek(0)
    response.write(byteio.getvalue())
    return response

# 生成伟星的采购订单
@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_purchase_generate_csv_wx(request):
    lab_order_list = []
    lab_order_pur_type_list = []
    rm = res_msg.response_dict()
    if request.method == "POST":
        is_full = request.POST.get("is_full")
        laborder_vd = request.POST.get("vd")
        id_list = request.POST.getlist("id_list")
        all_type = request.POST.get("all_type")

        try:
            with transaction.atomic():
                if is_full == "True":
                    q1 = Q()
                    q1.connector = 'OR'
                    q1.children.append(('status', 'FRAME_OUTBOUND'))
                    q1.children.append(('status', 'REQUEST_NOTES'))
                    laborders = LabOrder.objects.filter(is_enabled=True).filter(q1).filter(
                        vendor=laborder_vd).order_by("-id")
                    for lbo in laborders:
                        if not lbo.is_procured:
                            lab_order_list.append(lbo)
                    if all_type == "ASSEMBLED":
                        rm["code"] = "-1"
                        rm["message"] = "此厂家不支持【装配】类型采购订单"
                        return JsonResponse(rm)
                else:
                    for purchase_order_id in id_list:
                        po_info_arr = purchase_order_id.split('.')
                        lbo = LabOrder.objects.get(pk=po_info_arr[0])
                        lab_order_pur_type_list.append(po_info_arr[1])
                        lab_order_list.append(lbo)

                if len(lab_order_list) == 0:
                    rm["code"] = -1
                    rm["message"] = "请先选择要生成采购订单的订单项"
                    return JsonResponse(rm)

                for loptl in lab_order_pur_type_list:
                    if loptl == "ASSEMBLED":
                        rm["code"] = "-1"
                        rm["message"] = "此厂家不支持【装配】类型采购订单"
                        return JsonResponse(rm)

                # 生成 laborder_purchase_order
                lopo = laborder_purchase_order()
                lopo.count = len(lab_order_list)
                lopo.vendor = laborder_vd
                lopo.user_id = request.user.id
                lopo.user_name = request.user.username
                lopo.save()

                for idx in range(len(lab_order_list)):
                    lopol = laborder_purchase_order_line()
                    lopol.lpo = lopo
                    lopol.laborder_entity = lab_order_list[idx]
                    lopol.laborder_id = lab_order_list[idx].id
                    lopol.frame = lab_order_list[idx].frame
                    lopol.lab_number = lab_order_list[idx].lab_number
                    lopol.quantity = lab_order_list[idx].quantity
                    lopol.lens_type = lab_order_list[idx].lens_type
                    lopol.order_date = lab_order_list[idx].order_date
                    lopol.order_created_date = lab_order_list[idx].create_at
                    if is_full == "True":
                        lopol.purchase_type = all_type
                    else:
                        lopol.purchase_type = lab_order_pur_type_list[idx]
                    lopol.save()

                # 返回laborder_purchase_order的id 前端再通过id去get
                rm["message"] = lopo.id
                return JsonResponse(rm)

        except Exception as e:
            logging.debug('Exception: %s' % str(e))
            rm["code"] = -1
            rm["message"] = "%s" % str(e)
            return JsonResponse(rm)
    # GET
    else:
        try:
            import csv, codecs
            # Print Start
            lopo_id = request.GET.get("lopo_id")
            lrn = laborder_purchase_order.objects.get(id=lopo_id)
            lab_order_list = lrn.laborder_entities_frame_outbound

            response = HttpResponse(content_type='text/csv')
            file_name = 'laborder_numbers_csv'
            response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'

            response.write(codecs.BOM_UTF8)

            writer = csv.writer(response)

            writer.writerow([
                '', '', '', '', 'Frame', '', '', '', '',
                '', '', '', '', 'Right Eye', '', '', '', '',
                '', '', '', '', '', '', '', '', '', '', '', '', 'Left Eye'
            ])

            writer.writerow([
                'Imported', 'Customer', 'Order_num', 'Frame Type', 'Model', 'Color',
                'Size', 'SizeB', 'Index', 'Material', 'Dia[ref only]', 'Products', 'Treament',
                'SPH', 'CYL', 'Axis', 'ADD', 'Prism-H', 'Direction', 'BASE-H', 'Decenter1',
                'Prism-V', 'Direction', 'BASE-V', 'Decenter1',
                'CT', 'ET', 'PD1', 'PH1', 'Qty',
                'SPH', 'CYL', 'Axis', 'ADD', 'Prism-H', 'Direction', 'BASE-H', 'Decenter1',
                'Prism-V', 'Direction', 'BASE-V', 'Decenter1',
                'CT', 'ET', 'PD1', 'PH1', 'Qty',
                'Tinting', 'Remark', 'Lab Seg Height', 'Assemble Height', 'Sub Mirrors Height', 'Special Handling',
                'Purchase Type',
                'Pal Design Code', 'Pal Design'
            ])

            index = 0
            split_str = '装配瞳高按标准瞳高'
            for lbo in lab_order_list:
                index += 1
                lbo = __formate_prescripion(lbo)
                logging.debug(lbo.__dict__)
                lopol = laborder_purchase_order_line.objects.get(laborder_entity=lbo.id)

                comments = lbo.comments
                if lopol.purchase_type == 'GLASSES':
                    comments = lopol.get_purchase_type_display() + '|' + comments

                writer.writerow([
                    lbo.create_at, 'ZJ', lbo.lab_number, lbo.frame_type, lbo.frame[:-3], lbo.frame[-3:],
                    lbo.size, lbo.lens_height, lbo.lens_index, '-', lbo.dia_1, lbo.act_lens_name,
                    lbo.coating_name,
                    lbo.od_sph, lbo.od_cyl, lbo.od_axis, lbo.od_add, lbo.od_prism, '-', lbo.od_base, '-',
                    lbo.od_prism1, '-', lbo.od_base1, '-',
                    '-', '-', lbo.od_pd, lbo.lab_seg_height, '1',
                    lbo.os_sph, lbo.os_cyl, lbo.os_axis, lbo.os_add, lbo.os_prism, '-', lbo.os_base, '-',
                    lbo.os_prism1, '-', lbo.os_base1, '-',
                    '-', '-', lbo.os_pd, lbo.lab_seg_height, '1',
                    lbo.tint_name, comments, "%smm" % lbo.lab_seg_height, "按标准瞳高%smm" % lbo.assemble_height,
                                             "%smm" % lbo.sub_mirrors_height,
                    lbo.special_handling,
                    lopol.get_purchase_type_display(),
                    lbo.pal_design_sku,
                    lbo.pal_design_name
                ])
            return response

        except Exception as e:
            logging.debug('Exception: %s' % str(e))
            rm["code"] = -1
            rm["message"] = "生成采购订单遇到异常[ %s ], 请暂时手动生成，并联系系统支持...." % str(e)
            return JsonResponse(rm)


def __formate_prescripion(lbo):
    # lbo = LabOrder()

    if lbo.is_singgle_pd:
        lbo.od_pd = lbo.pd / 2
        lbo.os_pd = lbo.pd / 2

    lbo.od_sph = __formate_value(lbo.od_sph)
    logging.debug(type(lbo.od_sph))
    logging.debug(lbo.od_sph)
    lbo.od_cyl = __formate_value(lbo.od_cyl)
    lbo.od_add = __formate_value(lbo.od_add)
    lbo.od_prism = __formate_value(lbo.od_prism)
    lbo.od_prism1 = __formate_value(lbo.od_prism1)

    lbo.os_sph = __formate_value(lbo.os_sph)
    lbo.os_cyl = __formate_value(lbo.os_cyl)
    lbo.os_add = __formate_value(lbo.os_add)
    lbo.os_prism = __formate_value(lbo.os_prism)
    lbo.os_prism1 = __formate_value(lbo.os_prism1)

    if lbo.frame_type == None:
        lbo.frame_type = ''

    if lbo.coating_name == None:
        lbo.coating_name = ''

    if lbo.tint_name == None:
        lbo.tint_name = ''

    return lbo


def __formate_value(value):
    if value == "":
        return value

    try:
        return ("{:+.2f}".format(float(value)))
    except Exception as e:
        logging.debug(e.message)
        return ''


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_print(request):
    from wms.models import inventory_struct_warehouse
    _form_data = {}
    id = request.GET.get('id', '')
    lines = request.GET.get('lines', 30)
    items = []
    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    try:
        _form_data['id'] = id
        lrn = laborder_request_notes.objects.get(id=id)
        lbos = laborder_request_notes_line.objects.filter(lrn__id=id).order_by('location')

        # Print Start ================================

        items = lbos
        count = len(items)  # 数据总条数
        lines = int(lines)  # 每页总行数
        logging.debug('count: %s' % count)
        copies = count // lines
        logging.debug('copies: %s' % copies)
        remainder = count % lines
        if remainder > 0:
            copies += 1
        logging.debug('copies: %s' % copies)

        index = 0
        items = []

        for i in range(copies):
            pitms = ParentItems()
            pitms.index = i + 1
            pitms.count = copies
            pitms.created_at = datetime.date.today()
            pitms.warehouse_code = lrn.warehouse_code
            pitms.items = []

            for j in range(lines):
                itm = Items()
                itm.index = index

                invss = inventory_struct_warehouse.objects.filter(sku=lbos[index].frame,
                                                                  warehouse_code=lbos[index].warehouse_code)

                if len(invss) > 0:
                    itm.quantity = invss[0].quantity

                obj = Items()
                obj.index = index + 1
                obj.obj = lbos[index]
                itm.left = obj
                pitms.items.append(itm)

                index += 1
                if index == count:
                    break

            items.append(pitms)
            if index == count:
                break

        return render(request, "laborder_request_notes_print_old.html", {
            'list': items,
            'form_data': _form_data,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request_notes_print_old.html", {
            'list': items,
            'form_data': _form_data,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })


# 生成镜片出库申请单
@login_required
@permission_required('oms.RN_VIEW_LENS', login_url='/oms/forbid/')
def redirect_laborder_request_lens_notes_print(request):
    id = request.GET.get('id', '')

    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    try:
        items = None
        with connections['pg_oms_query'].cursor() as cursor:

            sql = 'SELECT t0.id,t0.lab_number,t1.act_lens_sku,t1.act_lens_name,t1.od_sph,t1.od_cyl,t1.os_sph,t1.os_cyl FROM oms_laborder_request_notes_line t0 LEFT JOIN oms_laborder t1 ON t0.laborder_entity_id=t1.id WHERE t0.lrn_id=%s ORDER BY t0.id' % id
            cursor.execute(sql)
            from util.db_helper import *

            items = namedtuplefetchall(cursor)
            logging.critical("items count:%s" % items.count())
        return render(request, "laborder_request__lens_notes_print_old.html", {
            'list': items,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request__lens_notes_print_old.html", {
            'list': items,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })


# 生成镜片出库申请汇总
@login_required
@permission_required('oms.RN_VIEW_LENS', login_url='/oms/forbid/')
def redirect_laborder_request_lens_notes_print_count(request):
    id = request.GET.get('id', '')

    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    try:
        items = None
        # 相同镜片合并
        with connections['pg_oms_query'].cursor() as cursor:
            from util.db_helper import *
            sql = '''
                SELECT A.sku as sku,A.name as name,A.sph as sph,A.cyl as cyl,count(*) as quantity FROM
                (SELECT t1.act_lens_sku as sku,t1.act_lens_name as name,t1.od_sph as sph,t1.od_cyl as cyl
                FROM oms_laborder_request_notes_line t0 LEFT JOIN oms_laborder t1 ON t0.laborder_entity_id=t1.id 
                WHERE t0.lrn_id=%s
                union all
                SELECT t1.act_lens_sku as sku,t1.act_lens_name as name,t1.os_sph as sph,t1.os_cyl as cyl
                FROM oms_laborder_request_notes_line t0 LEFT JOIN oms_laborder t1 ON t0.laborder_entity_id=t1.id 
                WHERE t0.lrn_id=%s) as A
                GROUP BY sku,sph,cyl
                ORDER BY sku
                ''' % (id, id)
            # 执行SQL
            cursor.execute(sql)
            items = namedtuplefetchall(cursor)

        return render(request, "laborder_request_lens_notes_print_count_old.html", {
            'list': items,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return render(request, "laborder_request_lens_notes_print_count_old.html", {
            'list': items,
            'requestUrl': '/oms/laborder_request_notes_print/',
        })


@login_required
@permission_required('oms.RN_VIEW', login_url='/oms/forbid/')
def redirect_laborder_request_notes_delivery(request):
    page_info = {}
    form_data = {}
    page = request.GET.get('page', 1)
    currentPage = int(page)
    laborder_filter = request.GET.get('filter', 'all')

    id = request.GET.get('id', '')
    try:
        id = int(id)
    except:
        return HttpResponse('系统获取参数ID遇到错误')

    form_data['id'] = id

    lines = request.GET.get('lines', 30)
    items = []
    paginator = None
    alias_name = 'FRAME_OUTBOUND'

    try:
        filter = {}

        lrn = laborder_request_notes.objects.get(id=id)
        lbos = lrn.laborder_entities

        # Print Start
        items = laborder_request_notes_line.objects.filter(lrn__id=id)

        lines = int(lines)
        count = len(items)
        logging.debug('count: %s' % count)

        for lbo in lbos:

            if lbo.status == None or lbo.status == '' or lbo.status == 'PRINT_DATE' or \
                            lbo.status == 'REQUEST_NOTES':
                laborder_status_backend(
                    request,
                    content='镜架出库',
                    labid=lbo.lab_number,
                    aliasname=alias_name,
                    carrier='',
                    ship='',
                    formDate=None
                )
            else:
                logging.debug('lab order: %s 状态未更新' % lbo.lab_number)
                logging.debug('lab order status: %s' % lbo.status)

        page_info['total'] = lbos.count

        paginator = Paginator(lbos, 5)  # Show 20 contacts per page

        return HttpResponseRedirect('/oms/laborder_request_notes_detail/?id=%s' % id)

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        page_info['total'] = e.message
        return render(request, "laborder_request_notes_detail.html",
                      {
                          'page_info': page_info,
                          'form_data': form_data,
                          'list': items,
                          'currentPage': currentPage, 'paginator': paginator,
                          'requestUrl': '/oms/laborder_request_notes_detail/',
                          'filter': laborder_filter,
                      })


@login_required
@permission_required('oms.OCVT_VIEW', login_url='/oms/forbid/')
def redirect_construction_voucher(request):
    _form_data = {}

    _form_data['request_feature'] = 'Construction Voucher'
    _items = []
    _paginator = None
    _id = -1

    _form_data['total'] = 0
    _id = request.GET.get('id', -1)
    production_flag = request.GET.get('production_flag', '')
    logging.debug('id: %s' % _id)
    _form_data['id'] = _id
    _form_data['production_flag'] = production_flag
    _item = {}

    try:
        if _id <> -1:
            logging.debug('开始')
            # _obj = LabOrder.objects.get(lab_number=id)
            loc = lab_order_controller()
            lbos = loc.get_by_entity(_id)



            if len(lbos) == 1:
                lbo = lbos[0]
                _id = lbo.lab_number

                # 用来到wms_product_frame中查找sku_specs字段（警示信息）
                caution_info = product_frame.objects.get(sku=lbo.frame)
                _form_data['caution_info'] = caution_info.sku_specs

            _lrn = laborder_request_notes_line.objects.get(laborder_entity__lab_number=_id)

            _obj = _lrn.laborder_entity
            _form_data['obj'] = _obj
            _form_data['created_at'] = timezone.now()

            ca = int(time.time())
            ca = time.localtime(ca)
            ca_s = time.strftime("%Y-%m-%d %H:%M", ca)
            _form_data['created_at'] = ca_s
            _item = _obj

        return render(request, "construction_voucher.html", {
            'form_data': _form_data,
            'item': _item,
            'requestUrl': '/oms/construction_voucher/',
        })

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html", {
            'form_data': _form_data,
        })


@login_required
@permission_required('oms.OCVT_VIEW', login_url='/oms/forbid/')
def redirect_construction_voucher_print(request):
    _form_data = {}
    _form_data['request_feature'] = 'Construction Voucher'
    _items = []
    _paginator = None
    _id = -1

    _form_data['total'] = 0
    _id = request.GET.get('id', -1)
    flag = request.GET.get('flag', '')
    production_flag = request.GET.get('production_flag', '')
    logging.debug('id: %s' % _id)
    _item = {}
    cargo_location = ''

    # 未定义先调用错误
    _obj_dict = None
    com_date = None
    locker_obj_num = None
    worknumber = None
    workshop = None
    com_date_str = None
    try:
        times = 0

        if _id != -1:
            logging.debug('开始')
            loc = lab_order_controller()
            lbos = loc.get_by_entity(_id)
            vs = loc.verify_status(_id)

            if len(lbos) == 1:
                lbo = lbos[0]
                _id = lbo.lab_number

            # _obj = LabOrder.objects.get(lab_number=id)
            _lrn = laborder_request_notes_line.objects.get(laborder_entity__lab_number=_id)
            _obj = _lrn.laborder_entity

            gwc = get_workshop_control()
            workshop = request.GET.get('workshop', None)
            logging.debug(workshop)
            if workshop:
                _ws = gwc.get_workshop(_obj.vendor, workshop)
            else:
                _ws = gwc.get_workshop(_obj.vendor)

            logging.debug(_ws)

            if _ws != -1:
                _obj.workshop = _ws
                _obj.save()

            _obj_dict = model_to_dict(_obj)
            for key, value in _obj_dict.items():
                if key == 'ship_direction':
                    _obj_dict[key] = _obj.get_ship_direction_display()
                if value is None or value == '':
                    _obj_dict[key] = '----'

            _obj_dict['lens_index'] = format_str(_obj.lens_index)
            _obj_dict['lens_type'] = format_str(_obj.lens_type)
            ca = int(time.time())
            ca = time.localtime(ca)
            ca_s = time.strftime("%Y-%m-%d %H:%M", ca)
            _form_data['created_at'] = ca_s
            _item = _obj

            lbo = LabOrder.objects.get(lab_number=_id)

            # 用来到wms_product_frame中查找sku_specs字段（警示信息）
            caution_info = product_frame.objects.get(sku=lbo.frame)
            _form_data['caution_info'] = caution_info.sku_specs

            # 仓位号在镜片生产时自动计算给出并打印于作业单；（仓位号计算规则是依次查找仓位是否达到glasses_max_limit(最大数量)，没有返回仓位号，达到了继续查找直到可以放置,同时更改Lockers中的qty）
            # 判断是上海还是dy，找不同得config
            logging.debug(lbo.vendor)
            locker = locker_controller()
            iswc = inventory_struct_warehouse_controller()
            locker_obj_num = ''

            if flag == 'printing':
                # 补打不限制状态
                if not (
                                    _obj.status == 'FRAME_OUTBOUND' or _obj.status == 'PRINT_DATE' or _obj.status == 'COLLECTION') and _obj.quantity == 1:
                    return HttpResponse('订单只有镜架出库状态，才能打印!')
                # 未生成采购订单不允许打印
                if not _obj.is_procured:
                    return HttpResponse('此订单未生成采购订单!')

                # 添加VD1 by zhutong
                if lbo.vendor == '2' or lbo.vendor == '3' or lbo.vendor == '4' or lbo.vendor == '7' \
                        or lbo.vendor == '11' or lbo.vendor == '12' or lbo.vendor == '13' or lbo.vendor == '14' or \
                                lbo.vendor == '15' or lbo.vendor == '1':
                    cargo_location = iswc.get_location(lbo.frame, 'W02')
                    if cargo_location is None:
                        cargo_location = ''
                    # 加急单
                    if lbo.act_ship_direction == 'EXPRESS' or lbo.act_ship_direction == 'CA_EXPRESS':
                        location = 'EPS'
                    else:
                        location = 'SH'
                    max_glass = locker.get_locker_config(location)
                    locker_obj = locker.locker_add(location, max_glass, request.user.username, lbo)
                    print(type(locker_obj.code))
                    if locker_obj.code == 0:
                        # with transaction.atomic():
                        cvc = construction_voucher_control()  # add 改变了lab_order的状态
                        # add中包含了事务
                        cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')
                        locker_obj_num = locker_obj.obj.storage_location + "-" + locker_obj.obj.locker_num
                    elif locker_obj.code == -1:
                        locker_obj_num = locker_obj.obj.storage_location + "-" + locker_obj.obj.locker_num
                        return HttpResponse('<h1>' + str(locker_obj_num) + '仓位中存在该订单，请人工确认！</h1>')
                    elif locker_obj.code == -2:
                        return HttpResponse('<h1>' + '仓位未设置' + location + '-VD' + str(lbo.vendor) + '的专属仓位!</h1>')
                elif lbo.vendor == '9' or lbo.vendor == '10':
                    cargo_location = iswc.get_location(lbo.frame, 'W02')
                    if cargo_location is None:
                        cargo_location = ''
                    # 加急单
                    if lbo.act_ship_direction == 'EXPRESS' or lbo.act_ship_direction == 'CA_EXPRESS':
                        location = 'EPS'
                    else:
                        location = 'SH'
                    max_glass = locker.get_locker_config(location)
                    locker_obj = locker.locker_add(location, max_glass, request.user.username, lbo)
                    if locker_obj.code == 0:
                        # with transaction.atomic():
                        cvc = construction_voucher_control()  # add 改变了lab_order的状态
                        # add中包含了事务
                        cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')
                        locker_obj_num = locker_obj.obj.storage_location + "-" + locker_obj.obj.locker_num
                    elif locker_obj.code == -1:
                        locker_obj_num = locker_obj.obj.storage_location + "-" + locker_obj.obj.locker_num
                        return HttpResponse('<h1>' + str(locker_obj_num) + '仓位中存在该订单，请人工确认！</h1>')
                    elif locker_obj.code == -2:
                        return HttpResponse('<h1>' + '仓位未设置' + location + '-VD' + str(lbo.vendor) + '的专属仓位!</h1>')
                else:
                    cargo_location = ''
                    locker_obj_num = ''
                    cvc = construction_voucher_control()  # add 改变了lab_order的状态
                    # add中包含了事务
                    cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')
            else:
                # 不打印，如果是从装配扫描过来的，修改为待装配状态
                if production_flag == '1':
                    logging.debug(production_flag)
                    if not lbo.status == 'LENS_RECEIVE' and not lbo.status == 'GLASSES_RETURN' and not lbo.status == 'COLLECTION' and not lbo.status == 'ASSEMBLING' and lbo.quantity == 1:
                        _form_data['code'] = -1
                        return HttpResponse(
                            "<h1 style="">订单只有 镜片收货或成镜返工 状态，才能更改状态!当前订单状态为" + format(lbo.status) + "</h1>")
                    lbo.status = 'ASSEMBLING'
                    lbo.save()
                    tloc = tracking_lab_order_controller()
                    tloc.tracking(lbo, request.user, 'ASSEMBLING')
                # if lbo.is_production_change or production_flag == '1':
                #     logging.debug(production_flag)
                #     if not lbo.status == 'LENS_RECEIVE' and not lbo.status == 'GLASSES_RETURN' and not lbo.status == 'COLLECTION' and not lbo.status == 'ASSEMBLING' and lbo.quantity == 1:
                #         _form_data['code'] = -1
                #         return HttpResponse(
                #             "<h1 style="">订单只有 镜片收货或成镜返工 状态，才能更改状态!当前订单状态为" + format(lbo.status) + "</h1>")
                #     lbo.status = 'ASSEMBLING'
                #     lbo.save()
                #     tloc = tracking_lab_order_controller()
                #     tloc.tracking(lbo, request.user, 'ASSEMBLING')

                # 添加VD1
                if _obj.status in ['FRAME_OUTBOUND', 'PRINT_DATE', 'COLLECTION'] and lbo.vendor in ['2', '3', '4',
                                                                                                    '7', '9', '10',
                                                                                                    '11', '12', '13',
                                                                                                    '14', '15','1']:
                    lis = LockersItem.objects.filter(lab_number=_id)
                    if len(lis) <= 0:
                        return HttpResponse('订单只有打印后，才能补打!')

                # with transaction.atomic():
                # 添加VD1
                if lbo.vendor == '2' or lbo.vendor == '3' or lbo.vendor == '4' or lbo.vendor == '7' \
                        or lbo.vendor == '11' or lbo.vendor == '12' or lbo.vendor == '13' or lbo.vendor == '14' or \
                                lbo.vendor == '15' or lbo.vendor == '1':
                    cargo_location = iswc.get_location(lbo.frame, 'W02')
                    if cargo_location is None:
                        cargo_location = ''
                elif lbo.vendor == '9' or lbo.vendor == '10':
                    cargo_location = iswc.get_location(lbo.frame, 'W02')
                    if cargo_location is None:
                        cargo_location = ''

                cvc = construction_voucher_control()  # add 改变了lab_order的状态
                # add中包含了事务
                cv = cvc.add_print_times(_id, request.user, request.user.id, request.user.username, '')

                locker_obj_num = locker.get_locker_num(_id)
                # if locker_obj_num == '' and lbo.vendor <> '5':
                #     return HttpResponse('<h1>仓位中不存在该订单，请点击打印作业单！</h1>')
            logging.debug(locker_obj_num)
            # 获取打印数量
            c_voucher = construction_voucher.objects.filter(request_notes_entity=_lrn)
            if len(c_voucher) > 0:
                times = c_voucher[0].print_times

            newlbo = LabOrder.objects.get(lab_number=_id)
            # 生成一维码
            flag = utilities.generate_code128(newlbo)
            if flag:
                _item = LabOrder.objects.get(lab_number=_id)

            worknumber = _id.split("-")[1]
            com_date = lbo.estimated_date
            workshop = '车间-' + lbo.workshop
        # 预计完成时间为空时 使作业单正常打印
        if com_date is None or com_date == 'None' or com_date == '':
            com_date_str = ''
        else:
            com_date_str = com_date.strftime('%m') + com_date.strftime('%d')
        _obj_dict['times'] = format_str(times)
        _obj_dict['locker_obj_num'] = format_str(locker_obj_num)
        _obj_dict['worknumber'] = format_str(worknumber)
        _obj_dict['com_date'] = format_str(com_date_str)
        _obj_dict['workshop'] = format_str(workshop)
        _obj_dict['cargo_location'] = format_str(cargo_location)
        _form_data['obj'] = _obj_dict
        # 判断事务是否完成
        # 从操作记录中判断

        # 2019.10.06 by guof.
        # 暂时去掉由于流程跳跃，未打印作业单的订单；暂时允许直接补打.
        # trackings = OrderTracking.objects.filter(order_number=_id, action='PRINT_DATE')
        # if trackings.count() < 1:
        #     return HttpResponse("<h1 style="">请重新打印作业单</h1>")

        return render(request, "construction_voucher_print.html", {
            'form_data': _form_data,
            'item': _item,
            'requestUrl': '/oms/construction_voucher_print/',
            'lbo_id': lbo.id,
            'media_base_url': settings.SSH_MEDIA_SERVER.get('MEDIA_BASE_URL'),
        })



    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        _form_data['exceptions'] = e
        _form_data['error_message'] = str(e)
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


class status_choice:
    key = ''
    value = ''
    permission = ''
    switch = 0


# 成镜施工单
@login_required
@permission_required('oms.CVFG_VIEW', login_url='/oms/forbid/')
def redirect_construction_voucher_finished_glasses(request):
    '''

    :param request:
    :return:
    '''
    _form_data = {}
    _form_data['request_feature'] = 'Finished Glasses'
    _items = []
    _paginator = None
    _id = -1

    _id = request.GET.get('id', -1)
    logging.debug('id: %s' % _id)

    _form_data['id'] = _id
    try:
        if _id != -1:
            loc = lab_order_controller()
            lbos = loc.get_by_entity(_id)
            lbo = lbos[0]
            lab_number = lbo.lab_number

            if not lbo.status == 'LENS_RECEIVE' and not lbo.status == 'GLASSES_RETURN' and not lbo.status == 'COLLECTION' and not lbo.status == 'ASSEMBLING' and lbo.quantity == 1:
                _form_data['code'] = -1
                _form_data['message'] = "订单只有 镜片收货或成镜返工 状态，才能更改状态!当前订单状态为{0}".format(lbo.status)
                return render(request, "construction_voucher_finished_glasses.html",
                              {
                                  'form_data': _form_data,
                              })
            elif lbo.status == 'ASSEMBLING':
                _form_data['code'] = -1
                _form_data['message'] = "该订单已是待装配状态！"
                return render(request, "construction_voucher_finished_glasses.html",
                              {
                                  'form_data': _form_data,
                              })

            if lbo.is_production_change:
                _form_data['message'] = "制作参数已经修改，请补打印作业单,补打完成后订单已是待装配状态"
                _form_data['flag'] = 0
                _form_data['code'] = 0
            elif lbo.is_production_change == False:
                lbo.status = 'ASSEMBLING'
                lbo.save()
                tloc = tracking_lab_order_controller()
                tloc.tracking(lbo, request.user, 'ASSEMBLING')
                _form_data['code'] = 0
                _form_data['message'] = "操作成功！"

            # 移除仓位
            lc = locker_controller()
            lc.deleteItem(lab_number, request.user.username)

        return render(request, "construction_voucher_finished_glasses.html",
                      {
                          'form_data': _form_data,
                      })

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.CVFG_VIEW', login_url='/oms/forbid/')
def redirect_construction_voucher_finished_glasses_print(request):
    '''
    成镜施工单打印
    :param request:
    :return:
    '''
    _form_data = {}
    _form_data['request_feature'] = 'Finished Glasses'
    _items = []
    _paginator = None
    _id = -1

    _form_data['total'] = 0

    _id = request.GET.get('id', -1)
    _ws = request.GET.get('ws', -1)

    logging.debug('id: %s' % _id)

    _item = {}

    try:
        if _id <> -1:

            loc = lab_order_controller()
            lbos = loc.get_by_entity(_id)

            vs = loc.verify_status(_id)

            if len(lbos) == 1:
                lbo = lbos[0]
                _id = lbo.lab_number

            # _obj = LabOrder.objects.get(lab_number=id)
            _lrn = laborder_request_notes_line.objects.get(laborder_entity__lab_number=_id)

            _obj = _lrn.laborder_entity

            if _obj.status == 'REQUEST_NOTES':
                return HttpResponse('订单在 出库申请 状态，不能打印!需要先执行镜架出库!')

            # VD3 直接可以直接到成镜施工？
            if _obj.vendor == '3' or _obj.vendor == '5':
                pass
            else:
                if not _obj.status == 'LENS_RECEIVE' and not _obj.status == 'GLASSES_RETURN' and not _obj.status == 'ASSEMBLING' and _obj.quantity == 1:
                    return HttpResponse('订单只有 镜片收货或成镜返工 状态，才能打印!')

            if _ws == -1:
                return HttpResponse('请先指定车间编号!')

            with transaction.atomic():
                _obj.workshop = _ws
                _obj.save()

                _form_data['obj'] = _obj

                ca = int(time.time())
                ca = time.localtime(ca)
                ca_s = time.strftime("%Y-%m-%d %H:%M", ca)
                _form_data['created_at'] = ca_s

                _item = _obj

                cvc = construction_voucher_finish_glasses_control()  # add中开启了事务
                cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')

                lbo = LabOrder.objects.get(lab_number=_id)
                # 获取打印数量
                cvfg = construction_voucher_finish_glasses.objects.filter(request_notes_entity=_lrn)
                if len(cvfg) > 0:
                    times = cvfg[0].print_times
                # 生成一维码
                flag = utilities.generate_code128(lbo)

                if flag:
                    _item = LabOrder.objects.get(lab_number=_id)

        # 装配中的状态 允许重复打印
        if lbo.status == "ASSEMBLING":
            return render(request, "construction_voucher_finished_glasses_print.html", {
                'form_data': _form_data,
                'item': _item,
                'requestUrl': '/oms/construction_voucher_finished_glasses_print/',
                'lbo_id': lbo.id,
                'times': times,
                'media_base_url': settings.SSH_MEDIA_SERVER.get('MEDIA_BASE_URL')
            })
        else:
            return HttpResponse("<h1 style="">只有 [装配中] 状态的订单允许打印成镜施工单,当前状态 [%s]</h1>" % lbo.get_status_display())

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.CVFG_VIEW', login_url='/oms/forbid/')
def redirect_construction_voucher_finished_glasses_print_v3(request):
    '''
    成镜施工单打印 VD3 专版
    :param request:
    :return:
    '''
    _form_data = {}
    _form_data['request_feature'] = 'Finished Glasses'
    _items = []
    _paginator = None
    _id = -1

    _form_data['total'] = 0

    _id = request.GET.get('id', -1)
    _ws = request.GET.get('ws', -1)

    logging.debug('id: %s' % _id)

    _item = {}

    try:
        if _id <> -1:

            loc = lab_order_controller()
            lbos = loc.get_by_entity(_id)

            vs = loc.verify_status(_id)

            if len(lbos) == 1:
                lbo = lbos[0]
                _id = lbo.lab_number
            # _obj = LabOrder.objects.get(lab_number=id)
            _lrn = laborder_request_notes_line.objects.get(laborder_entity__lab_number=_id)

            _obj = _lrn.laborder_entity

            if _obj.status == 'REQUEST_NOTES':
                return HttpResponse('订单在 出库申请 状态，不能打印!需要先执行镜架出库!')

            if not _obj.vendor == '3':
                if not _obj.status == 'LENS_RECEIVE' and not _obj.status == 'GLASSES_RETURN' and not _obj.status == 'ASSEMBLING' and _obj.quantity == 1:
                    return HttpResponse('订单只有 镜片收货或成镜返工 状态，才能打印!')

            if _ws == -1:
                return HttpResponse('请先指定车间编号!')

            _form_data['obj'] = _obj

            with transaction.atomic():
                _obj.workshop = _ws
                _obj.save()

                ca = int(time.time())
                ca = time.localtime(ca)
                ca_s = time.strftime("%Y-%m-%d %H:%M", ca)
                _form_data['created_at'] = ca_s

                if float(_obj.od_sph) > 0:
                    _obj.od_sph = "+" + str(_obj.od_sph)

                if float(_obj.os_sph) > 0:
                    _obj.os_sph = "+" + str(_obj.os_sph)

                _item = _obj
                cvc = construction_voucher_finish_glasses_control()
                cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')
                lbo = LabOrder.objects.get(lab_number=_id)
                # 生成一维码
                flag = utilities.generate_code128(lbo)
                if flag:
                    _item = LabOrder.objects.get(lab_number=_id)

        if lbo.status == "ASSEMBLING":
            return render(request, "construction_voucher_finished_glasses_print_v3.html", {
                'form_data': _form_data,
                'item': _item,
                'requestUrl': '/oms/construction_voucher_finished_glasses_print_v3/',
                'lbo_id': lbo.id,
                'media_base_url': settings.SSH_MEDIA_SERVER.get('MEDIA_BASE_URL')
            })
        else:
            return HttpResponse("<h1 style="">请重新打印成镜施工单</h1>")

    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


class status_choice:
    key = ''
    value = ''


@login_required
@permission_required('oms.OLBO_ALL_VIEW', login_url='/oms/forbid/')
def redirect_laborder_list_v2(request):
    '''
    LAB Order 订单列表 V2
    :param request:
    :return:
    '''
    _form_data = {}
    _form_data['request_feature'] = 'All LAB Orders V2'

    order_number = request.GET.get('order_number', '')
    page = request.GET.get('page', 1)
    currentPage = int(page)
    filter = request.GET.get('filter', 'week')
    vendor = request.GET.get('vendor', 'all')
    ship = request.GET.get('ship', 'all')
    status = request.GET.get('status', 'all')
    id = request.GET.get('id', '')
    index = request.GET.get('index', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    sorted = request.GET.get('sorted', '-id')

    logging.debug(sorted)
    if sorted == 'set_time':
        logging.debug('sorted: %s' % sorted)
        filter = 'all'

    _form_data['sorted'] = sorted
    _form_data['id'] = id
    _form_data['index'] = index
    _form_data['filter'] = filter
    _form_data['ship'] = ship

    items = []
    paginator = None

    timedel = date_delta()  # N天前的日期
    timedel_week = date_delta_week()
    timedel_month = date_delta_month()

    # 状态列表
    STATUS_CHOICES = LabOrder.STATUS_CHOICES
    status_choices_list = []
    for sta in STATUS_CHOICES:
        sc = status_choice()
        sc.key = sta[0]
        sc.value = sta[1]
        status_choices_list.append(sc)

    # 发货方式列表
    SHIP_DIRECTION_CHOICES = LabOrder.SHIP_DIRECTION_CHOICES
    ship_direction_list = []
    for dir in SHIP_DIRECTION_CHOICES:
        sc = status_choice()
        sc.key = dir[0]
        sc.value = dir[1]
        ship_direction_list.append(sc)

    # VD列表
    vendors_choice_list = []
    for vcl in LabOrder.VENDOR_CHOICES:
        vc = status_choice()
        vc.key = vcl[0]
        vc.value = vcl[1]
        vendors_choice_list.append(vc)

    try:
        _filter = {}
        if order_number <> '':
            objs = []
            loc = lab_order_controller()
            objs = loc.get_by_entity(order_number)

            if len(objs) > 0:
                items = objs

        else:
            if id <> '' and index <> '':
                obj = LabOrder.objects.get(id=id)
                obj.vendor = index
                obj.save()

            # filter["status"] = 'PRINT_DATE'
            items = LabOrder.objects.filter(
                is_enabled=True,
                create_at__gte=timedel)

            if filter == 'week':
                items = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel_week)

            if filter == 'month':
                items = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel_month)

            if filter == 'all':
                items = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel)
            items = items.order_by(sorted)
            if filter == 'new':
                items = items.filter(**_filter).filter(Q(status='') | Q(status=None))
            if filter == '':
                items = items.filter(**_filter).filter(Q(status='') | Q(status=None))
            else:
                if not vendor == 'all':
                    _filter['vendor'] = vendor

                if not ship == 'all':
                    _filter['ship_direction'] = ship

                if not status == 'all':
                    _filter['status'] = status

                items = items.filter(**_filter)

        count = items.count

        if sorted == 'set_time':
            logging.debug('-------------------------------------')

            _filter['is_enabled'] = True
            _filter['create_at__gte'] = timedel
            _filter['set_time__lt'] = 0

            _filter_exceed_dimentions = {}
            _filter_exceed_dimentions['is_enabled'] = True
            _filter_exceed_dimentions['create_at__gte'] = timedel
            _filter_exceed_dimentions['set_time__lt'] = 0

            items_origin = LabOrder.objects.filter(**_filter_exceed_dimentions)
            logging.debug(items_origin.query)

            from report.models import exceed_laborders
            el = exceed_laborders()
            _form_data['exceed_dimentions'] = el.exceed_dimentions(items_origin)

            _form_data['vendor_dimentions'] = el.vendor_dimentions(items_origin)

            # items = LabOrder.objects.filter(**_filter) \
            #     .filter(Q(status='')
            #             | Q(status='REQUEST_NOTES')
            #             | Q(status='FRAME_OUTBOUND')
            #             | Q(status='PRINT_DATE')
            #             | Q(status='INITIAL_INSPECTION')
            #             | Q(status='LENS_RECEIVE')
            #             | Q(status='ASSEMBLING')
            #             | Q(status='FINAL_INSPECTION')
            #             | Q(status='PICKING')
            #             | Q(status='ORDER_MATCH')
            #             | Q(status='REDO')
            #             | Q(status='ONHOLD')
            #             | Q(status='REDO')
            #             )

            items = LabOrder.objects.filter(**_filter) \
                .filter(
                ~Q(status='SHIPPING'),
                ~Q(status='CANCELLED')
            )
            items.order_by('set_time')

            count = items.count()

        # 筛选重做单、转单和变更单
        ltype = request.GET.get('ltype', 'all')
        if not ltype == 'all':
            if ltype == 'c':
                items = items.filter(lab_number__contains='C')
            elif ltype == 'r':
                items = items.filter(lab_number__contains='R')
            elif ltype == 'z':
                items = items.filter(lab_number__contains='Z')
            else:
                items = items.filter(lab_number__contains='B')
            count = items.count()
        # 按时间区间筛选
        if not start_date == '':
            items = items.filter(create_at__range=(start_date, end_date))
            count = items.count()
        # 获取页码模板需要的参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))

        if query_string:
            query_string = '&' + query_string
        if count > 0:
            _form_data['total'] = count

        if sorted == 'set_time':
            paginator = Paginator(items, count)
        else:
            paginator = Paginator(items, const.PAGE_SIZE)  # Show 20 contacts per page

        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)
        print(items)




        return render(request, "laborder_list_v2.html",
                      {
                          'form_data': _form_data,
                          'list': items,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/redirect_laborder_list_v2/',
                          'filter': filter,
                          'vendor': vendor,
                          'status': status,
                          'status_choices': status_choices_list,
                          'vendors_choices': vendors_choice_list,
                          'ship_choices': ship_direction_list,
                          'ltype': ltype,
                          'query_string': query_string,
                          'start_date': start_date,
                          'end_date': end_date,



                      })
    except Exception as e:
        logging.debug('Exception: %s' % e)
        _form_data['exceptions'] = str(e)
        _form_data['error_message'] = str(e)
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.LO_EXP', login_url='/oms/forbid/')
def export_excel(request):
    entities = request.GET.get('entities', '')
    # items = []
    try:
        lbos = None
        # 前台选中的 entity id
        entities = request.GET.get('entities', False)
        if entities:
            arry_entities = entities.split(',')
            for entity in arry_entities:
                logging.debug(entity)
            lbos = LabOrder.objects.filter(id__in=arry_entities)
        else:  # 获取符合筛选条件的所有对象
            order_number = request.GET.get('order_number', '')
            filter = request.GET.get('filter', 'all')
            vendor = request.GET.get('vendor', 'all')
            ship = request.GET.get('ship', 'all')
            status = request.GET.get('status', 'all')
            start_date = request.GET.get('start_date', '')
            end_date = request.GET.get('end_date', '')
            sorted = request.GET.get('sorted', '-id')
            ltype = request.GET.get('ltype', 'all')
            if sorted == '':
                sorted = '-id'

            if order_number <> '':
                loc = lab_order_controller()
                objs = loc.get_by_entity(order_number)

                if len(objs) > 0:
                    lbos = objs
            else:
                _filter = {}
                timedel = date_delta()
                timedel_week = date_delta_week()
                timedel_month = date_delta_month()
                items = LabOrder.objects.filter(
                    is_enabled=True,
                    create_at__gte=timedel)

                if filter == 'week':
                    items = LabOrder.objects.filter(
                        is_enabled=True,
                        create_at__gte=timedel_week)

                if filter == 'month':
                    items = LabOrder.objects.filter(
                        is_enabled=True,
                        create_at__gte=timedel_month)

                if filter == 'all':
                    items = LabOrder.objects.filter(
                        is_enabled=True,
                        create_at__gte=timedel)
                print ('999')
                items = items.order_by(sorted)

                if filter == 'new':
                    items = items.filter(**_filter).filter(Q(status='') | Q(status=None))
                if filter == '':
                    items = items.filter(**_filter).filter(Q(status='') | Q(status=None))
                else:
                    if not vendor == 'all':
                        _filter['vendor'] = vendor

                    if not ship == 'all':
                        _filter['ship_direction'] = ship

                    if not status == 'all':
                        _filter['status'] = status

                    items = items.filter(**_filter)
                if sorted == 'set_time':
                    logging.debug('-------------------------------------')

                    _filter['is_enabled'] = True
                    _filter['create_at__gte'] = timedel
                    _filter['set_time__lt'] = 0

                    _filter_exceed_dimentions = {}
                    _filter_exceed_dimentions['is_enabled'] = True
                    _filter_exceed_dimentions['create_at__gte'] = timedel
                    _filter_exceed_dimentions['set_time__lt'] = 0

                    items_origin = LabOrder.objects.filter(**_filter_exceed_dimentions)
                    logging.debug(items_origin.query)

                    from report.models import exceed_laborders
                    el = exceed_laborders()
                    _form_data['exceed_dimentions'] = el.exceed_dimentions(items_origin)

                    _form_data['vendor_dimentions'] = el.vendor_dimentions(items_origin)

                    items = LabOrder.objects.filter(**_filter) \
                        .filter(
                        ~Q(status='SHIPPING'),
                        ~Q(status='CANCELLED')
                    )
                    items.order_by('set_time')
                if not ltype == 'all':
                    if ltype == 'c':
                        items = items.filter(lab_number__contains='C')
                    elif ltype == 'r':
                        items = items.filter(lab_number__contains='R')
                    elif ltype == 'z':
                        items = items.filter(lab_number__contains='Z')
                    else:
                        items = items.filter(lab_number__contains='B')
                if not start_date == '':
                    items = items.filter(create_at__range=(start_date, end_date))
                # 筛选条件生成列表
                lbos = items

        import csv, codecs
        # for lbo in lbos:
        #   items.append(lbo)

        # logging.debug('items count: %s' % len(items))

        response = HttpResponse(content_type='text/csv')
        file_name = 'redirect_laborder_list'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'
        response.write(codecs.BOM_UTF8)

        writer = csv.writer(response)
        # 在下面添加要导出的属性即可
        writer.writerow([
            '#', '实际发货', '订单号', '镜架', '镜框',
            '数量', '镜片', '高散', '染色', '镜片类型',
            '订单日期', '下达日期', '更新时间', '生产天数', '预计完成', '规定完成',
            '距离规定', '状态', 'VIP', 'VD', 'WS', '内部备注'
        ])
        # count = len(items)
        # index = 0
        for item in lbos:
            print((item.create_at + datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S"))
            if item.estimated_date:
                estimated_date = (item.estimated_date + datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S")
            else:
                estimated_date = ''

            if item.targeted_date:
                targeted_date = (item.targeted_date + datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S")
            else:
                targeted_date = ''

            writer.writerow([
                item.id, item.get_act_ship_direction_display(), item.lab_number, item.frame, item.frame_type,
                item.quantity, item.act_lens_name, item.is_cyl_high, item.tint_name, item.lens_type,
                item.order_date, (item.create_at + datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S"), (item.update_at + datetime.timedelta(hours=8)).strftime("%Y-%m-%d %H:%M:%S"), item.days_of_production, estimated_date,
                targeted_date,
                item.set_time_1, item.get_status_display(), item.is_vip, item.vendor, item.workshop, item.comments_inner

            ])

        return response
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        return HttpResponse('生成工厂订单遇到异常[ %s ], 请暂时手动生成，并联系系统支持....' % e.message)


@login_required
@permission_required('oms.OLBO_ALL_VIEW', login_url='/oms/forbid/')
def redirect_laborder_detail_v2(request):
    '''
    LAB Order 订单详情 V2
    :param request:
    :return:
    '''
    _form_data = {}
    _items = []
    _paginator = None
    _id = -1

    _form_data['total'] = 0

    _id = request.GET.get('id', -1)

    logging.debug('id: %s' % _id)

    _item = {}

    try:
        if _id <> -1:
            logging.debug('开始')
            # _obj = LabOrder.objects.get(lab_number=id)
            _lrn = laborder_request_notes_line.objects.get(laborder_entity__lab_number=_id)

            _obj = _lrn.laborder_entity

            _form_data['obj'] = _obj
            _form_data['created_at'] = timezone.now()

            _item = _obj

            cvc = construction_voucher_control()
            cv = cvc.add(_id, request.user, request.user.id, request.user.username, '')

        return render(request, "laborder_detail_v2.html",
                      {
                          'form_data': _form_data,
                          'item': _item,
                          'requestUrl': '/oms/redirect_laborder_detail_v2/',
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.OLBO_ALL_VIEW', login_url='/oms/forbid/')
def redirect_laborder_detail_pspf(request):
    '''
    订单详情
    :param request:
    :return:
    '''

    try:

        vendors_choice_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            # if int(vc.key) > 0 and int(vc.key) <= 10:
            vendors_choice_list.append(vc)

        lab_number = request.POST.get("lab_number")
        laborder = LabOrder.objects.get(lab_number=lab_number)

        lbo = LabOrder.objects.get(lab_number=lab_number)
        ot = tracking_lab_order_controller()
        rm = ot.get_order_history(lbo.lab_number)
        gfits = glasses_final_inspection_technique.objects.filter(laborder_id=lbo.id)
        gfit_id = ''
        if gfits.count() > 0:
            gfit_id = gfits[0].id

        tracking = rm.obj

        data = {}
        data['vendor'] = lbo.vendor
        data['order_number'] = lbo.lab_number
        tracking_vendor = None

        if lbo.vendor == '4':
            wosc = WcOrderStatusController()
            rm = wosc.get_order_history(request, data)
            tracking_vendor = rm.obj
        elif lbo.vendor == '9':
            wxosc = WxOrderStatusController()
            rm = wxosc.get_order_history(request, data)
            tracking_vendor = rm.obj

        # 判断类型
        category_name = ''
        category_id = lbo.category_id
        if category_id == CategoryType.Woman.value:
            category_name = '女款光学镜'
        elif category_id == CategoryType.Man.value:
            category_name = '男款光学镜'
        elif category_id == CategoryType.Child.value:
            category_name = '儿童款光学镜'
        elif category_id == CategoryType.WomanSunGlasses.value:
            category_name = '女款太阳镜'
        elif category_id == CategoryType.ManSunGlasses.value:
            category_name = '男款太阳镜'
        elif category_id == CategoryType.ChildSunGlasses.value:
            category_name = '儿童款太阳镜'

        if request.method == 'GET':
            return render(request, 'laborder_detail_v2.pspf.html',
                          {
                              'vendors_choices': vendors_choice_list,
                          }
                          )
        elif request.method == 'POST':
            obj = lbo

            return render(request, "laborder_detail_v2.pspf.html",
                          {
                              'item': lbo,
                              'tracking': tracking,
                              'tracking_vendor': tracking_vendor,
                              'vendors_choices': vendors_choice_list,
                              'category_name': category_name,
                              'gfit_id': gfit_id
                          })
    except Exception as e:
        logging.debug(e)
        return HttpResponse('No Data')


@login_required
@permission_required('oms.OLBO_ALL_VIEW', login_url='/oms/forbid/')
def redirect_laborder_detail(request, parameters=''):
    '''
    订单详情
    :param request:
    :return:
    '''

    try:
        lab_number = parameters
        logging.debug(lab_number)
        lbo = LabOrder.objects.get(lab_number=lab_number)
        ot = tracking_lab_order_controller()
        rm = ot.get_order_history(lbo.lab_number)
        tracking = rm.obj

        if request.method == 'GET':
            return render(request, 'laborder_detail_v2.pspf.html',
                          {
                              'item': lbo,
                              'tracking': tracking
                          })
        elif request.method == 'POST':
            logging.debug('----------------------------------------')
            return render(request, "laborder_detail_v2.pspf.html",
                          {
                              'item': lbo,
                              'tracking': tracking
                          })
    except Exception as e:
        logging.debug(e.message)
        return HttpResponse('No Data')


# 单独添加一个权限
@login_required
@permission_required('oms.CHANGE_ORDER_VALUE', login_url='/oms/forbid/')
def ajax_edit_act_lens(request):
    rm = {'code': '0', 'message': ''}
    m_cur_lab = request.POST.get('cur_lab')
    m_field = request.POST.get('field')
    m_value = request.POST.get('value')
    logging.debug('=========>%s' % m_cur_lab)
    logging.debug('=========>%s: %s' % (m_field, m_value))

    try:
        count = 0
        status_tuple = (
            None,
            '',
            'REQUEST_NOTES',
            'FRAME_OUTBOUND',
            'PRINT_DATE',  # 镜片生产之后修改需重新打印作业单
            'LENS_OUTBOUND',
            'LENS_REGISTRATION',
            'LENS_RETURN',
            'LENS_RECEIVE'
        )
        lbos = LabOrder.objects.filter(lab_number=m_cur_lab)
        lbo = lbos[0]
        cur_status = lbo.status
        last_status = lbo.current_status
        if lbos.exists() and (cur_status == 'R2HOLD' or cur_status == 'ONHOLD'):
            if last_status in status_tuple:
                tloc = tracking_lab_order_controller()
                # 框型
                if 'frame_type' == m_field:
                    # 原值
                    old_frame_type = lbo.frame_type
                    # 更新
                    count = lbos.update(frame_type=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改框型", '原值=' + old_frame_type)
                # 尺寸
                if 'size' == m_field:
                    # 原值
                    old_size = lbo.size
                    # 更新
                    count = lbos.update(size=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改尺寸", '原值=' + old_size)
                # 框宽
                if 'lens_width' == m_field:
                    # 原值
                    old_lens_width = lbo.lens_width
                    # 更新
                    count = lbos.update(lens_width=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改框宽", '原值=' + str(old_lens_width))
                # 框高
                if 'lens_height' == m_field:
                    # 原值
                    old_lens_height = lbo.lens_height
                    # 更新
                    count = lbos.update(lens_height=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改框高", '原值=' + str(old_lens_height))
                # 订单镜片SKU
                if 'lens_sku' == m_field:
                    # 原值
                    old_lens_sku = lbo.lens_sku
                    # 查询镜片名称
                    lens_names = LabProduct.objects.filter(sku=m_value)
                    if lens_names.count() < 1:
                        rm['code'] = '-1'
                        rm['message'] = 'SKU错误'
                        return JsonResponse(rm)
                    lens_name = lens_names[0]
                    # 更新
                    count = lbos.update(lens_sku=m_value, lens_name=lens_name.name)
                    # 如果已经分单，同步更新lens_order表
                    if not lbo.vendor == '0':
                        los = lens_order.objects.filter(lab_number=lbo.lab_number)
                        count = los.update(lens_sku=m_value, lens_name=lens_name.name)
                    tloc.tracking(lbo, request.user, m_value, "修改订单镜片SKU", '原值=' + old_lens_sku)
                # 实际镜片SKU
                if 'act_lens_sku' == m_field:
                    # 原值
                    old_act_lens_sku = lbo.act_lens_sku
                    # 查询镜片名称
                    lens_names = LabProduct.objects.filter(sku=m_value)
                    if lens_names.count() < 1:
                        rm['code'] = '-1'
                        rm['message'] = 'SKU错误'
                        return JsonResponse(rm)
                    lens_name = lens_names[0]
                    # 更新
                    count = lbos.update(act_lens_sku=m_value, act_lens_name=lens_name.name)
                    # 如果已经分单，同步更新lens_order表
                    if not lbo.vendor == '0':
                        los = lens_order.objects.filter(lab_number=lbo.lab_number)
                        count = los.update(act_lens_sku=m_value, act_lens_name=lens_name.name)
                    tloc.tracking(lbo, request.user, m_value, "修改实际镜片SKU", '原值=' + old_act_lens_sku)
                # ------------------由SKU关联更新-------------------------
                # # 订单镜片名称
                # if 'lens_name' == m_field:
                #     # 原值
                #     old_lens_name = lbo.lens_name
                #     # 更新
                #     count = lbos.update(lens_name=m_value)
                #     # 如果已经分单，同步更新lens_order表
                #     if not lbo.vendor == '0':
                #         los = lens_order.objects.filter(lab_number=lbo.lab_number)
                #         count = los.update(lens_name=m_value)
                #     tloc.tracking(lbo, request.user, m_value, "修改订单镜片", '原值=' + old_lens_name)
                # # 实际镜片名称
                # if 'act_lens_name' == m_field:
                #     # 原值
                #     old_act_lens_name = lbo.act_lens_name
                #     # 更新
                #     count = lbos.update(act_lens_name=m_value)
                #     # 如果已经分单，同步更新lens_order表
                #     if not lbo.vendor == '0':
                #         los = lens_order.objects.filter(lab_number=lbo.lab_number)
                #         count = los.update(act_lens_name=m_value)
                #     tloc.tracking(lbo, request.user, m_value, "修改实际镜片", '原值=' + old_act_lens_name)
                # 染色
                # ------------------由SKU关联更新-------------------------
                if 'tint_name' == m_field:
                    # 原值
                    old_tint_name = lbo.tint_name
                    # 更新
                    count = lbos.update(tint_name=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改染色", '原值=' + old_tint_name)
                # 染色SKU
                if 'tint_sku' == m_field:
                    # 原值
                    old_tint_sku = lbo.tint_sku
                    # 更新
                    count = lbos.update(tint_sku=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改染色SKU", '原值=' + old_tint_sku)
                # 设计
                if 'pal_design_name' == m_field:
                    # 原值
                    old_pal_design_name = lbo.pal_design_name
                    # 更新
                    count = lbos.update(pal_design_name=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改设计", '原值=' + old_pal_design_name)
                # 设计SKU
                if 'pal_design_sku' == m_field:
                    # 原值
                    old_pal_design_sku = lbo.pal_design_sku
                    # 更新
                    count = lbos.update(pal_design_sku=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改设计SKU", '原值=' + old_pal_design_sku)
                # 涂层
                if 'coating_name' == m_field:
                    # 原值
                    old_coating_name = lbo.coating_name
                    # 更新
                    count = lbos.update(coating_name=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改涂层", '原值=' + old_coating_name)
                # 涂层SKU
                if 'coating_sku' == m_field:
                    # 原值
                    old_coating_sku = lbo.coating_sku
                    # 更新
                    count = lbos.update(coating_sku=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改涂层SKU", '原值=' + old_coating_sku)
                # 加工瞳高
                if 'lab_seg_height' == m_field:
                    # 原值
                    old_lab_seg_height = lbo.lab_seg_height
                    # 更新
                    count = lbos.update(lab_seg_height=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改加工瞳高", '原值=' + old_lab_seg_height)
                # 装配瞳高
                if 'assemble_height' == m_field:
                    # 原值
                    old_assemble_height = lbo.assemble_height
                    # 更新
                    count = lbos.update(assemble_height=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改装配瞳高", '原值=' + old_assemble_height)
                # 子镜高度
                if 'sub_mirrors_height' == m_field:
                    # 原值
                    old_sub_mirrors_height = lbo.sub_mirrors_height
                    # 更新
                    count = lbos.update(sub_mirrors_height=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改子镜高度", '原值=' + old_sub_mirrors_height)
                # 加工要求
                if 'special_handling' == m_field:
                    # 原值
                    old_special_handling = lbo.special_handling
                    # 更新
                    count = lbos.update(special_handling=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改加工要求", '原值=' + old_special_handling)
                # 发运方式
                if 'ship_direction' == m_field:
                    # 原值
                    old_ship_direction = lbo.ship_direction
                    # 更新
                    count = lbos.update(ship_direction=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改发运方式", '原值=' + old_ship_direction)
                # 实际发运方式
                if 'act_ship_direction' == m_field:
                    # 原值
                    old_act_ship_direction = lbo.act_ship_direction
                    # 更新
                    count = lbos.update(act_ship_direction=m_value)
                    tloc.tracking(lbo, request.user, m_value, "修改实际发运方式", '原值=' + old_act_ship_direction)
            else:
                rm['code'] = '1'
                rm['message'] = '待装配前状态才可以修改订单信息,暂停前状态【%s】' % last_status
        else:
            rm['code'] = '1'
            rm['message'] = '请先申请暂停再修改订单信息,当前状态【%s】' % cur_status

        # 经片生产之后的状态设置重新打印作业单
        if last_status in ('PRINT_DATE', 'LENS_OUTBOUND', 'LENS_REGISTRATION', 'LENS_RETURN', 'LENS_RECEIVE'):
            lbos.update(is_production_change=True)

    except Exception as e:
        logging.debug('err=============>%s' % e)
        rm['code'] = '-1'
        rm['message'] = str(e)

    return JsonResponse(rm)


@login_required
@permission_required('oms.LAB_DIST_VENDOR', login_url='/oms/forbid/')
def distribute_vendor(request):
    '''
    分配供应商 VD
    :param request:
    :return:
    '''
    logging.debug('----------------------------------------')
    order_entities = request.POST.get('entities', 0)
    index = request.POST.get('index', 0)
    rm = res_msg.response_dict()

    logging.debug('order_entities: %s' % order_entities)
    logging.debug('index: %s' % index)

    try:
        entities = order_entities
        lbo = LabOrder.objects.get(id=entities)
        if lbo.status == "" or lbo.status == None:
            origin_value = lbo.vendor
            lbo.vendor = index
            lbo.save()

            from api.controllers.tracking_controllers import tracking_operation_controller
            tc = tracking_operation_controller()
            tc.tracking(
                lbo.type, lbo.id,
                lbo.lab_number,
                "分配VD",
                "vendor",
                request.user,
                origin_value,
                index,
                None,
                None,

            )
            rm["message"] = "已成功分配VD"
            return JsonResponse(rm)
        else:
            rm["code"] = 4
            rm["message"] = "只有新订单可以分配VD"
            return JsonResponse(rm)
    except Exception as e:
        logging.debug(e.message)
        rm["code"] = -1
        rm["message"] = "分配VD异常"
        return JsonResponse(rm)


from oms.models.glasses_models import received_glasses_control


@login_required
@permission_required('oms.REFG_VIEW', login_url='/oms/forbid/')
def redirect_received_glasses(request):
    '''
    成镜收货单
    :param request:
    :return:
    '''

    rm = res_msg()
    _form_data = {}

    _form_data['request_feature'] = 'Received Glasses'
    items = []
    lbo = None
    try:
        if request.method == 'POST':
            res = {}
            lab_number = request.POST.get('lab_nubmer', '')

            if lab_number == '':
                res['code'] = -1
                res['message'] = '请输入订单号!!'
                return HttpResponse(json.dumps(res))

            try:
                logging.debug('----------------------------------------')
                rgc = received_glasses_control()
                # add 中已做事务处理
                rm = rgc.add(
                    request,
                    lab_number
                )

                res['code'] = rm.code
                res['message'] = rm.message

                logging.debug('----------------------------------------')

            except Exception as e:
                res['code'] = -999
                res['message'] = '数据遇到异常: ' + e.message

            return HttpResponse(json.dumps(res))

        entity_id = request.GET.get('entity_id', '')
        # _form_data["search_entity"] = entity_id

        if not entity_id == '':
            loc = lab_order_controller()
            lbos = loc.get_by_entity(entity_id)

            if len(lbos) == 1:
                lbo = lbos[0]
                lab_number = lbo.lab_number

            # 用来到wms_product_frame中查找sku_specs字段（警示信息）
            caution_info = product_frame.objects.get(sku=lbo.frame)
            _form_data['caution_info'] = caution_info.sku_specs

            lbo = LabOrder.objects.get(lab_number=lab_number)
            _form_data['laborder'] = lbo

        return render(request, "received_glasses.html",
                      {
                          'form_data': _form_data,
                          'requestUrl': reverse('received_glasses'),
                          'item': lbo,
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                          'requestUrl': reverse('received_glasses'),
                      })


# 成镜收货清单
@login_required
@permission_required('oms.REFG_VIEW', login_url='/oms/forbid/')
def redirect_received_glasses_list(request):
    # 定义变量
    rm = response_message()
    _form_data = {}
    # 获取参数
    lab_number = request.GET.get('lab_number', '')
    _form_data['lab_number'] = lab_number
    start_time = request.GET.get('start_time', '')
    _form_data['start_time'] = start_time
    end_time = request.GET.get('end_time', '')
    _form_data['end_time'] = end_time
    page = request.GET.get('page', 1)
    query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
    if query_string:
        query_string = '&' + query_string
    # 条件搜索
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    where = 'where 1= 1 '
    # 开始时间不为空
    if start_time == '':
        where += ' and created_at >= "%s" ' % today
        _form_data['start_time'] = today
    else:
        where += ' and created_at >= "%s" ' % start_time
    # 结束时间不为空
    if not end_time == '':
        where += ' and created_at <= "%s" ' % end_time
    try:
        sql = """
            SELECT id,lab_number,user_name,created_at,STATUS,comments FROM oms_received_glasses
            %s
            order by id desc
        """ % where
        logging.debug(sql)
        with connections['pg_oms_query'].cursor() as cursor:
            cursor.execute(sql)
            received_list = namedtuplefetchall(cursor)

        _form_data['total'] = len(received_list)
        paginator = Paginator(received_list, 20)  # Show 20 contacts per page

        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)

        return render(request, "received_glasses_list.html",
                      {
                          "list": items,
                          'form_data': _form_data,
                          'paginator': paginator,
                          'requestUrl': '/oms/received_glasses_list/',
                          'filter': filter,
                          'query_string': query_string,
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                          'requestUrl': reverse('received_glasses_list'),
                      })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def get_prescription(request):
    rm = res_msg()
    m_item_id = request.POST.get('m_item_id')
    sql = sql_query_prescription % m_item_id
    with connections['pg_mg_query'].cursor() as cursor:
        cursor.execute(sql)
        results = namedtuplefetchall(cursor)

        rm.obj = {}
        for i in range(len(results)):
            rm.obj['rsph'] = float(results[i].rsph)
            rm.obj['lsph'] = float(results[i].lsph)
            rm.obj['rcyl'] = float(results[i].rcyl)
            rm.obj['lcyl'] = float(results[i].lcyl)
            rm.obj['rax'] = float(results[i].rax)
            rm.obj['lax'] = float(results[i].lax)

            rm.obj['radd'] = float(results[i].radd)
            rm.obj['ladd'] = float(results[i].ladd)
            rm.obj['pd'] = float(results[i].pd)
            rm.obj['single_pd'] = float(results[i].single_pd)
            rm.obj['rpd'] = float(results[i].rpd)
            rm.obj['lpd'] = float(results[i].lpd)

            rm.obj['rpri'] = float(results[i].rpri)
            rm.obj['lpri'] = float(results[i].lpri)
            rm.obj['rbase'] = results[i].rbase
            rm.obj['lbase'] = results[i].lbase

            # 字段没有确定 这里的字段直接取的 megento的字段名 字段名更改可能会直接影响到这里
            rm.obj['rpri1'] = float(results[i].rpri_1)
            rm.obj['lpri1'] = float(results[i].lpri_1)
            rm.obj['rbase1'] = results[i].rbase_1
            rm.obj['lbase1'] = results[i].lbase_1

            # 如果查不到可以通过product_options中的内容查
            # rm.obj['product_options'] = results[i].product_options

    return render(request, 'primitive_prescription.html', {
        'item': rm.obj,
    })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def get_mg_ed(request):
    # 获取 ED
    rm = res_msg.response_dict()
    item_id = request.POST.get('item_id')
    try:
        item_frame = PgOrderItem.objects.values('frame').get(item_id=item_id)
        frame_sku = ''
        m_data = json.dumps({'sku': item_frame['frame']})
        m_req = urllib2.Request(
            url=MG_ED_URL,
            data=m_data,
            headers={'Content-Type': 'application/json'}
        )
        m_res = urllib2.urlopen(m_req)
        resp = m_res.read()
        rm['obj'] = json.loads(resp)

    except Exception as e:
        rm['code'] = '-1'
        rm['message'] = 'ERR===>%s' % e

    return JsonResponse(rm)


@login_required
@permission_required('oms.PG_REORDER', login_url='/oms/forbid/')
def re_order_v3(request):
    req_info = {}
    rm = res_msg.response_dict()
    order_number = request.POST.get('order_number', '')
    redo_list = request.POST.getlist('redo_list[]', [])
    if len(redo_list) == 0:
        rm['code'] = '-2'
        rm['message'] = ' 未指定 PG Order Item.'
        return JsonResponse(rm)

    try:
        pg_order = PgOrder.objects.only('id', 'is_inlab', 'reorder_number').get(order_number=order_number)
        if not pg_order.is_inlab:
            rm['code'] = '-2'
            rm['message'] = 'is_inlab is False.'
            return JsonResponse(rm)

        # 拼接数据
        req_info = json.dumps({
            "order_id": pg_order.id,
            "order_number": order_number,
            "item_id": redo_list,
            "is_redo": "1"
        })

        # getToken()

        # 请求接口 重做
        api_response = requests.post(PG_SYNC_REORDER_URL, data=req_info, headers=token_header, timeout=5)
        res_json = json.loads(api_response.text)
        stat = res_json.get('success', '-1')
        if stat == '1':
            # re_order成功
            with transaction.atomic():
                tloc = tracking_lab_order_controller()
                for item in redo_list:
                    pgoi = PgOrderItem.objects.only('id', 'comments').get(item_id=item)

                    # 如果生成工厂订单 要暂停工厂订单
                    if pg_order.is_inlab:
                        lbos = LabOrder.objects.filter(base_entity=pgoi.id).order_by('create_at')
                        # 如果订单时已发货状态则不改变状态
                        max_len = lbos.count()
                        logging.debug(lbos[max_len - 1].lab_number)
                        if not lbos[max_len - 1].status == 'SHIPPING':
                            logging.debug('暂停')
                            lbo_ctrl = lab_order_controller(order_query=lbos)
                            lbo_ctrl.hold_orders(tloc, '订单重做 系统暂停')

                    # 重做成功增加标识
                    pgoi.comments = '该订单子项已重做;%s' % pgoi.comments
                    pgoi.save()

                pg_order.reorder_number = '%s%s;' % (pg_order.reorder_number, res_json.get('data').get('increment_id'))
                pg_order.save()

            rm['obj'] = res_json.get('data')
            # 添加日志
            order_activity = OrderActivity()
            order_activity.add_activity('OPOR', pgoi.pg_order_entity.id, order_number, '', request.user.id,
                                        request.user.username, 'PG Order ReOrder...')

        elif stat in ('0', '-1'):
            # 无库存等情况
            rm['code'] = '-2'
            rm['message'] = res_json.get('message')
        else:
            # 未知情况 超时等情况
            rm['code'] = '-1'
            rm['obj'] = res_json.get('messages').get('error')

    except Exception as e:
        logging.debug(e)
        return HttpResponse(str(e))

    return JsonResponse(rm)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def edit_channel(request):
    rm = res_msg.response_dict()
    is_get_type = request.POST.get('is_get_type', '')
    channel_type = request.POST.get('channel', '')
    item_id = request.POST.get('item_id', None)
    product_index = request.POST.get('product_index', None)

    if not is_get_type == '':
        return JsonResponse(dict(LabOrder.CHANNEL_CHOICES))

    try:
        with transaction.atomic():
            pgoi = PgOrderItem.objects.get(item_id=item_id, product_index=product_index)
            pgoi.channel = channel_type
            pgoi.save()
            lbos = LabOrder.objects.filter(base_entity=pgoi.id)
            if len(lbos) > 0:
                for lbo in lbos:
                    lbo.channel = channel_type
                    lbo.save()

            rm['message'] = channel_type

    except Exception as e:
        logging.debug(str(e))
        rm['code'] = '-1'
        return JsonResponse(rm)

    return JsonResponse(rm)


def pgOrderSetInLab(request):
    try:
        rm = res_msg.response_dict()
        if request.method == 'POST':
            order_number = request.POST.get("order_number")
            ocresponse = order_controller.PgOrderController().set_to_inlab(order_number)
            ocresponse_json = json.loads(ocresponse)
            if ocresponse_json['code'] == 0:
                rm['code'] = '0'
                rm['message'] = '执行成功'
                return JsonResponse(rm)
            rm['code'] = '-1'
            rm['message'] = '执行失败'
            return JsonResponse(rm)
    except Exception as e:
        logging.debug(str(e))
        rm['code'] = '-1'
        rm['message'] = str(e)
        return JsonResponse(rm)


# 更改订单状态的申请
@login_required
@permission_required('oms.HOLD_CANCEL_REQUEST', login_url='/oms/forbid/')
def hold_cancel_request_list(request):
    rm = response_message()
    _page_info = {}
    _item = None
    action_handel = '通过暂停申请'
    action = '暂停'
    # post请求
    if request.method == 'POST':
        rm_js = {}
        logging.debug('进入POST')
        application_id = request.POST.get('application_id', '')
        reply_text = request.POST.get('reply_text', '')
        handle_result = request.POST.get('handle_result', '')
        try:
            # 获取实体
            application_entitys = hold_cancel_request.objects.filter(id=application_id)
            application_entity = application_entitys[0]
            user_entity = User.objects.get(id=application_entity.user_id)
            lab_order_entity = LabOrder.objects.get(lab_number=application_entity.lab_number)
            tloc = tracking_lab_order_controller()
            if handle_result == 'ALLOW':

                if not lab_order_entity.status == 'R2CANCEL' and not lab_order_entity.status == 'R2HOLD':
                    logging.debug('订单不是申请状态')
                    rm_js['code'] = '-1'
                    rm_js['message'] = '订单不是申请状态'
                    rm_js = json.dumps(rm_js)
                    return HttpResponse(rm_js)
                # 如果是取消单,调用cancel_order（）方法和，update_inventory_qty（）方法
                if application_entity.order_status_future == 'CANCELLED':
                    action_handel = '通过取消申请'
                    action = '取消'
                    request_dict = {
                        'user': {
                            'id': application_entity.user_id,
                            'username': application_entity.user_name
                        }
                    }
                    request_obj = DictToObj(request_dict)

                    # 重做单取消原单不更新库存
                    if application_entity.reason == '重做订单取消原单(自动申请)' or application_entity.reason == '转单取消原单(自动申请)':
                        logging.debug('下重做单取消，不更新库存')
                        cancel_order(request, lab_order_entity)
                    # 其它取消方式更新库存再取消
                    else:
                        rst_data = update_inventory_qty(request_obj, lab_order_entity)

                        if rst_data.code == 0:
                            logging.debug('更新库存成功')
                            cancel_order(request, lab_order_entity)
                        else:
                            logging.debug('更新库存失败')
                            rm_js['code'] = '-1'
                            rm_js['message'] = rst_data.message
                            rm_js = json.dumps(rm_js)
                            return HttpResponse(rm_js)

                # 更新申请单
                application_entity.reply = reply_text
                application_entity.handle_result = handle_result
                application_entity.is_handle = True
                application_entity.reply_username = request.user.username
                application_entity.save()

                # 写处理申请TRACKING
                rm_r = tloc.tracking(lab_order_entity, request.user, application_entity.order_status_future,
                                     action_handel, reply_text)
                rm_t = None
                if rm_r.code == 0:
                    # 写变更状态TRACKING
                    rm_t = tloc.tracking(lab_order_entity, user_entity, application_entity.order_status_future, action,
                                         application_entity.reason)
                else:
                    logging.debug('写tracking失败')
                    rm_js['code'] = '-1'
                    rm_js['message'] = rm_r.message
                    rm_js = json.dumps(rm_js)
                    return HttpResponse(rm_js)

                # 更新lab_order状态
                if rm_t.code == 0:
                    # 获取需要写入lab_order的数据
                    status = application_entity.order_status_future
                    current_status = application_entity.order_status_now
                    comments_inner = lab_order_entity.comments_inner + ';' + application_entity.reason
                    lab_number = application_entity.lab_number
                    with connections['default'].cursor() as cursor:
                        update_sql = '''
                            UPDATE oms_laborder SET `status`='%s', `current_status`='%s', `comments_inner`='%s'WHERE lab_number='%s'
                        ''' % (status, current_status, comments_inner, lab_number)
                        cursor.execute(update_sql)
                else:
                    logging.debug('写tracking失败2')
                    rm_js['code'] = '-1'
                    rm_js['message'] = rm_t.message
                    rm_js = json.dumps(rm_js)
                    return HttpResponse(rm_js)

            # 关闭订单
            elif handle_result == 'CLOSE_ORDER':
                action_handel = '关闭订单'
                # 更新申请单
                application_entity.reply = reply_text
                application_entity.handle_result = handle_result
                application_entity.is_handle = True
                application_entity.reply_username = request.user.username
                application_entity.save()

                # 写TRACKING
                rm_r = tloc.tracking(lab_order_entity, request.user, application_entity.order_status_future,
                                     action_handel, reply_text)

                # 更新lab_order状态
                if rm_r.code == 0:
                    # 获取需要写入lab_order的数据
                    status = 'CLOSED'
                    current_status = application_entity.order_status_now
                    comments_inner = lab_order_entity.comments_inner + ';' + application_entity.reason
                    lab_number = application_entity.lab_number
                    with connections['default'].cursor() as cursor:
                        update_sql = '''
                                    UPDATE oms_laborder SET `is_enabled`=False,`status`='%s', `current_status`='%s', `comments_inner`='%s'WHERE lab_number='%s'
                                    ''' % (status, current_status, comments_inner, lab_number)
                        cursor.execute(update_sql)
                else:
                    logging.debug('写tracking失败')
                    rm_js['code'] = '-1'
                    rm_js['message'] = rm_r.message
                    rm_js = json.dumps(rm_js)
                    return HttpResponse(rm_js)

            # 不通过,写入申请单，将订单的申请状态改回原来状态 并写TRACKING
            else:
                if application_entity.order_status_future == 'CANCELLED':
                    action_handel = '拒绝取消申请'
                else:
                    action_handel = '拒绝暂停申请'
                # 更新申请单
                application_entity.reply = reply_text
                application_entity.handle_result = handle_result
                application_entity.is_handle = True
                application_entity.reply_username = request.user.username
                application_entity.save()
                # 写tracking
                tloc = tracking_lab_order_controller()
                logging.debug("写tracking")
                rm_t = tloc.tracking(lab_order_entity, request.user, application_entity.order_status_future,
                                     action_handel,
                                     reply_text)
                if rm_t.code == 0:
                    # 获取需要写入lab_order的数据
                    status = application_entity.order_status_now
                    current_status = ''
                    comments_inner = lab_order_entity.comments_inner + ';' + reply_text
                    lab_number = application_entity.lab_number
                    with connections['default'].cursor() as cursor:
                        update_sql = '''
                            UPDATE oms_laborder SET `status`='%s', `current_status`='%s', `comments_inner`='%s'WHERE lab_number='%s'
                         ''' % (status, current_status, comments_inner, lab_number)
                        cursor.execute(update_sql)
                else:
                    logging.debug('写tracking失败2')
                    rm_js['code'] = '-1'
                    rm_js['message'] = rm_t.message
                    rm_js = json.dumps(rm_js)
                    return HttpResponse(rm_js)
            rm_js['code'] = '0'
            rm_js['message'] = '成功'
            rm_js = json.dumps(rm_js)
            return HttpResponse(rm_js)
        except Exception as e:
            logging.debug('ex=%s' % str(e))
            rm_js['code'] = '-1'
            rm_js['message'] = str(e)
            rm_js = json.dumps(rm_js)
            return HttpResponse(rm_js)

    # GET请求
    else:
        filter = {}
        try:
            # 传值列表
            form_data = {}
            # 获取请求参数
            lab_number = request.GET.get('lab_number', '')
            page = request.GET.get('page', 1)
            user_filter = request.GET.get('user_filter', 'all')
            action_filter = request.GET.get('action_filter', 'all')
            is_handle_filter = request.GET.get('is_handle_filter', '0')
            # 生成用户列表
            user_list = hold_cancel_request.objects.values('user_name').distinct()
            # 添加到列表
            form_data['user_filter'] = user_filter
            form_data['action_filter'] = action_filter
            form_data['user_list'] = user_list
            form_data['is_handle_filter'] = is_handle_filter
            # 搜索
            if not lab_number == '':
                loc = lab_order_controller()
                lbos = loc.get_by_entity(lab_number)
                if len(lbos) == 1:
                    lbo = lbos[0]
                    lab_number = lbo.lab_number
                _item = hold_cancel_request.objects.filter(lab_number=lab_number).order_by('is_handle', '-updated_at')
            else:
                # 二次筛选
                if not user_filter == 'all':
                    filter['user_name'] = user_filter
                if not action_filter == 'all':
                    filter['order_status_future'] = action_filter
                if not is_handle_filter == '0':
                    filter['is_handle'] = False
                _item = hold_cancel_request.objects.filter(**filter).order_by('is_handle', '-updated_at')

            # 获取URL中除page外的其它参数
            query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
            if query_string:
                query_string = '&' + query_string
            _page_info['total'] = _item.count()
            # 分页对象，设置每页20条数据
            paginator = Paginator(_item, 20)
            try:
                contacts = paginator.page(page)
            except PageNotAnInteger:
                # If page is not an integer, deliver first page.
                contacts = paginator.page(1)
            except EmptyPage:
                # If page is out of range (e.g. 9999), deliver last page of results.
                contacts = paginator.page(paginator.num_pages)

            return render(request, 'hold_cancel_request_list.html', {
                'list': contacts,
                'page_info': _page_info,
                'query_string': query_string,
                'paginator': paginator,
                'page': page,
                'requestUrl': reverse('hold_cancel_request'),
                'form_data': form_data
            })
        except Exception as e:
            logging.debug('ex=' + str(e))
            rm.code = -1
            rm.capture_execption(e)
            return HttpResponse(rm)


# 更改申请暂停或申请取消的订单（参数校验）
@login_required
@permission_required('oms.HOLD_CANCEL_REQUEST', login_url='/oms/forbid/')
def edit_apply_hold_cancel_request_list(request):
    # 获取搜索框参数
    try:
        data = {}
        lab_number = request.POST.get('lab_number', '')
        status = request.POST.get('status', '')

        # 判断hold_cancel_request里有没有
        hcr = hold_cancel_request.objects.filter(lab_number=lab_number)
        print(hcr)

        if hcr:
            return json_response(code=1, msg='已经处理过了', data=data)
        else:
            lo = LabOrder.objects.get(lab_number=lab_number,status=status)
            data['lab_number'] = lo.lab_number
            data['status'] = lo.status
            data['is_handle'] = 0

            if lo.status == 'R2HOLD':
                data['status_future'] = 'ONHOLD'
            elif lo.status == 'R2CANCEL':
                data['status_future'] = 'CANCELLED'

        return json_response(code=0, msg='Success', data=data)
    except Exception as e:
        return json_response(code=-1, msg=e)


# 更改申请暂停或申请取消的订单（参数存储）
@login_required
@permission_required('oms.HOLD_CANCEL_REQUEST', login_url='/oms/forbid/')
def apply_hold_cancel_request_list(request):
    rm = response_message()
    _page_info = {}
    _form_data = {}
    _item = None

    # post请求
    if request.method == 'POST':

        lab_number = request.POST.get('lab_number', '')
        is_handle = request.POST.get('is_handle', 0)
        status_now = request.POST.get('status_now', '')
        status_future = request.POST.get('status_future', '')
        reason = request.POST.get('reason', '')

        try:

            logging.debug('----开始创建新的hold_cancel_request订单-----')

            hcr = hold_cancel_request()
            hcr.lab_number = lab_number
            hcr.is_handle = int(is_handle)
            hcr.order_status_now = status_now
            hcr.order_status_future = status_future
            hcr.user_name = request.user.username
            hcr.user_id = request.user.id
            hcr.reason = reason
            hcr.save()
            logging.debug('----创建新的hold_cancel_request订单成功-----')
            return json_response(code=0,msg='处理成功')

        except Exception as e:

            return json_response(code=-999,msg='数据遇到异常:%s'%e)

    # GET请求
    else:
        filter = {}
        try:
            # 传值列表
            form_data = {}
            # 获取请求参数
            lab_number = request.GET.get('lab_number', '')
            page = request.GET.get('page', 1)
            action_filter = request.GET.get('action_filter', 'all')

            form_data['action_filter'] = action_filter
            # 搜索
            if not lab_number == '':
                status_tuple = ('R2HOLD', 'R2CANCEL')
                _item = LabOrder.objects.filter(lab_number=lab_number).filter(status__in=status_tuple).order_by('-update_at')
            else:
                # 二次筛选
                if not action_filter == 'all':
                    filter['status'] = action_filter
                    _item = LabOrder.objects.filter(**filter).order_by('-update_at')
                else:
                    status_tuple = ('R2HOLD','R2CANCEL')
                    _item = LabOrder.objects.filter(status__in=status_tuple).order_by('-update_at')
            # 获取URL中除page外的其它参数
            query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
            if query_string:
                query_string = '&' + query_string
            _page_info['total'] = _item.count()
            # 分页对象，设置每页20条数据
            paginator = Paginator(_item, 20)
            try:
                contacts = paginator.page(page)
            except PageNotAnInteger:
                # If page is not an integer, deliver first page.
                contacts = paginator.page(1)
            except EmptyPage:
                # If page is out of range (e.g. 9999), deliver last page of results.
                contacts = paginator.page(paginator.num_pages)

            return render(request, 'apply_hold_cancel_request_list.html', {
                'list': contacts,
                'page_info': _page_info,
                'query_string': query_string,
                'paginator': paginator,
                'page': page,
                'requestUrl': reverse('apply_hold_cancel_request'),
                'form_data': form_data
            })
        except Exception as e:
            logging.debug('ex=' + str(e))
            rm.code = -1
            rm.capture_execption(e)
            return HttpResponse(rm)


# 查询异常镜架
@login_required
@permission_required('oms.OLBO_ALL_VIEW', login_url='/oms/forbid/')
def abnormal_list(request):
    '''
    LAB Order 异常镜架订单列表
    :param request:
    :return:
    '''
    _form_data = {}
    _form_data['request_feature'] = 'All LAB ABnormal Orders'

    frame = request.GET.get('frame', '')
    od_sph = request.GET.get('od_sph', '')
    os_sph = request.GET.get('os_sph', '')
    od_cyl = request.GET.get('od_cyl', '')
    os_cyl = request.GET.get('os_cyl', '')

    page = request.GET.get('page', 1)
    currentPage = int(page)
    id = request.GET.get('id', '')
    index = request.GET.get('index', '')
    sorted = request.GET.get('sorted', '-id')

    logging.debug(sorted)
    if sorted == 'set_time':
        logging.debug('sorted: %s' % sorted)
        filter = 'all'

    _form_data['sorted'] = sorted
    _form_data['id'] = id
    _form_data['index'] = index

    items = []
    paginator = None

    timedel = date_delta()  # N天前的日期
    timedel_week = date_delta_week()
    timedel_month = date_delta_month()

    try:

        _filter = {}

        if frame == '' and od_sph == '' and os_sph == '' and od_cyl == '' and os_cyl == '':
            items = items
        else:
            items = LabOrder.objects.filter(
                is_enabled=True,
                create_at__gte=timedel)

        if frame != '':
            items = items.filter(frame=frame)
        if od_sph != '':
            items = items.filter(od_sph=od_sph)
        if os_sph != '':
            items = items.filter(os_sph=os_sph)
        if od_cyl != '':
            items = items.filter(od_cyl=od_cyl)
        if os_cyl != '':
            items = items.filter(os_cyl=os_cyl)

        count = items.count

        # 获取页码模板需要的参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))

        if query_string:
            query_string = '&' + query_string
        if count > 0:
            _form_data['total'] = count

        if sorted == 'set_time':
            paginator = Paginator(items, count)
        else:
            paginator = Paginator(items, const.PAGE_SIZE)  # Show 20 contacts per page

        try:
            items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            items = paginator.page(paginator.num_pages)

        return render(request, "laborder_abnormal_list.html",
                      {
                          'form_data': _form_data,
                          'list': items,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/laborder_abnormal_list/',
                          'query_string': query_string,
                          'frame': frame,
                          'od_sph': od_sph,
                          'os_sph': os_sph,
                          'od_cyl': od_cyl,
                          'os_cyl': os_cyl
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


def get_frame_img(request):
    try:
        rm = {}
        sku = request.GET.get('sku', '')
        res_list = []
        with connections['default'].cursor() as cursor:
            sql = """SELECT
                        p.image AS image,
                        p.thumbnail AS thumbnail,
                        s.quantity AS quantity,
                        s.al_quantity AS al_quantity
                    FROM
                        wms_product_frame AS p
                    LEFT JOIN wms_inventory_struct AS s ON p.sku = s.sku
                    WHERE
                        p.sku ='%s'
            """ % sku
            cursor.execute(sql)
            res_data = namedtuplefetchall(cursor)
            if len(res_data) > 0:
                for item in res_data:
                    res_list.append({
                        "image": item[0],
                        "thumbnail": item[1],
                        "quantity": item[2]
                    })
                rm['code'] = '0'
                rm['message'] = '执行成功'
                rm['img_url'] = settings.SSH_MEDIA_SERVER.get('MEDIA_BASE_URL', '')
                rm['list'] = res_list
                return JsonResponse(rm)
            else:
                rm['code'] = '-1'
                rm['message'] = '未查询到该sku,请查看是否存在'
                rm['list'] = res_list
                return JsonResponse(rm)
    except Exception as e:
        logging.debug(e)
        rm['code'] = '-1'
        rm['message'] = '执行失败'
        rm['list'] = res_list
        return JsonResponse(rm)


# 查询dashbord processing
@login_required
@permission_required('oms.DASHBORD_PROCESSING_LIST', login_url='/oms/forbid/')
def dashbord_processing_list(request):
    _form_data = {}
    _form_data['request_feature'] = 'All DASHBORD PROCESSING LIST'

    page = request.GET.get('page', 1)
    currentPage = int(page)
    id = request.GET.get('id', '')
    index = request.GET.get('index', '')
    sorted = request.GET.get('sorted', '-id')

    _form_data['sorted'] = sorted
    _form_data['id'] = id
    _form_data['index'] = index
    diff = {}
    items = []
    itemnews = []
    paginator = None

    try:

        _filter = {}
        with connections['default'].cursor() as cursor:
            sql_processing = """
            select * from oms_pgorder where create_at>'2019.01.01'
                    and is_enabled=True
                    and status='processing'
                    and is_inlab=False order by create_at desc"""
            cursor.execute(sql_processing)
            items = namedtuplefetchall(cursor)

        d1 = datetime.datetime.now()
        for item_info in items:
            dt = datetime.datetime.now()
            dt1 = item_info.create_at
            order_datetime = item_info.order_datetime.strftime('%Y-%m-%d %H:%M:%S')
            start = dt.strftime("%Y-%m-%d %H:%M:%S")
            # end = dt1.strftime("%Y-%m-%d %H:%M:%S")
            end = (dt1 + datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            d1 = datetime.datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
            d2 = datetime.datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
            delta = d1 - d2
            hours = float(delta.seconds) / 3600
            days = delta.days
            if days > 0:
                hours_tmp = float(days) * 24
                hours = float(hours_tmp) + float(hours)
            hours = '%.1f' % hours
            hours = decimal.Decimal(hours)
            if hours < 1.0:
                diff['hours_of_purchase'] = '1'
                diff['hours'] = hours
                diff['order_number'] = item_info.order_number
                diff['id'] = item_info.id
                diff['shipping_method'] = item_info.shipping_method
                diff['order_datetime'] = order_datetime
                diff['create_at'] = end
                itemnews.append(diff)
                diff = {}
            if hours >= 1.0 and hours <= 2.0:
                diff['hours_of_purchase'] = '2'
                diff['hours'] = hours
                diff['order_number'] = item_info.order_number
                diff['id'] = item_info.id
                diff['shipping_method'] = item_info.shipping_method
                diff['order_datetime'] = order_datetime
                diff['create_at'] = end
                itemnews.append(diff)
                diff = {}
            if hours > 2.0 and hours <= 3.0:
                diff['hours_of_purchase'] = '3'
                diff['hours'] = hours
                diff['order_number'] = item_info.order_number
                diff['id'] = item_info.id
                diff['shipping_method'] = item_info.shipping_method
                diff['order_datetime'] = order_datetime
                diff['create_at'] = end
                itemnews.append(diff)
                diff = {}
            if hours > 3.0:
                diff['hours_of_purchase'] = '4'
                diff['hours'] = hours
                diff['order_number'] = item_info.order_number
                diff['id'] = item_info.id
                diff['shipping_method'] = item_info.shipping_method
                diff['order_datetime'] = order_datetime
                diff['create_at'] = end
                itemnews.append(diff)
                diff = {}

        count = itemnews.count

        # 获取页码模板需要的参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))

        if query_string:
            query_string = '&' + query_string
        if count > 0:
            _form_data['total'] = count

        if sorted == 'set_time':
            paginator = Paginator(itemnews, count)
        else:
            paginator = Paginator(itemnews, const.PAGE_SIZE)  # Show 20 contacts per page

        try:
            itemnews = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            itemnews = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            itemnews = paginator.page(paginator.num_pages)

        return render(request, "dashbord_processing_list.html",
                      {
                          'form_data': _form_data,
                          'list': itemnews,
                          'currentPage': currentPage,
                          'paginator': paginator,
                          'requestUrl': '/oms/dashbord_processing_list/',
                          'query_string': query_string
                      })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirect_overdue_orders(request):
    _form_data = {}
    lbo_list = []
    try:
        lab_number = request.GET.get('lab_number', '')
        vendor = request.GET.get('vendor', 'all')
        filter = request.GET.get('filter', 'all')
        status = request.GET.get('status', '')
        page = request.GET.get('page', 1)
        currentPage = int(page)
        vendors_choice_list = []
        status_choice_list = []
        exclusion_status = ['ASSEMBLED',
                            'SHIPPING', 'DELIVERED', 'CANCELLED', 'CLOSED']

        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            vendors_choice_list.append(vc)

        for vcl in LabOrder.STATUS_CHOICES:
            if vcl[0] not in exclusion_status:
                vc = status_choice()
                vc.key = vcl[0]
                vc.value = vcl[1]
                status_choice_list.append(vc)

        sql = """SELECT
                        l.id AS id,
                        l.lab_number AS lab_number,
                        l.act_ship_direction AS act_ship_direction,
                        l.frame AS frame,
                        l.quantity AS quantity,
                        l.lens_name AS lens_name,
                        l.act_lens_name AS act_lens_name,
                        CONVERT_TZ(l.update_at,@@session.time_zone,'+8:00') AS update_at,
                        l.comments_inner as comments_inner,
                        l.`status` AS `status`,
                        l.vendor AS vendor,
                        CONVERT_TZ(l.create_at,@@session.time_zone,'+8:00') AS create_at,
						CONVERT_TZ(lpol.created_at,@@session.time_zone,'+8:00') AS delivery_date,
                        l.cur_progress AS cur_progress,
                        CONVERT_TZ(l.estimated_date,@@session.time_zone,'+8:00') AS estimated_date,
                        l.overdue_reasons AS overdue_reasons,
                        (TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S"))) AS overdue_hours
                FROM
                  oms_laborder AS l LEFT JOIN oms_laborder_purchase_order_line AS lpol ON l.lab_number = lpol.lab_number
                WHERE
                    l.`status` NOT IN (
                        'ASSEMBLED',
                        'SHIPPING',
                        'DELIVERED',
                        'CANCELLED',
                        'CLOSED'
                    ) AND ((TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S")) >24 AND (l.vendor = '2' OR l.vendor = '3' OR l.vendor = '6' OR l.vendor = '8' OR l.vendor = '10')) 
                OR (TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S"))>72 AND (l.vendor = '4' OR l.vendor= '5' OR l.vendor = '7' OR l.vendor = '9'))) """
        if lab_number != '':
            sql = sql + """ AND l.lab_number LIKE '%%%s%%' """ % lab_number
        if vendor != '' and vendor != 'all':
            sql = sql + """ AND l.vendor='%s' """ % vendor

        if status != '' and status != 'all':
            sql = sql + """ AND l.status='%s' """ % status
        elif status == '':
            sql = sql + """ AND l.status='' """

        if filter == 'HANDLE':
            sql = sql + """ AND l.overdue_reasons<>'' """
        elif filter == 'UNTREATED':
            sql = sql + """ AND l.overdue_reasons='' """

        sql = sql + """ ORDER BY l.create_at """
        print (sql)
        import pytz
        with connections['pg_oms_query'].cursor() as cursor:
            cursor.execute(sql)
            laborders = namedtuplefetchall(cursor)
            for item in laborders:
                # create_at = item.create_at.replace(tzinfo=pytz.timezone('UTC'))
                if item.act_ship_direction == 'STANDARD':
                    act_ship_direction_name = '普通'
                elif item.act_ship_direction == 'EXPRESS':
                    act_ship_direction_name = '加急'
                elif item.act_ship_direction == 'EMPLOYEE':
                    act_ship_direction_name = '内部'
                elif item.act_ship_direction == 'FLATRATE':
                    act_ship_direction_name = '批量'
                elif item.act_ship_direction == 'CA_EXPRESS':
                    act_ship_direction_name = '加急-加拿大'
                lbo_list.append({
                    "id": item.id,
                    "lab_number": item.lab_number,
                    "status": LAB_STATUS.get(item.status, ''),
                    "vendor": item.vendor,
                    "create_at": item.create_at,
                    "delivery_date": item.delivery_date,
                    "cur_progress": item.cur_progress,
                    "estimated_date": item.estimated_date,
                    "overdue_reasons": item.overdue_reasons,
                    "overdue_hours": item.overdue_hours,
                    "days_of_production": int(item.overdue_hours) / 24,
                    # (datetime.datetime.now() - item.create_at).days
                    "act_ship_direction_name": act_ship_direction_name,
                    "frame": item.frame,
                    "quantity": item.quantity,
                    "lens_name": item.lens_name,
                    "act_lens_name": item.act_lens_name,
                    "update_at": item.update_at,
                    "comments_inner": item.comments_inner
                })
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        _form_data['total'] = len(lbo_list)
        paginator = Paginator(lbo_list, const.PAGE_SIZE_MORE)

        try:
            lbo_list = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            lbo_list = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            lbo_list = paginator.page(paginator.num_pages)

        _form_data['lbo_list'] = lbo_list
        return render(request, 'laborder_overdue_list.html', {
            'form_data': _form_data,
            'currentPage': currentPage,
            'paginator': paginator,
            'vendors_choices': vendors_choice_list,
            'status_choices': status_choice_list,
            'requestUrl': '/oms/overdue_orders/',
            'query_string': query_string,
            'filter': filter,
            'vendor': vendor,
            'status': status
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


# custom_tag
from django.template.defaulttags import register


@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)


from report.overdue_orders_report import OverdueOrderReport


@register.filter
def get_query_string_exclude_self(query_string, key):
    parameters = []

    qs = query_string.replace('?','').split('&')

    for item in qs:
        ps = item.split('=')
        p = ps[0]
        if p != key and p != '' and p != 'lab_number':
            if not parameters:
                parameters.append( p + '=' + ps[1])
            else:
                parameters.append('&' + p + '=' + ps[1])
    return ''.join(parameters)


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirect_overdue_orders_v2(request):
    _form_data = {}
    _items = []
    try:
        lab_number = request.GET.get('lab_number','')
        overdue_days = request.GET.get('overdue_days', '2')
        if overdue_days == '':
            overdue_days = '2'
        status = request.GET.get('status', 'all')
        vendor = request.GET.get('vendor', 'all')
        priority = request.GET.get('priority', 'all')
        ship_direction = request.GET.get('ship_direction', 'all')

        parameters = {}
        parameters['lab_number'] = lab_number
        parameters['overdue_days'] = overdue_days
        parameters['status'] = status
        parameters['vendor'] = vendor
        parameters['priority'] = priority
        parameters['ship_direction'] = ship_direction

        oor = OverdueOrderReport()
        rm = oor.get_overdue_orders_summary_report(request, parameters)

        _form_data['overdue_orders_summary'] = rm.obj
        rm = oor.get_overdue_orders_list_report(request, parameters)

        _items = rm.obj.get('items')

        request.session['overdue_orders_list_sql'] = rm.obj.get('sql_script')

        page = request.GET.get('page', 1)
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        query_string_complex = request.META.get('QUERY_STRING', None)

        _form_data['total'] = len(_items)

        vendor_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            vendor_list.append(vc)
        _form_data['vendor_list'] = vendor_list

        priority_list = []
        for i in range(6):
            vc = status_choice()
            vc.key = str(i)
            vc.value = str(i)
            if i != 0:
                priority_list.append(vc)
        _form_data['priority_list'] = priority_list

        ship_direction_list = []
        for vcl in LabOrder.SHIP_DIRECTION_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            ship_direction_list.append(vc)
        _form_data['ship_direction_list'] = ship_direction_list

        _form_data['vendor'] = vendor
        _form_data['overdue_days'] = overdue_days
        _form_data['status'] = status
        _form_data['priority'] = priority
        _form_data['ship_direction'] = ship_direction

        paginator = Paginator(_items, 50)

        try:
            _items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            _items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            _items = paginator.page(paginator.num_pages)

        return render(request, 'laborder_overdue_list_v2.html', {
            'form_data': _form_data,
            'list': _items,
            'paginator': paginator,
            'requestUrl': '/oms/overdue_orders_v2/',
            'query_string': query_string,
            'query_string_complex': query_string_complex,
            'filter': filter,
        })
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirect_overdue_orders_csv_v2(request):
    _form_data = {}
    lbo_list = []
    try:
        import csv, codecs

        sql = request.session.get('overdue_orders_list_sql','')

        data = DbHelper.query_with_titles(sql)

        response = HttpResponse(content_type='text/csv')
        time_now = time.strftime('%Y%m%d', time.localtime(time.time()))
        file_name = 'overdue_laborder_csv-%s' % time_now
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response)
        writer.writerow(data['titles'])
        import pytz

        for item in data['results']:
            writer.writerow(item)

        return response
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirect_sub_delay_data(request):
    rm = res_msg.response_dict()
    try:
        form_data = request.POST.get("form_data")
        _data = json.loads(form_data)
        lbo = LabOrder.objects.get(lab_number=_data['lab_number'])
        cur_progress = lbo.cur_progress
        overdue_reasons = lbo.overdue_reasons
        lbo.overdue_reasons = overdue_reasons + '-->' + _data['overdue_reasons']
        lbo.cur_progress = cur_progress + '-->' + _data['cur_progress']
        if _data['delay_time'] == '':
            now_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        else:
            now_date = _data['delay_time']
        lbo.estimated_date = now_date
        lbo.save()
        remark = "【{0}】：{1}".format(now_date, _data['overdue_reasons'])
        # 添加订单记录
        tloc = tracking_lab_order_controller()
        tloc.tracking(lbo, request.user, "", "催单", remark)
        rm['code'] = '0'
        rm['message'] = '执行成功'
        return JsonResponse(rm)
    except Exception as e:
        rm['code'] = '-1'
        rm['message'] = e
        return JsonResponse(rm)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def edit_laborder_frame(request):
    m_order_number = request.POST.get('order_number')
    rm = res_msg()
    try:
        # poi = PgOrderItem.objects.get(lab_order_number=m_order_number)
        lab = LabOrder.objects.get(lab_number=m_order_number)
        logging.debug(lab)
    except Exception as e:
        rm.capture_execption(e)
        return JsonResponse({'code': rm.code, 'message': rm.message})

    prescripion_num = generate_prescripion_tuple()
    # item = PgOrderItemFormDetailParts(instance=poi)
    # logging.debug(item)
    return render(request, 'laborder_prescripion_detail.html', {
        'instance': lab,
        'item': lab,
        'prescripion_num': prescripion_num,
    })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def modify_laborder_frame(request):
    rm = res_msg.response_dict()
    try:
        order_number = request.POST.get("order_number")
        form_data = request.POST.get("form_data")
        form_json = json.loads(form_data)
        lab_orders = LabOrder.objects.filter(lab_number=order_number)
        if len(lab_orders) > 1:
            rm['code'] = 100
            rm['message'] = u'已有多个工厂订单 需手动修改'
            return JsonResponse(rm)

        with transaction.atomic():
            # 如果有工厂订单
            if not len(lab_orders) == 0:
                lbo = lab_orders[0]
                if lbo.status not in (
                        '', None, 'REQUEST_NOTES', 'FRAME_OUTBOUND', 'PRINT_DATE', 'LENS_OUTBOUND', 'LENS_REGISTRATION',
                        'LENS_RETURN', 'LENS_RECEIVE', 'ONHOLD', 'R2HOLD'):
                    rm['code'] = 300
                    rm['message'] = u'当前状态【%s】不允许修改验光单数据' % lbo.get_status_display()
                    return JsonResponse(rm)
                logging.debug(form_json.get('od_pd'))
                logging.debug(form_json.get('os_pd'))
                logging.debug(form_json.get('pd'))
                lbo.od_prism = form_json.get('od_prism')
                lbo.od_base = form_json.get('od_base')
                lbo.od_prism1 = form_json.get('od_prism1')
                lbo.od_base1 = form_json.get('od_base1')
                lbo.od_axis = form_json.get('od_axis')
                lbo.od_sph = form_json.get('od_sph')
                lbo.od_cyl = form_json.get('od_cyl')
                if form_json.get('is_singgle_pd') == 'True':
                    lbo.pd = form_json.get('pd')
                    lbo.is_singgle_pd = True
                else:
                    lbo.is_singgle_pd = False

                lbo.od_pd = form_json.get('od_pd')
                lbo.od_add = form_json.get('od_add')

                lbo.os_prism = form_json.get('os_prism')
                lbo.os_base = form_json.get('os_base')
                lbo.os_prism1 = form_json.get('os_prism1')
                lbo.os_base1 = form_json.get('os_base1')
                lbo.os_axis = form_json.get('os_axis')
                lbo.os_sph = form_json.get('os_sph')
                lbo.os_cyl = form_json.get('os_cyl')
                lbo.os_pd = form_json.get('os_pd')
                lbo.os_add = form_json.get('os_add')

                lbo.is_production_change = True
                lbo.save()

                l_orders = lens_order.objects.filter(lab_number=order_number)
                for item in l_orders:
                    if item.rl_identification == 'R':
                        item.sph = form_json.get('od_sph')
                        item.cyl = form_json.get('od_cyl')
                        item.axis = form_json.get('od_axis')

                        item.add = form_json.get('od_add')
                        item.prism = form_json.get('od_prism')
                        item.base = form_json.get('od_base')
                        item.prism1 = form_json.get('od_prism1')
                        item.base1 = form_json.get('od_base1')

                        if lbo.is_singgle_pd:
                            item.pd = float(form_json.get('pd'))/2
                        else:
                            item.pd = form_json.get('od_pd')
                        item.save()
                    else:
                        item.sph = form_json.get('os_sph')
                        item.cyl = form_json.get('os_cyl')
                        item.axis = form_json.get('os_axis')

                        item.add = form_json.get('os_add')
                        item.prism = form_json.get('os_prism')
                        item.base = form_json.get('os_base')
                        item.prism1 = form_json.get('os_prism1')
                        item.base1 = form_json.get('os_base1')

                        if lbo.is_singgle_pd:
                            item.pd = float(form_json.get('pd'))/2
                        else:
                            item.pd = form_json.get('os_pd')
                        item.save()

                tloc = tracking_lab_order_controller()
                tloc.tracking(lbo, request.user, 'CHANGED_RX', '修改验光单&制作参数')

        rm['code'] = '0'
        rm['message'] = '执行成功'

        return JsonResponse(rm)

    except Exception as e:
        logging.debug('Exception: %s' % e)
        rm['code'] = 404
        rm['message'] = u'数据存在异常'
        return JsonResponse(rm)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def edit_laborder_sku(request):
    m_order_number = request.POST.get('order_number')
    rm = res_msg()
    try:
        lab = LabOrder.objects.get(lab_number=m_order_number)
    except Exception as e:
        rm.capture_execption(e)
        return JsonResponse({'code': rm.code, 'message': rm.message})

    return render(request, 'laborder_sku_detail.html', {
        'item': lab
    })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def modify_laborder_sku(request):
    rm = res_msg.response_dict()
    try:
        order_number = request.POST.get("order_number")
        form_data = request.POST.get("form_data")
        form_json = json.loads(form_data)
        lab_orders = LabOrder.objects.filter(lab_number=order_number)
        if len(lab_orders) > 1:
            rm['code'] = 100
            rm['message'] = u'已有多个工厂订单 需手动修改'
            return JsonResponse(rm)
        with transaction.atomic():
            # 如果有工厂订单
            if not len(lab_orders) == 0:
                lbo = lab_orders[0]
                if lbo.status not in (
                        '', None, 'REQUEST_NOTES', 'FRAME_OUTBOUND', 'PRINT_DATE', 'LENS_OUTBOUND', 'LENS_REGISTRATION',
                        'LENS_RETURN', 'LENS_RECEIVE', 'ONHOLD', 'R2HOLD'):
                    rm['code'] = 300
                    rm['message'] = u'当前状态【%s】不允许修改加工参数数据' % lbo.get_status_display()
                    return JsonResponse(rm)
                tint_sku = str(form_json.get('tint_sku'))
                if tint_sku.find("%2B") >= 0:
                    tint_sku = tint_sku.replace("%2B", "+")
                lbo.tint_sku = tint_sku

                tint_name = str(form_json.get('tint_name'))
                if tint_name.find("%2B") >= 0:
                    tint_name = tint_name.replace("%2B", "+")
                lbo.tint_name = tint_name

                pal_design_sku = str(form_json.get('pal_design_sku'))
                if pal_design_sku.find("%2B") >= 0:
                    pal_design_sku = pal_design_sku.replace("%2B", "+")
                lbo.pal_design_sku = pal_design_sku

                pal_design_name = str(form_json.get('pal_design_name'))
                if pal_design_name.find("%2B") >= 0:
                    pal_design_name = pal_design_name.replace("%2B", "+")
                lbo.pal_design_name = pal_design_name

                coating_sku = str(form_json.get('coating_sku'))
                if coating_sku.find("%2B") >= 0:
                    coating_sku = coating_sku.replace("%2B", "+")
                lbo.coating_sku = coating_sku

                coating_name = str(form_json.get('coating_name'))
                if coating_name.find("%2B") >= 0:
                    coating_name = coating_name.replace("%2B", "+")
                lbo.coating_name = coating_name

                lab_seg_height = str(form_json.get('lab_seg_height'))
                if lab_seg_height.find("%2B") >= 0:
                    lab_seg_height = lab_seg_height.replace("%2B", "+")
                lbo.lab_seg_height = lab_seg_height

                assemble_height = str(form_json.get('assemble_height'))
                if assemble_height.find("%2B") >= 0:
                    assemble_height = assemble_height.replace("%2B", "+")
                lbo.assemble_height = assemble_height

                sub_mirrors_height = str(form_json.get('sub_mirrors_height'))
                if sub_mirrors_height.find("%2B") >= 0:
                    sub_mirrors_height = sub_mirrors_height.replace("%2B", "+")
                lbo.sub_mirrors_height = sub_mirrors_height

                special_handling = str(form_json.get('special_handling'))
                if special_handling.find("%2B") >= 0:
                    special_handling = special_handling.replace("%2B", "+")
                lbo.special_handling = special_handling

                lbo.is_production_change = True
                lbo.save()

                tloc = tracking_lab_order_controller()
                tloc.tracking(lbo, request.user, 'CHANGED_RX', '修改验光单&制作参数')

        rm['code'] = '0'
        rm['message'] = '执行成功'
        return JsonResponse(rm)

    except Exception as e:
        logging.debug('Exception: %s' % e)
        rm['code'] = 404
        rm['message'] = u'数据存在异常'
        return JsonResponse(rm)


@login_required
@permission_required('oms.OLOR_VIEW', login_url='/oms/forbid/')
def redirect_overdue_orders_csv(request):
    _form_data = {}
    lbo_list = []
    try:
        import csv, codecs
        lab_number = request.GET.get('lab_number', '')
        vendor = request.GET.get('vendor', 'all')
        filter = request.GET.get('filter', 'all')
        status = request.GET.get('status', '')

        sql = """SELECT
                        l.id AS id,
                        l.lab_number AS lab_number,
                        l.`status` AS `status`,
                        l.act_ship_direction AS act_ship_direction,
                        l.frame AS frame,
                        l.quantity AS quantity,
                        l.lens_name AS lens_name,
                        l.act_lens_name AS act_lens_name,
                        CONVERT_TZ(l.update_at,@@session.time_zone,'+8:00') AS update_at,
                        l.comments_inner as comments_inner,
                        l.vendor AS vendor,
                        CONVERT_TZ(l.create_at,@@session.time_zone,'+8:00') AS create_at,
						CONVERT_TZ(lpol.created_at,@@session.time_zone,'+8:00') AS delivery_date,
                        l.cur_progress AS cur_progress,
                        CONVERT_TZ(l.estimated_date,@@session.time_zone,'+8:00') AS estimated_date,
                        l.overdue_reasons AS overdue_reasons,
                        (TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S"))) AS overdue_hours
                FROM
                  oms_laborder AS l LEFT JOIN oms_laborder_purchase_order_line AS lpol ON l.lab_number = lpol.lab_number
                WHERE
                    l.`status` NOT IN (
                        'ASSEMBLED',
                        'SHIPPING',
                        'DELIVERED',
                        'CANCELLED',
                        'CLOSED'
                    ) AND ((TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S")) >24 AND (l.vendor = '2' OR l.vendor = '3' OR l.vendor = '6' OR l.vendor = '8' OR l.vendor = '10')) 
                OR (TIMESTAMPDIFF(HOUR, l.create_at, date_format(NOW(), "%Y-%m-%d %H:%i:%S"))>72 AND (l.vendor = '4' OR l.vendor= '5' OR l.vendor = '7' OR l.vendor = '9'))) """
        if lab_number != '':
            sql = sql + """ AND l.lab_number LIKE '%%%s%%' """ % lab_number
        if vendor != '' and vendor != 'all':
            sql = sql + """ AND l.vendor='%s' """ % vendor

        if status != '' and status != 'all':
            sql = sql + """ AND l.status='%s' """ % status
        elif status == '':
            sql = sql + """ AND l.status='' """

        if filter == 'HANDLE':
            sql = sql + """ AND l.overdue_reasons<>'' """
        elif filter == 'UNTREATED':
            sql = sql + """ AND l.overdue_reasons='' """

        sql = sql + """ ORDER BY l.create_at """

        response = HttpResponse(content_type='text/csv')
        file_name = 'overdue_laborder_csv'
        response['Content-Disposition'] = 'attachment;filename=' + file_name + '.csv'
        response.write(codecs.BOM_UTF8)
        writer = csv.writer(response)
        writer.writerow([
            u'#', u'订单号', u'实际发货', u'镜架', u'数量', u'计划镜片', u'镜片', u'当前状态', u'vendor', u'创建日期', u'下达日期', u'更新时间', u'目前进程',
            u'生产天数', u'预计完成', u'内部备注', u'超期原因', u'超期时间(小时)'
        ])
        import pytz
        with connections['pg_oms_query'].cursor() as cursor:
            cursor.execute(sql)
            laborders = namedtuplefetchall(cursor)
            for item in laborders:
                create_at = item.create_at.replace(tzinfo=pytz.timezone('UTC'))
                if item.act_ship_direction == 'STANDARD':
                    act_ship_direction_name = '普通'
                elif item.act_ship_direction == 'EXPRESS':
                    act_ship_direction_name = '加急'
                elif item.act_ship_direction == 'EMPLOYEE':
                    act_ship_direction_name = '内部'
                elif item.act_ship_direction == 'FLATRATE':
                    act_ship_direction_name = '批量'
                elif item.act_ship_direction == 'CA_EXPRESS':
                    act_ship_direction_name = '加急-加拿大'

                if item.create_at is not None:
                    created_at = item.create_at.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    created_at = ''

                if item.delivery_date is not None:
                    delivery_date = item.delivery_date.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    delivery_date = ''

                if item.estimated_date is not None:
                    estimated_date = item.estimated_date.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    estimated_date = ''

                if item.update_at is not None:
                    update_at = item.update_at.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    update_at = ''
                writer.writerow([
                    item.id, item.lab_number, act_ship_direction_name, item.frame, item.quantity, item.lens_name,
                    item.act_lens_name, LAB_STATUS.get(item.status, ''), item.vendor, created_at, delivery_date,
                    update_at,
                    item.cur_progress, int(item.overdue_hours) / 24, estimated_date, item.comments_inner,
                    item.overdue_reasons,
                    item.overdue_hours
                ])

        return response
    except Exception as e:
        logging.debug('Exception: %s' % e.message)
        _form_data['exceptions'] = e
        _form_data['error_message'] = e.message
        return render(request, "exceptions.html",
                      {
                          'form_data': _form_data,
                      })


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def change_frame_sku(request):
    lab_number = request.POST.get('lab_number')
    new_frame = request.POST.get('frame')
    new_image = request.POST.get('image')
    print(lab_number)
    print(new_frame)
    print(new_image)
    rm = res_msg()
    try:
        with transaction.atomic():
            labs = LabOrder.objects.filter(lab_number=lab_number)
            print(len(labs))
            if len(labs) > 0:
                lbo = labs[0]
                log_str = '镜架sku由【' + lbo.frame + '】改为【' + new_frame + '】'
                if lbo.status in ('ONHOLD', 'R2HOLD'):
                    oiss = inventory_struct.objects.filter(sku=lbo.frame)
                    if len(oiss) > 0:
                        ois = oiss[0]
                        ois.reserve_quantity = ois.reserve_quantity - lbo.quantity if ois.reserve_quantity - lbo.quantity >= 0 else 0
                        ois.save()

                    niss = inventory_struct.objects.filter(sku=new_frame)
                    if len(niss) > 0:
                        nis = niss[0]
                        nis.reserve_quantity = nis.reserve_quantity + lbo.quantity
                        nis.save()

                    lbo.frame = new_frame
                    lbo.image = new_image
                    lbo.thumbnail = new_image
                    lbo.save()

                    tloc = tracking_lab_order_controller()
                    tloc.tracking(lbo, request.user, 'CHANGED_RX', log_str)

                    return json_response(code=0, msg='执行成功!')
                else:
                    return json_response(code=-1, msg='当前状态为【' + lbo.status + '】不允许修改')
            return json_response(code=-1, msg='未查找到LabOrder')
    except Exception as e:
        rm.capture_execption(e)
        json_response(code=-1, msg=e)


@login_required
@permission_required('oms.POL_VIEW', login_url='/oms/forbid/')
def redirect_pgorder_item_detail_v3_nrx(request):
    m_order_number = request.POST.get("order_number")
    m_item_id = request.POST.get("item_id")
    m_product_index = request.POST.get("product_index")
    logging.debug('order_number: %s' % m_order_number)
    logging.debug('item_id: %s' % m_item_id)
    poc = PgOrderController()
    poi = poc.get_pgorder_item_detail_v3(m_order_number, m_item_id, m_product_index)

    item = PgOrderItemFormDetailParts(instance=poi)
    prd_img_pre = const.PRODUCT_IMAGE_PREPATH
    order_image_urls = poi.order_image_urls

    # get lab order info corresponding to PG order
    accs_orders = AccsOrder.objects.filter(base_entity=poi.id)
    lab_orders = LabOrder.objects.filter(base_entity=poi.id)
    return render(
        request, "pgorder_item_detail_v3_nrx.pspf.html", {
            'obj_type': poi.type,
            'obj_id': poi.id,
            'instance': poi,
            'item': item,
            'prd_img_pre': prd_img_pre,
            'order_image_urls': order_image_urls,
            'accs_orders': accs_orders,
            'lab_orders': lab_orders
        }
    )


# 根据验光单ID获取验光单数据 update
def get_origin_prescription(entity_id):
    sql = "select * from prescription_entity where entity_id=%s" % entity_id
    logging.debug(sql)
    with connections['pg_mg_query'].cursor() as cursor:
        cursor.execute(sql)
        # results = namedtuplefetchall(cursor)
        results = dictfetchall(cursor)

        if len(results) > 0:
            return results[0]
        else:
            return False


@login_required
def redirect_sync_shipment_status(request):
    try:
        lab_number = request.POST.get("lab_number", "")
        entity = request.POST.get("entity", "")
        lab_order = LabOrder.objects.get(lab_number=lab_number)
        if lab_order.status not in ['SHIPPING', 'DELIVERED']:
            return json_response(code=-1, msg='当前状态不能进行同步')

        if lab_order.act_ship_direction == 'STANDARD':
            res_data = get_ship2_tracking_number(lab_number)
            logging.debug(res_data)
            if res_data['code'] == 0:
                ship_data = res_data['ship_data']
                if ship_data != '' and len(ship_data) > 0:
                    ep_tracking_code = ship_data[1]
                    ep_created_at = ship_data[0]
                    ep_label_url = ship_data[2]
                    logging.debug(ep_created_at)
                    if ep_tracking_code != '':
                        # jobs.shipping_number = ep_tracking_code
                        form_data = read_ups_by_track_numbers(ep_tracking_code)
                        logging.debug(form_data)
                        if form_data != '' and len(form_data) == 3:
                            delivery_year = form_data[0]
                            delivery_hour = form_data[1]
                            status = form_data[2]
                            if status != '' and "Delivered" in status:
                                lab_order.status = 'DELIVERED'
                                lab_order.save()
                                # 写TRACKING
                                tloc = tracking_lab_order_controller()
                                tloc.tracking(lab_order, request.user, 'DELIVERED', '妥投')
            else:
                return json_response(code=-1, msg='接口错误:{0}'.format(res_data['msg']))
        elif lab_order.act_ship_direction == 'EXPRESS' or lab_order.act_ship_direction == 'CA_EXPRESS':
            from util.ups import run
            if lab_order.shipping_number is not None and not lab_order.shipping_number == '':
                shipping_number = lab_order.shipping_number
                shipping_number_list = shipping_number.split(",")
                for shipping_number in shipping_number_list:
                    data = run(shipping_number)
                    if data != '' and len(data) == 2:
                        if not data[0] == '':
                            lab_order.status = 'DELIVERED'
                            lab_order.save()
                            # 写TRACKING
                            tloc = tracking_lab_order_controller()
                            tloc.tracking(lab_order, request.user, 'DELIVERED', '妥投')
                    else:
                        return json_response(code=-1, msg='接口错误')
            else:
                return json_response(code=-1, msg='为获取到Shipping Number')

        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
def redirect_edit_laborder_shipment(request):
    try:
        lab_number = request.POST.get("lab_number", "")
        act_ship_direction = request.POST.get("act_ship_direction", "")
        msg = u'原发货方式【%s】,发货方式状态【%s】'
        lab = LabOrder.objects.get(lab_number=lab_number)
        msg = msg % (lab.act_ship_direction, act_ship_direction)
        lab.act_ship_direction = act_ship_direction
        lab.save()
        tloc = tracking_lab_order_controller()
        tloc.tracking(lab, request.user, "CHANGE_SHIP_DIRECTION", "发货状态调整", msg)
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
def redirect_edit_hold_to_processing(request):
    try:
        order_number = request.POST.get("order_number", "")
        action = '【%s】To Processing'
        if order_number == '':
            return json_response(code=-1, msg='参数错误！')
        pgorder = PgOrder.objects.get(order_number=order_number)
        if pgorder.status == 'r2hold' or pgorder.status == 'pending' or pgorder.status == 'holded':
            action = action % (pgorder.status)
            pgorder.status = 'processing'
            pgorder.save()
            poc = PgOrderController()
            poc.tracking_operation_log(request, pgorder, action, "", "")
        else:
            return json_response(code=-1, msg='只有【r2hold】【pending】【holded】状态可以更改状态为【processing】')
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


#根据验光单ID获取验光单数据 update
def get_origin_prescription(entity_id):
    sql = "select * from prescription_entity where entity_id=%s" % entity_id
    logging.debug(sql)
    with connections['pg_mg_query'].cursor() as cursor:
        cursor.execute(sql)
        # results = namedtuplefetchall(cursor)
        results = dictfetchall(cursor)

        if len(results)>0:
           results[0]['is_nonrx']=0
           return results[0]
        else:
            return False


@login_required
def redirect_pg_order_reorder(request):
    m_order_number = request.GET.get("order_number", '')
    po = PgOrder.objects.get(order_number=m_order_number)
    poi_items = PgOrderItem.objects.filter(order_number = m_order_number).defer('product_options')
    reorder_list = RemakeOrder.objects.filter(order_number = m_order_number)

    # 所有运输方式
    for item in poi_items:
        if(item.profile_prescription_id):
            item.origin_prescription = get_origin_prescription(item.profile_prescription_id)
        else:
            item.origin_prescription = None
        pg_order_remake_carts_list = RemakeOrderCart.objects.filter(item_id=item.item_id,is_remake=0).defer('user_id','user_name','created_at','updated_at')
        pg_ordered_remake_carts_list = RemakeOrderCart.objects.filter(item_id=item.item_id,is_remake=1).defer('user_id','user_name','created_at','updated_at')
        item.remake_cart_list= json.dumps(utilities.convert_to_dicts(pg_order_remake_carts_list),cls=DateEncoder)
        item.remake_cart_list1 = json.dumps(utilities.convert_to_dicts(pg_order_remake_carts_list), cls=DateEncoder)
        item.ordered_remake_cart_list= json.dumps(utilities.convert_to_dicts(pg_ordered_remake_carts_list),cls=DateEncoder)
    # print(json.dumps(poi_items,cls=DateEncoder))
    data_list = []
    for k in reorder_list:
        reorder = {"remake_order":k.remake_order,"created_at":k.created_at.strftime("%Y-%m-%d %H:%M:%S")}
        data_list.append(reorder)
    #所有运输方式
    sel_shipping_methods = DICT_SHIPPING_METHODS.get(po.country_id,'')
    return render(
        request, "pg_order_reorder.html", {
            'sel_shipping_methods': json.dumps(sel_shipping_methods),
            'po': po,
            'items':json.dumps(utilities.convert_to_dicts(poi_items),cls=DateEncoder),
            'reorder_list': data_list,
            # 'origin_prescription':json.dumps(origin_prescription,cls=DateEncoder),
            'prd_img_pre': const.PRODUCT_IMAGE_PREPATH,
            # 'order_image_urls': order_image_urls,
            # 'lab_orders': lab_orders,
        }
    )


@login_required
@permission_required('oms.CHANGE_ORDER_VALUE', login_url='/oms/forbid/')
def redirect_edit_laborder_parameters(request):
    try:
        lab_number = request.POST.get("lab_number", "")
        lab_seg_height = request.POST.get("lab_seg_height", "")
        assemble_height = request.POST.get("assemble_height", "")
        sub_mirrors_height = request.POST.get("sub_mirrors_height", "")
        special_handling = request.POST.get("special_handling", "")

        lab = LabOrder.objects.get(lab_number=lab_number)
        m_value = u''
        m_key = u'修改加工参数'
        msg = u''
        old_lab_seg_height = lab.lab_seg_height
        old_assemble_height = lab.assemble_height
        old_sub_mirrors_height = lab.sub_mirrors_height
        old_special_handling = lab.special_handling

        if not old_sub_mirrors_height:
            old_sub_mirrors_height = ''

        if not old_special_handling:
            old_special_handling = ''

        # # 加工瞳高
        if lab_seg_height != old_lab_seg_height:
            # 原值
            m_value = m_value + "&" + lab_seg_height
            msg = msg + "&" + '加工瞳高原值=' + old_lab_seg_height

        if assemble_height != old_assemble_height:
            m_value = m_value + "&" + assemble_height
            msg = msg + "&" + '装配瞳高原值=' + old_assemble_height

        if sub_mirrors_height != old_sub_mirrors_height:
            m_value = m_value + "&" + sub_mirrors_height
            msg = msg + "&" + '子镜高度原值=' + old_sub_mirrors_height

        if special_handling != old_special_handling:
            m_value = m_value + "&" + sub_mirrors_height
            msg = msg + "&" + '加工要求原值=' + old_special_handling

        lab.lab_seg_height = lab_seg_height
        lab.assemble_height = assemble_height
        lab.sub_mirrors_height = sub_mirrors_height
        lab.special_handling = special_handling
        lab.save()
        tloc = tracking_lab_order_controller()
        tloc.tracking(lab, request.user, m_value, m_key, msg)
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.EDIT_LENS_SPECMAP', login_url='/oms/forbid/')
def redirect_lens_specmap_add_edit(request):
    _form_data = {}  # 回传参数
    _items = []  # 回传列表
    rm = response_message()  # 响应
    # 获取GET参数
    page = request.GET.get('page', 1)
    name = request.GET.get('name', '')
    technology_type = request.GET.get('technology_type', 'all')
    vendor = request.GET.get('vendor', 'all')
    active = request.GET.get('active', 'all')
    # 获取搜索框参数
    try:
        _filter = {}
        TECHNOLOGY_CHOICES = LensSpecmap.TECHNOLOGY_CHOICES
        technology_type_list = []
        # 工艺列表
        for tlt in TECHNOLOGY_CHOICES:
            sc = {}
            sc['key'] = tlt[0]
            sc['value'] = tlt[1]
            technology_type_list.append(sc)

        # VD列表
        vendors_choice_list = []
        for vcl in LabOrder.VENDOR_CHOICES:
            vc = status_choice()
            vc.key = vcl[0]
            vc.value = vcl[1]
            vendors_choice_list.append(vc)

        # Active列表
        activate_choice_list = []
        for ac in LensSpecmap.ACTIVE_CHOICES:
            vc = status_choice()
            vc.key = ac[0]
            vc.value = ac[1]
            activate_choice_list.append(vc)

        if technology_type != 'all':
            _filter['technology_type'] = technology_type

        if vendor != 'all':
            _filter['vendor'] = vendor

        if active != 'all':
            _filter['active'] = active

        if name:
            lensspecmap_list = LensSpecmap.objects.filter(
                Q(inner_name__icontains=name) | Q(outer_name__icontains=name)).order_by("-id")
            lensspecmap_list = lensspecmap_list.filter(**_filter)
        else:
            lensspecmap_list = LensSpecmap.objects.filter(**_filter).order_by("-id")

        _items = lensspecmap_list

        # --页码-- 获取URL中除page外的其它参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        _form_data['total'] = _items.count()

        paginator = Paginator(_items, oms.const.PAGE_SIZE)  # Show 20 contacts per page

        try:
            _items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            _items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            _items = paginator.page(paginator.num_pages)

        return render(request, "lens_specmap_edit.html", {
            'form_data': _form_data,
            'list': _items,
            'technology_type_list': technology_type_list,
            'vendors_choices': vendors_choice_list,
            'activate_choice_list': activate_choice_list,
            'response_message': rm,
            'requestUrl': reverse('oms_lens_specmap_add_edit'),
            'query_string': query_string,
            'paginator': paginator,
            'technology_type': technology_type,
            'vendor': vendor,
            'active': active,
            'name': name
        })
    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        _form_data['exceptions'] = e
        _form_data['error_message'] = str(e)
        _form_data['request_feature'] = 'Lens Delivery'
        return render(request, "exceptions.html", {
            'form_data': _form_data,
            'requestUrl': reverse('oms_lens_specmap_add_edit'),
        })


@login_required
@permission_required('oms.EDIT_LENS_SPECMAP', login_url='/oms/forbid/')
def redirect_add_lens_specmap(request):
    try:
        technology_type = request.POST.get('technology_type', '')
        inner_code = request.POST.get('inner_code', '')
        inner_name = request.POST.get('inner_name', '')
        vendor = request.POST.get('vendor', '')
        outer_code = request.POST.get('outer_code', '')
        outer_name = request.POST.get('outer_name', '')
        active = request.POST.get('active', '')
        ls_list = LensSpecmap.objects.filter(technology_type=technology_type, inner_code=inner_code, vendor=vendor,
                                             outer_code=outer_code)
        if len(ls_list) > 0:
            return json_response(code=-1, msg='该工艺已存在该内部代码请！')
        else:
            ls = LensSpecmap()

        ls.inner_code = inner_code
        ls.inner_name = inner_name
        ls.vendor = vendor
        ls.outer_code = outer_code
        ls.outer_name = outer_name
        ls.technology_type = technology_type
        ls.active = active
        ls.user_id = request.user.id
        ls.user_name = request.user.username
        ls.save()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.EDIT_LENS_SPECMAP', login_url='/oms/forbid/')
def redirect_edit_lens_specmap(request):
    # 获取搜索框参数
    try:
        data = {}
        id = request.POST.get('id', '')
        ls = LensSpecmap.objects.get(id=id)
        data['inner_code'] = ls.inner_code
        data['inner_name'] = ls.inner_name
        data['vendor'] = ls.vendor
        data['outer_code'] = ls.outer_code
        data['outer_name'] = ls.outer_name
        data['technology_type'] = ls.technology_type
        data['active'] = ls.active
        data['id'] = ls.id
        return json_response(code=0, msg='Success', data=data)
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.EDIT_LENS_SPECMAP', login_url='/oms/forbid/')
def redirect_save_lens_specmap(request):
    # 获取搜索框参数
    try:
        id = request.POST.get('id', '')
        technology_type = request.POST.get('technology_type', '')
        inner_code = request.POST.get('inner_code', '')
        inner_name = request.POST.get('inner_name', '')
        vendor = request.POST.get('vendor', '')
        outer_code = request.POST.get('outer_code', '')
        outer_name = request.POST.get('outer_name', '')
        active = request.POST.get('active', '')
        ls = LensSpecmap.objects.get(id=id)
        ls.technology_type = technology_type
        ls.inner_code = inner_code
        ls.inner_name = inner_name
        ls.vendor = vendor
        ls.outer_code = outer_code
        ls.outer_name = outer_name
        ls.active = active
        ls.user_id = request.user.id
        ls.user_name = request.user.username
        ls.save()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.BLACK_LIST', login_url='/oms/forbid/')
def redirect_black_list_edit(request):
    _form_data = {}  # 回传参数
    _items = []  # 回传列表
    rm = response_message()  # 响应
    # 获取GET参数
    page = request.GET.get('page', 1)
    name = request.GET.get('name', '')
    # 获取搜索框参数
    try:
        _filter = {}
        if name:
            blacklists = BlackList.objects.filter(Q(customer_name__icontains=name) | Q(email__icontains=name)).order_by("-id")
        else:
            blacklists = BlackList.objects.filter().order_by("-id")

        _items = blacklists

        # --页码-- 获取URL中除page外的其它参数
        query_string = re.sub(r'page=[0-9]+[&]?', '', request.META.get('QUERY_STRING', None))
        if query_string:
            query_string = '&' + query_string

        _form_data['total'] = _items.count()

        paginator = Paginator(_items, oms.const.PAGE_SIZE)  # Show 20 contacts per page

        try:
            _items = paginator.page(page)
        except PageNotAnInteger:
            # If page is not an integer, deliver first page.
            _items = paginator.page(1)
        except EmptyPage:
            # If page is out of range (e.g. 9999), deliver last page of results.
            _items = paginator.page(paginator.num_pages)

        return render(request, "black_list_edit.html", {
            'form_data': _form_data,
            'list': _items,
            'response_message': rm,
            'requestUrl': reverse('oms_lens_specmap_add_edit'),
            'query_string': query_string,
            'paginator': paginator,
            'name': name
        })
    except Exception as e:
        logging.debug('Exception: %s' % str(e))
        _form_data['exceptions'] = e
        _form_data['error_message'] = str(e)
        _form_data['request_feature'] = 'Lens Delivery'
        return render(request, "exceptions.html", {
            'form_data': _form_data,
            'requestUrl': reverse('oms_lens_specmap_add_edit'),
        })


@login_required
@permission_required('oms.BLACK_LIST', login_url='/oms/forbid/')
def redirect_black_list_modify(request):
    # 获取搜索框参数
    try:
        data = {}
        id = request.POST.get('id', '')
        flag = request.POST.get('flag', '')
        blacklist = BlackList.objects.get(id=id)
        if flag == 'enable':
            blacklist.is_enabled = True
        else:
            blacklist.is_enabled = False
        blacklist.save()
        #清空set()
        poc = PgOrderController()
        poc.black_set.clear()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.BLACK_LIST', login_url='/oms/forbid/')
def redirect_black_list_add(request):
    try:
        name = request.POST.get('name', '')
        first_name = request.POST.get('first_name', '')
        last_name = request.POST.get('last_name', '')
        phone = request.POST.get('phone', '')
        email = request.POST.get('email', '')
        blacklists = BlackList.objects.filter(email=email)
        if len(blacklists)>0:
            return json_response(code=-1, msg=e)

        blacklist = BlackList()
        blacklist.customer_name = name
        blacklist.first_name = first_name
        blacklist.last_name = last_name
        blacklist.phone = phone
        blacklist.email = email
        blacklist.save()
        #清空set()
        poc = PgOrderController()
        poc.black_set.clear()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
@permission_required('oms.ACT_SHIPMENT', login_url='/oms/forbid/')
def redirect_delivered_time_add(request):
    try:
        delivered_time = request.POST.get('delivered_time', '')
        lab_number = request.POST.get('lab_number', '')
        lab = LabOrder.objects.get(lab_number=lab_number)
        if lab.status == 'DELIVERED':
            return json_response(code=-1, msg="该单已是妥投状态")


        lab.status = 'DELIVERED'
        lab.delivered_at = delivered_time
        lab.save()
        tloc = tracking_lab_order_controller()
        tloc.tracking(lab, request.user, 'DELIVERED', '妥投', '')

        order_number = lab.order_number
        poc = pg_order_controller()
        s2delivered = poc.set_order2delivered(request, order_number)

        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
def redirect_pause_duration_add(request):
    try:
        lab_number = request.POST.get('lab_number', '')
        exclude_days = request.POST.get('exclude_days', '0')
        lab = LabOrder.objects.get(lab_number=lab_number)
        new_exclude_days = float(lab.exclude_days) + float(exclude_days)
        lab.exclude_days = new_exclude_days
        lab.save()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


@login_required
def redirect_edit_pgorderitem_dia(request):
    try:
        order_number = request.POST.get('order_number', '')
        item_id = request.POST.get('item_id', '')
        product_index = request.POST.get('product_index', '')
        pgis = PgOrderItem.objects.filter(item_id=item_id, order_number=order_number, product_index=product_index)
        if len(pgis) < 0:
            return json_response(code=-1, msg='未查找到PgOrderItem')

        pgorderitem = pgis[0]
        pre = pgorderitem.get_prescritpion
        pre_dict = utilities.convert_to_dict(pre)
        body = {
            "product_sku": pgorderitem.frame,
            "prescription": pre_dict
        }
        body = json.dumps(body, cls=DateEncoder)
        logging.debug("body==>%s" % body)

        req = urllib2.Request(url=CALCULATE_DIAMETER_URL_V2, data=body, headers=token_header)
        res = urllib2.urlopen(req)
        resp = res.read()

        resp = json.loads(resp)
        logging.debug('respones: %s' % resp)

        pgorderitem.dia_1 = resp['dia_1']
        pgorderitem.dia_2 = resp['dia_2']
        pgorderitem.save()
        return json_response(code=0, msg='Success')
    except Exception as e:
        return json_response(code=-1, msg=e)


# 伟星下单(库存片)
@login_required
@permission_required('oms.WX_ORDER', login_url='/oms/forbid/')
def wx_meta_purchasing_order(request):
    rm = res_msg.response_dict()
    stat_dict = {}
    line_id = request.POST.get('sub_list', '')
    try:
        # 获取采购订单line对象
        purchase_order_line_entitys = laborder_purchase_order_line.objects.filter(id=line_id)

        # 检查
        if purchase_order_line_entitys.count() < 1:  # 采购订单不存在
            stat_dict[line_id] = {'Success': False, 'Message': '采购订单不存在'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })
        # 取出对象
        purchase_order_line_entity = purchase_order_line_entitys[0]
        if purchase_order_line_entity.is_purchased:  # 已经下单
            stat_dict[purchase_order_line_entity.lab_number] = {'Success': False, 'Message': '该订单已下单'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })
        if purchase_order_line_entity.laborder_entity.status not in (
                'PRINT_DATE', 'LENS_REGISTRATION', 'LENS_RECEIVE', 'ASSEMBLING', 'FRAME_OUTBOUND', 'REQUEST_NOTES'):
            stat_dict[purchase_order_line_entity.lab_number] = {'Success': False, 'Message': '该状态不能下单'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })

        # wx_meta_purchase_controller类
        wx_pur_ctrl = wx_meta_purchase_controller()

        lab = LabOrder.objects.get(lab_number=purchase_order_line_entity.lab_number)

        data_dict = wx_pur_ctrl.pack_request_value(lab)
        
        if len(data_dict) == 0:
            stat_dict[purchase_order_line_entity.lab_number] = {'Success': False, 'Message': '封装数据错误'}
            return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
            })

        res = wx_pur_ctrl.add_meta_order(data_dict)
        stat_dict = wx_pur_ctrl.analysis_result(request, lab, purchase_order_line_entity, res)
        return render(request, 'purchase_res_template.html', {
                'stat_list': stat_dict
        })

    except Exception as e:
        rm['code'] = -1
        rm['message'] = 'Error Message: %s' % str(e)

    return JsonResponse(rm)

# add gyf 2020.8.31
@login_required
@permission_required('wms.RETIRED', login_url='/oms/forbid/')
def oms_file_download(request):
    """ 下载QA文件 """
    qa_url = request.GET.get('qa_url', '')
    qa_url = qa_url.replace("%20"," ")
    sql = "SELECT * from qc_laborder_accessories WHERE object_url ='{}';"
    nowTime = "_"+ str(datetime.datetime.now())[-5:-1]
    try:
        with connections['pg_oms_query'].cursor() as cursor:
            sql = sql.format(qa_url)

            cursor.execute(sql)

            results = namedtuplefetchall(cursor)
            if (results.__len__()!=0):

                file_path = str(results[0].object_url)
                name = results[0].lab_number
                type = results[0].accessories_type
                res = requests.get(file_path)

                response = HttpResponse(res, content_type="application/octet-stream")
                from django.utils.encoding import escape_uri_path
                if file_path.endswith('.jpg'):
                    response['Content-Disposition'] = "attachment; filename={};".format(escape_uri_path(name + nowTime + '.jpg'))
                elif file_path.endswith('.pdf'):
                    response['Content-Disposition'] = "attachment; filename={};".format(escape_uri_path(name + nowTime + '.pdf'))
                else:
                    response['Content-Disposition'] = "attachment; filename={};".format(escape_uri_path(name + nowTime + '.mp4'))
                return response
            else:
                return json_response(code=-1, msg='文件不存在')


    except Exception as e:
        return json_response(code=-1,msg='文件不存在')


@login_required
def redirect_laborder_list_v2_qa(request):
    # QA 列表

    qa_order_number = request.POST.get('qa_order_number', '')
    print(qa_order_number)

    sql = "SELECT accessories_type,object_url FROM qc_laborder_accessories  WHERE lab_number = '{0}'"

    qa_video_list = []
    qa_img_list = []
    qa_pdf_list = []

    with connections['pg_oms_query'].cursor() as cursor:
        sql = sql.format(qa_order_number)
        cursor.execute(sql)
        results = dictfetchall(cursor)
        print(results)

        if (results.__len__() != 0):

            for i in range(len(results)):
                type = str(results[i]["accessories_type"])

                url = str(results[i]["object_url"]).replace(" ","%20")

                if url.endswith('.jpg'):
                    qa_img_list.append(url)
                elif url.endswith('.pdf'):
                    qa_pdf_list.append(url)
                else:
                    qa_video_list.append(url)
        else:
            pass

        return json_response(code=1,data={'qa_img_list':qa_img_list,
                                          'qa_video_list':qa_video_list,
                                          'qa_pdf_list':qa_pdf_list})

