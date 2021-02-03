# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import logging

from django.core.management.base import BaseCommand
from django.utils.http import urlquote
from django.http import HttpResponse
from django.utils import timezone
from django.db import connections

from util.db_helper import *
from io import BytesIO
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill
from oms.views import add_tracklog, add_order_tracking_report, addOrderTrackingReportCs
from oms.models.order_models import LabOrder, PgOrderItem, PgOrder
from api.models import DingdingChat
from pg_oms import settings


class Command(BaseCommand):
    def handle(self, *args, **options):
        logging.critical('start ....')
        # chat_id="chat1fa6ffa170a762c040b5f0cc042ade7c" #keep touch chat
        chat_id="chat63968dc1c4da30c32130a0542470744f" #生产chat
        try:
            ddc = DingdingChat()
            file_name = settings.UPLOAD_BASE + settings.UPLOAD_CONFIG['UPLOAD_EXCEL_URL'] +'ups_number.xlsx'
            logging.critical(file_name)
            workbook1 = load_workbook(file_name)
            sheets = workbook1.get_sheet_names()
        except Exception as e:
            msg = "未找到ups_number.xlsx"
            ddc.send_text_to_chat(chat_id, "{0}".format(msg))
            return

        try:
            for sheet in sheets:
                live_sheet = workbook1[sheet]
                activate = live_sheet['A1']
                if not activate.value:
                    max_row = live_sheet.max_row
                    max_column = live_sheet.max_column
                    data_list = []
                    for row_num in range(2, max_row+1):
                        if live_sheet.cell(row_num, 1).value:
                            d={
                                "box_id": str(live_sheet.cell(row_num, 1).value).replace("L", ""),
                                "order_number": str(live_sheet.cell(row_num, 2).value).replace("L", ""),
                                "carrier": str(live_sheet.cell(row_num, 3).value),
                                "tracking_number": live_sheet.cell(row_num, 4).value,
                                "qty": str(live_sheet.cell(row_num, 5).value).replace("L", ""),
                                "status": "SHIPPING"
                            }
                            data_list.append(d)

                    message = ''
                    for item in data_list:
                        try:
                            lbos = LabOrder.objects.filter(order_number__contains=item['order_number']).exclude(status='CANCELLED')
                            pg_order_number = lbos[0].order_number
                            nowtime = timezone.now().strftime('%Y-%m-%d %H:%M:%S')

                            with connections['pg_oms_query'].cursor() as cursor:
                                pois = PgOrderItem.objects.filter(order_number=pg_order_number)
                                for laborder in lbos:
                                    sql = """SELECT id FROM shipment_pre_delivery_line WHERE lab_order_entity_id='%s' AND pre_delivery_entity_id='%s'""" % (laborder.id, item['box_id'])
                                    cursor.execute(sql)
                                    pdl = namedtuplefetchall(cursor)
                                    if len(pdl) > 0:
                                        if laborder.shipping_number != '' and laborder.status == 'SHIPPING':
                                            message = message + '-' + str(laborder.lab_number) + '该订单已处于已发货状态,请人工确认' + '\t'
                                            logging.critical(str(laborder.id)+","+laborder.shipping_number+","+laborder.status)
                                            continue
                                        laborder.final_time = nowtime
                                        laborder.carriers = item['carrier']
                                        laborder.shipping_number = item['tracking_number']
                                        laborder.status = item['status']
                                        laborder.save()

                                        # 写入shipment_pre_delivery_line表的快递公司
                                        update_sql="update shipment_pre_delivery_line set ship_carrier='{0}',tracking_number='{1}' where id={2}".format(item["carrier"],item['tracking_number'],pdl[0].id)
                                        connections["default"].cursor().execute(update_sql)

                                        labnumber = laborder.lab_number
                                        lablist = labnumber.split("-")[:3]
                                        order_number = "-".join(lablist)
                                        for poi in pois:
                                            if poi.lab_order_number == order_number:
                                                poi.final_date = nowtime
                                                poi.status = "shipped"
                                                poi.save()
                                                # add order_tracking log
                                                add_tracklog(laborder, laborder.lab_number, laborder.frame,
                                                             laborder.order_date,
                                                             None, '发货',
                                                             laborder.status)
                                                # add order_tracking_report
                                                add_order_tracking_report(laborder.lab_number,
                                                                          laborder,
                                                                          laborder.status, item['carrier'],
                                                                          item['tracking_number'],
                                                                          "null")
                                                # add order_tracking_report_cs
                                                addOrderTrackingReportCs(poi.order_number,
                                                                         poi.shipping_method,
                                                                         laborder.lab_number,
                                                                         laborder.frame,
                                                                         laborder.order_date, laborder.status,
                                                                         item['carrier'], item['tracking_number'],
                                                                         'null')
                                    else:
                                        message = message + '-' + '该box中不存在该订单-' + str(laborder.lab_number) + ',请人工确认' + '\t'
                                        continue

                                pg_order = PgOrder.objects.get(order_number=pg_order_number)
                                pg_order.final_date = nowtime
                                pg_order.status = 'shipped'
                                pg_order.save()
                                message = message + "OrderNumber:" + item['order_number']+",laborder_id="+ str(laborder.id)+ "发货成功" + '\t'
                                logging.critical(message)
                        except Exception as e:
                            logging.critical(e)
                            message = message + "OrderNumber:" + item['order_number']+",laborder_id="+ str(laborder.id)+ " 发货异常请人工确认" + '\t'
                            logging.critical(message)
                            continue
                    live_sheet['A1'].value = 'TRUE'
                    workbook1.save(file_name)
                    ddc.send_text_to_chat(chat_id,  "BOX:{0}: {1}".format(item['box_id'], message))
        except Exception as e:
            logging.critical(e)
            for sheet in sheets:
                live_sheet = workbook1[sheet]
                activate = live_sheet['A1']
                box_id = live_sheet['A2']
                if not activate.value:
                    live_sheet['A1'].value = 'TRUE'
                    workbook1.save(file_name)
                msg = "发货失败人工处理"
                ddc.send_text_to_chat(chat_id, "BOX:{0}: {1}".format(box_id.value, msg))


