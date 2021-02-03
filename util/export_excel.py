# -*-coding:utf-8 -*-
#
# Created on 2016-03-15, by wangj
#

__author__ = 'wangj'

import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill
from io import BytesIO
from django.utils.http import urlquote
from django.http import HttpResponse


class ExcelResponse(HttpResponse):
    """
    excel文件导出
    支持xls和csv格式文件
    支持多sheet页导出
    """
    def export_excel(self, file_name, head_data=[], content_data=[]):
        wb = Workbook()		# 生成一个工作簿（即一个Excel文件）
        wb.encoding = 'utf-8'
        font = Font(u'DengXian', size=14, bold=True, color='000000')
        body_font = Font(u'DengXian', size=14, bold=False, color='000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill("solid", fgColor="d1cbcb")
        sheet1 = wb.active # 获取第一个工作表（sheet1）
        for i in range(1, len(head_data)+1): # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
            # 从row=1，column=1开始写，即将head_data的数据依次写入第一行
            sheet1.cell(row=1, column=i).value = head_data[i - 1]
            sheet1.cell(row=1, column=i).font = font
            sheet1.cell(row=1, column=i).alignment = alignment

        sheet1.row_dimensions[1].height = 30

        for row in sheet1.rows:
            for cell in row:
                sheet1[cell.coordinate].fill = fill
                colum_name = cell.coordinate[:-1]
                sheet1.column_dimensions[colum_name].width = 20

        for obj in content_data:
            max_row = sheet1.max_row + 1 # 获取到工作表的最大行数并加1
            for x in range(1, len(obj)+1):		# 将每一个对象的所有字段的信息写入一行内
                sheet1.cell(row=max_row, column=x).value = obj[x-1]
                sheet1.cell(row=max_row, column=x).font = body_font
                sheet1.cell(row=max_row, column=x).alignment = alignment
        # 准备写入到IO中
        output = BytesIO()
        wb.save(output)	 # 将Excel文件内容保存到IO中
        output.seek(0)	 # 重新定位到开始
        # 设置HttpResponse的类型
        response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
        file_name = '%s.xls' % file_name # 给文件名中添加日期时间
        file_name = urlquote(file_name)	 # 使用urlquote()方法解决中文无法使用的问题
        response['Content-Disposition'] = 'attachment; filename=%s' % file_name
        return response


